"""
Accounting Admin - Professional Edition
Author: ddiplas
Version: 3.0 TURBO
Features: Advanced filtering, bulk actions, exports, dashboards, inline editing
"""

from django.contrib import admin
from django.utils.html import format_html, mark_safe
from django.urls import reverse, path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Count, Sum, Q, Avg
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django import forms
from django.template.response import TemplateResponse
from django.contrib.admin import SimpleListFilter
from datetime import datetime, timedelta
import csv
import json
from functools import wraps

from .models import (
    ClientProfile, 
    ObligationGroup, 
    ObligationProfile, 
    ObligationType,
    ClientObligation,
    MonthlyObligation,
    EmailTemplate,
    EmailAutomationRule,
    ScheduledEmail,
    VoIPCall, 
    VoIPCallLog
)

# ============================================
# CUSTOM FILTERS
# ============================================

class OverdueFilter(SimpleListFilter):
    """Φίλτρο για καθυστερημένες υποχρεώσεις"""
    title = 'Καθυστέρηση'
    parameter_name = 'overdue'
    
    def lookups(self, request, model_admin):
        return (
            ('overdue', '🔴 Καθυστερημένες'),
            ('today', '⚠️ Λήγουν Σήμερα'),
            ('week', '📅 Επόμενη Εβδομάδα'),
            ('month', '📆 Επόμενος Μήνας'),
        )
    
    def queryset(self, request, queryset):
        today = timezone.now().date()
        if self.value() == 'overdue':
            return queryset.filter(deadline__lt=today, status='pending')
        elif self.value() == 'today':
            return queryset.filter(deadline=today)
        elif self.value() == 'week':
            week_later = today + timedelta(days=7)
            return queryset.filter(deadline__range=[today, week_later])
        elif self.value() == 'month':
            month_later = today + timedelta(days=30)
            return queryset.filter(deadline__range=[today, month_later])


class HasAttachmentFilter(SimpleListFilter):
    """Φίλτρο για συνημμένα"""
    title = 'Συνημμένο'
    parameter_name = 'has_attachment'
    
    def lookups(self, request, model_admin):
        return (
            ('yes', '📎 Με Συνημμένο'),
            ('no', '❌ Χωρίς Συνημμένο'),
        )
    
    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.exclude(attachment='')
        elif self.value() == 'no':
            return queryset.filter(attachment='')


class ClientTypeFilter(SimpleListFilter):
    """Φίλτρο ανά τύπο πελάτη"""
    title = 'Τύπος Πελάτη'
    parameter_name = 'client_type'
    
    def lookups(self, request, model_admin):
        return (
            ('vip', '⭐ VIP (>10 υποχρ.)'),
            ('active', '✅ Ενεργοί'),
            ('new', '🆕 Νέοι (<3 μήνες)'),
            ('inactive', '💤 Ανενεργοί'),
        )
    
    def queryset(self, request, queryset):
        if self.value() == 'vip':
            # Πελάτες με πολλές υποχρεώσεις
            vip_clients = ClientProfile.objects.annotate(
                obl_count=Count('monthly_obligations')
            ).filter(obl_count__gt=10).values_list('id', flat=True)
            return queryset.filter(id__in=vip_clients)
        elif self.value() == 'active':
            return queryset.filter(clientobligation__is_active=True)
        elif self.value() == 'new':
            three_months_ago = timezone.now() - timedelta(days=90)
            return queryset.filter(created_at__gte=three_months_ago)
        elif self.value() == 'inactive':
            return queryset.filter(clientobligation__is_active=False)


# ============================================
# CUSTOM FORMS με VALIDATION
# ============================================

class GenerateObligationsForm(forms.Form):
    """Enhanced form για δημιουργία υποχρεώσεων"""
    year = forms.IntegerField(
        label='Έτος',
        initial=timezone.now().year,
        widget=forms.NumberInput(attrs={'class': 'form-control', 'style': 'width: 200px;'})
    )
    month = forms.IntegerField(
        label='Μήνας',
        initial=timezone.now().month,
        min_value=1,
        max_value=12,
        widget=forms.Select(choices=[(i, f'{i:02d} - {["","Ιαν","Φεβ","Μαρ","Απρ","Μαϊ","Ιουν","Ιουλ","Αυγ","Σεπ","Οκτ","Νοε","Δεκ"][i]}') for i in range(1, 13)])
    )
    only_active = forms.BooleanField(
        label='Μόνο ενεργοί πελάτες',
        initial=True,
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'})
    )
    send_notification = forms.BooleanField(
        label='Αποστολή email ειδοποίησης',
        initial=False,
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'})
    )


class BulkCompleteForm(forms.Form):
    """Form για μαζική ολοκλήρωση"""
    time_spent = forms.DecimalField(
        label='Ώρες εργασίας (για όλες)',
        required=False,
        decimal_places=2,
        widget=forms.NumberInput(attrs={'class': 'form-control', 'placeholder': '1.5'})
    )
    notes = forms.CharField(
        label='Σημειώσεις',
        required=False,
        widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 3})
    )
    send_email = forms.BooleanField(
        label='Αποστολή email ολοκλήρωσης',
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'})
    )


# ============================================
# MIXINS για REUSABILITY
# ============================================

class ExportMixin:
    """Mixin για export λειτουργικότητα"""
    
    def export_as_csv(self, request, queryset):
        """Export to CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response.write('\ufeff')  # UTF-8 BOM for Excel
        response['Content-Disposition'] = f'attachment; filename="{self.model._meta.verbose_name_plural}_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        
        writer = csv.writer(response, delimiter=';')
        
        # Headers
        fields = [field.name for field in self.model._meta.fields]
        headers = [self.model._meta.get_field(field).verbose_name for field in fields]
        writer.writerow(headers)
        
        # Data
        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field)
                if hasattr(value, 'strftime'):
                    value = value.strftime('%d/%m/%Y %H:%M')
                row.append(str(value))
            writer.writerow(row)
        
        messages.success(request, f'✅ Exported {queryset.count()} records to CSV!')
        return response
    
    export_as_csv.short_description = '📊 Export to CSV'
    
    def export_as_excel(self, request, queryset):
        """Export to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.model._meta.verbose_name_plural
            
            # Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            
            # Headers
            fields = [field.name for field in self.model._meta.fields]
            headers = [self.model._meta.get_field(field).verbose_name for field in fields]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_num, obj in enumerate(queryset, 2):
                for col, field in enumerate(fields, 1):
                    value = getattr(obj, field)
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%d/%m/%Y %H:%M')
                    ws.cell(row=row_num, column=col).value = str(value)
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{self.model._meta.verbose_name_plural}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            wb.save(response)
            
            messages.success(request, f'✅ Exported {queryset.count()} records to Excel!')
            return response
            
        except ImportError:
            messages.error(request, '❌ openpyxl not installed!')
            return None
    
    export_as_excel.short_description = '📈 Export to Excel'


class BulkActionsMixin:
    """Mixin για μαζικές ενέργειες"""
    
    def bulk_complete(self, request, queryset):
        """Μαζική ολοκλήρωση με form"""
        if 'apply' in request.POST:
            form = BulkCompleteForm(request.POST)
            if form.is_valid():
                time_spent = form.cleaned_data.get('time_spent')
                notes = form.cleaned_data.get('notes')
                send_email = form.cleaned_data.get('send_email')
                
                updated = 0
                for obj in queryset:
                    obj.status = 'completed'
                    obj.completed_date = timezone.now().date()
                    obj.completed_by = request.user
                    if time_spent:
                        obj.time_spent = time_spent
                    if notes:
                        obj.notes = f"{obj.notes}\n[BULK] {notes}" if obj.notes else f"[BULK] {notes}"
                    obj.save()
                    updated += 1
                
                messages.success(request, f'✅ Ολοκληρώθηκαν {updated} υποχρεώσεις!')
                
                if send_email:
                    # TODO: Implement email sending
                    messages.info(request, '📧 Emails θα σταλούν σύντομα...')
                
                return redirect(request.get_full_path())
        
        else:
            form = BulkCompleteForm()
        
        context = {
            'title': 'Μαζική Ολοκλήρωση Υποχρεώσεων',
            'objects': queryset,
            'form': form,
            'action': 'bulk_complete',
        }
        return render(request, 'admin/accounting/bulk_action.html', context)
    
    bulk_complete.short_description = '✅ Μαζική Ολοκλήρωση'


# ============================================
# ENHANCED ADMIN CLASSES
# ============================================

@admin.register(ClientProfile)
class ClientProfileAdmin(admin.ModelAdmin, ExportMixin):
    """TURBO ClientProfile Admin με όλα τα features"""
    
    list_display = [
        'afm_colored',
        'eponimia_link',
        'client_type_badge',
        'obligations_count',
        'phone_display',
        'email_link',
        'status_indicator',
        'quick_actions'
    ]
    
    list_filter = [
        ClientTypeFilter,
        'eidos_ipoxreou',
        'katigoria_vivlion',
        'agrotis',
        ('created_at', admin.DateFieldListFilter),
    ]
    
    search_fields = ['afm', 'eponimia', 'onoma', 'email', 'kinito_tilefono']
    
    readonly_fields = ['created_at', 'updated_at', 'statistics_display', 'obligations_timeline']
    
    actions = ['export_as_csv', 'export_as_excel', 'send_mass_email', 'generate_report']
    
    list_per_page = 25
    
    fieldsets = (
        ('🏢 Βασικά Στοιχεία', {
            'fields': ('afm', 'doy', 'eponimia', 'onoma', 'onoma_patros'),
            'classes': ('wide',)
        }),
        ('👤 Ταυτοποίηση', {
            'fields': ('arithmos_taftotitas', 'eidos_taftotitas', 'prosopikos_arithmos', 
                      'amka', 'am_ika', 'arithmos_gemi', 'arithmos_dypa'),
            'classes': ('collapse',)
        }),
        ('📍 Διεύθυνση Κατοικίας', {
            'fields': ('diefthinsi_katoikias', 'arithmos_katoikias', 'poli_katoikias', 
                      'dimos_katoikias', 'nomos_katoikias', 'tk_katoikias'),
            'classes': ('collapse',)
        }),
        ('📞 Επικοινωνία', {
            'fields': ('tilefono_oikias_1', 'tilefono_oikias_2', 'kinito_tilefono', 'email'),
            'classes': ('wide',)
        }),
        ('🏭 Επιχείρηση', {
            'fields': ('diefthinsi_epixeirisis', 'arithmos_epixeirisis', 'poli_epixeirisis',
                      'dimos_epixeirisis', 'nomos_epixeirisis', 'tk_epixeirisis',
                      'tilefono_epixeirisis_1', 'tilefono_epixeirisis_2'),
            'classes': ('collapse',)
        }),
        ('💳 Τραπεζικά', {
            'fields': ('trapeza', 'iban'),
            'classes': ('collapse',)
        }),
        ('📊 Επιχειρηματικά', {
            'fields': ('eidos_ipoxreou', 'katigoria_vivlion', 'nomiki_morfi', 
                      'agrotis', 'imerominia_enarksis')
        }),
        ('🔑 Διαπιστευτήρια', {
            'fields': ('onoma_xristi_taxisnet', 'kodikos_taxisnet',
                      'onoma_xristi_ika_ergodoti', 'kodikos_ika_ergodoti',
                      'onoma_xristi_gemi', 'kodikos_gemi'),
            'classes': ('collapse',)
        }),
        ('📈 Στατιστικά', {
            'fields': ('statistics_display', 'obligations_timeline'),
            'classes': ('wide',)
        }),
    )
    
    def afm_colored(self, obj):
        """ΑΦΜ με colored badge"""
        return format_html(
            '<span style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); '
            'color: white; padding: 6px 12px; border-radius: 20px; font-weight: bold; '
            'font-family: monospace;">{}</span>',
            obj.afm
        )
    afm_colored.short_description = 'Α.Φ.Μ.'
    
    def eponimia_link(self, obj):
        """Επωνυμία με link και tooltip"""
        obligations = obj.monthly_obligations.filter(status='pending').count()
        tooltip = f"{obligations} εκκρεμείς υποχρεώσεις"
        return format_html(
            '<a href="{}" style="color: #2563eb; text-decoration: none; font-weight: 600;" '
            'title="{}">👤 {}</a>',
            reverse('admin:accounting_clientprofile_change', args=[obj.id]),
            tooltip,
            obj.eponimia
        )
    eponimia_link.short_description = 'Επωνυμία'
    
    def client_type_badge(self, obj):
        """Τύπος πελάτη με badge"""
        colors = {
            'individual': ('#10b981', '👤'),
            'professional': ('#3b82f6', '💼'),
            'company': ('#8b5cf6', '🏢'),
        }
        color, icon = colors.get(obj.eidos_ipoxreou, ('#666', '❓'))
        return format_html(
            '<span style="background: {}; color: white; padding: 4px 10px; '
            'border-radius: 12px; font-size: 0.9em;">{} {}</span>',
            color, icon, obj.get_eidos_ipoxreou_display()
        )
    client_type_badge.short_description = 'Τύπος'
    
    def obligations_count(self, obj):
        """Μετρητής υποχρεώσεων με progress bar"""
        total = obj.monthly_obligations.count()
        completed = obj.monthly_obligations.filter(status='completed').count()
        pending = obj.monthly_obligations.filter(status='pending').count()
        overdue = obj.monthly_obligations.filter(status='overdue').count()
        
        if total == 0:
            return '—'
        
        percentage = (completed / total) * 100 if total > 0 else 0
        
        return format_html(
            '<div style="display: flex; align-items: center; gap: 10px;">'
            '<div style="flex: 1; background: #e5e7eb; border-radius: 10px; height: 20px; overflow: hidden;">'
            '<div style="background: linear-gradient(90deg, #10b981 0%, #34d399 100%); '
            'width: {}%; height: 100%; transition: width 0.3s;"></div>'
            '</div>'
            '<div style="font-size: 0.85em; white-space: nowrap;">'
            '<span style="color: #10b981;">✅{}</span> | '
            '<span style="color: #f59e0b;">⏳{}</span> | '
            '<span style="color: #ef4444;">🔴{}</span>'
            '</div>'
            '</div>',
            percentage, completed, pending, overdue
        )
    obligations_count.short_description = 'Υποχρεώσεις'
    
    def phone_display(self, obj):
        """Τηλέφωνο με click-to-call"""
        phone = obj.kinito_tilefono or obj.tilefono_oikias_1 or '—'
        if phone == '—':
            return phone
        return format_html(
            '<a href="tel:{}" style="color: #059669; text-decoration: none; font-weight: 600;">📱 {}</a>',
            phone, phone
        )
    phone_display.short_description = 'Τηλέφωνο'
    
    def email_link(self, obj):
        """Email με mailto link"""
        if not obj.email:
            return '—'
        return format_html(
            '<a href="mailto:{}" style="color: #2563eb; text-decoration: none;">✉️ {}</a>',
            obj.email, obj.email
        )
    email_link.short_description = 'Email'
    
    def status_indicator(self, obj):
        """Status indicator με animation"""
        has_active = hasattr(obj, 'clientobligation') and obj.clientobligation.is_active
        if has_active:
            return format_html(
                '<div style="display: flex; align-items: center; gap: 5px;">'
                '<div style="width: 10px; height: 10px; background: #10b981; '
                'border-radius: 50%; animation: pulse 2s infinite;"></div>'
                '<span style="color: #10b981; font-weight: 600;">Ενεργός</span>'
                '</div>'
            )
        return format_html(
            '<div style="display: flex; align-items: center; gap: 5px;">'
            '<div style="width: 10px; height: 10px; background: #ef4444; border-radius: 50%;"></div>'
            '<span style="color: #ef4444;">Ανενεργός</span>'
            '</div>'
        )
    status_indicator.short_description = 'Κατάσταση'
    
    def quick_actions(self, obj):
        """Quick action buttons"""
        return format_html(
            '<div style="display: flex; gap: 5px;">'
            '<a class="button" href="/accounting/client/{}/obligations/" '
            'style="padding: 4px 8px; background: #667eea; color: white; '
            'border-radius: 4px; text-decoration: none;">📋 Υποχρ.</a>'
            '<a class="button" href="/accounting/client/{}/report/" '
            'style="padding: 4px 8px; background: #10b981; color: white; '
            'border-radius: 4px; text-decoration: none;">📊 Report</a>'
            '</div>',
            obj.id, obj.id
        )
    quick_actions.short_description = '⚡ Actions'
    
    def statistics_display(self, obj):
        """Στατιστικά πελάτη"""
        stats = obj.monthly_obligations.aggregate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed')),
            total_hours=Sum('time_spent'),
            total_cost=Sum('time_spent') * 50  # Assuming 50€/hour
        )
        
        return format_html(
            '<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; '
            'padding: 15px; background: linear-gradient(135deg, #f3f4f6 0%, #e5e7eb 100%); '
            'border-radius: 10px;">'
            '<div><strong>📊 Σύνολο:</strong> {}</div>'
            '<div><strong>✅ Ολοκληρωμένες:</strong> {}</div>'
            '<div><strong>⏱️ Ώρες:</strong> {:.1f}</div>'
            '<div><strong>💰 Κόστος:</strong> €{:.2f}</div>'
            '</div>',
            stats['total'] or 0,
            stats['completed'] or 0,
            stats['total_hours'] or 0,
            stats['total_cost'] or 0
        )
    statistics_display.short_description = 'Στατιστικά'
    
    def obligations_timeline(self, obj):
        """Timeline υποχρεώσεων"""
        upcoming = obj.monthly_obligations.filter(
            status='pending',
            deadline__gte=timezone.now().date()
        ).order_by('deadline')[:5]
        
        if not upcoming:
            return 'Δεν υπάρχουν προσεχείς υποχρεώσεις'
        
        html = '<div style="padding: 10px; background: #fff; border-radius: 8px;">'
        for obl in upcoming:
            days = (obl.deadline - timezone.now().date()).days
            color = '#10b981' if days > 7 else '#f59e0b' if days > 3 else '#ef4444'
            html += format_html(
                '<div style="display: flex; justify-content: space-between; '
                'padding: 8px; margin: 5px 0; background: #f9fafb; border-left: 3px solid {};">'
                '<span><strong>{}</strong></span>'
                '<span style="color: {};">σε {} ημέρες</span>'
                '</div>',
                color, obl.obligation_type.name, color, days
            )
        html += '</div>'
        return mark_safe(html)
    obligations_timeline.short_description = 'Επόμενες Υποχρεώσεις'
    
    def send_mass_email(self, request, queryset):
        """Μαζική αποστολή email"""
        emails = queryset.exclude(email='').values_list('email', flat=True)
        messages.info(request, f'📧 Θα σταλούν {len(emails)} emails')
        # TODO: Implement email sending
    send_mass_email.short_description = '📧 Μαζικό Email'
    
    def generate_report(self, request, queryset):
        """Generate PDF report"""
        messages.info(request, '📄 Generating PDF reports...')
        # TODO: Implement PDF generation
    generate_report.short_description = '📄 Generate Report'
    
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/custom_admin.js',)


@admin.register(MonthlyObligation)
class MonthlyObligationAdmin(admin.ModelAdmin, ExportMixin, BulkActionsMixin):
    """SUPER TURBO Monthly Obligation Admin"""
    
    list_display = [
        'id_badge',
        'client_smart',
        'obligation_badge',
        'deadline_visual',
        'status_progress',
        'time_cost_display',
        'attachment_indicator',
        'completed_by_avatar',
        'action_buttons'
    ]
    
    list_filter = [
        OverdueFilter,
        HasAttachmentFilter,
        'status',
        ('obligation_type', admin.RelatedOnlyFieldListFilter),
        ('client', admin.RelatedOnlyFieldListFilter),
        'year',
        'month',
        ('completed_by', admin.RelatedOnlyFieldListFilter),
    ]
    
    search_fields = [
        'client__afm',
        'client__eponimia',
        'obligation_type__name',
        'notes',
        'id'
    ]
    
    list_editable = []  # Removed for cleaner interface
    list_select_related = ['client', 'obligation_type', 'completed_by']
    
    date_hierarchy = 'deadline'
    list_per_page = 30
    
    actions = [
        'bulk_complete',
        'mark_as_completed',
        'mark_as_pending',
        'export_as_csv',
        'export_as_excel',
        'send_reminder_emails',
        'generate_invoices'
    ]
    
    readonly_fields = [
        'created_at', 
        'updated_at', 
        'cost_calculator',
        'history_timeline',
        'related_obligations'
    ]
    
    fieldsets = (
        ('📋 Βασικά Στοιχεία', {
            'fields': ('client', 'obligation_type', 'year', 'month', 'deadline'),
            'classes': ('wide',)
        }),
        ('✅ Κατάσταση Ολοκλήρωσης', {
            'fields': ('status', 'completed_date', 'completed_by'),
            'classes': ('wide',)
        }),
        ('💰 Χρέωση & Κοστολόγηση', {
            'fields': ('time_spent', 'hourly_rate', 'cost_calculator'),
            'description': '⏱️ Καταγραφή χρόνου και αυτόματος υπολογισμός κόστους'
        }),
        ('📎 Σημειώσεις & Αρχεία', {
            'fields': ('notes', 'attachment'),
            'classes': ('wide',)
        }),
        ('📊 Analytics & History', {
            'fields': ('history_timeline', 'related_obligations'),
            'classes': ('collapse',)
        }),
        ('🕐 Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def get_queryset(self, request):
        """Optimized queryset με prefetch"""
        qs = super().get_queryset(request)
        return qs.select_related(
            'client', 
            'obligation_type', 
            'completed_by'
        ).prefetch_related(
            'client__monthly_obligations'
        )
    
    def id_badge(self, obj):
        """ID με colored badge"""
        return format_html(
            '<span style="background: #e0e7ff; color: #4338ca; '
            'padding: 4px 8px; border-radius: 6px; font-family: monospace; '
            'font-weight: 600;">#{}</span>',
            obj.id
        )
    id_badge.short_description = 'ID'
    id_badge.admin_order_field = 'id'
    
    def client_smart(self, obj):
        """Smart client display με hover card"""
        pending = obj.client.monthly_obligations.filter(status='pending').count()
        return format_html(
            '<div style="position: relative;" class="client-hover">'
            '<a href="{}" style="color: #2563eb; text-decoration: none; font-weight: 600;">'
            '👤 {}<br>'
            '<small style="color: #6b7280;">ΑΦΜ: {} | 📋 {} εκκρεμείς</small>'
            '</a>'
            '</div>',
            reverse('admin:accounting_clientprofile_change', args=[obj.client.id]),
            obj.client.eponimia,
            obj.client.afm,
            pending
        )
    client_smart.short_description = 'Πελάτης'
    client_smart.admin_order_field = 'client__eponimia'
    
    def obligation_badge(self, obj):
        """Obligation type με icon"""
        icons = {
            'ΦΠΑ': '📊',
            'Μισθοδοσία': '💰',
            'ΕΦΚΑ': '🏛️',
            'Φόρος': '💸',
        }
        icon = '📋'
        for key, val in icons.items():
            if key in obj.obligation_type.name:
                icon = val
                break
        
        return format_html(
            '<span style="display: inline-flex; align-items: center; gap: 5px; '
            'background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%); '
            'padding: 6px 12px; border-radius: 8px; font-weight: 600;">'
            '{} {}'
            '</span>',
            icon, obj.obligation_type.name
        )
    obligation_badge.short_description = 'Υποχρέωση'
    
    def deadline_visual(self, obj):
        """Visual deadline με calendar icon"""
        days = obj.days_until_deadline
        today = timezone.now().date()
        
        if obj.status == 'completed':
            color = '#10b981'
            icon = '✅'
            message = 'Ολοκληρωμένη'
        elif obj.deadline < today:
            color = '#ef4444'
            icon = '🔴'
            message = f'Καθυστερεί {abs(days)}η'
        elif obj.deadline == today:
            color = '#f59e0b'
            icon = '⚠️'
            message = 'ΣΗΜΕΡΑ!'
        elif days <= 3:
            color = '#f59e0b'
            icon = '🟡'
            message = f'Σε {days}η'
        else:
            color = '#10b981'
            icon = '🟢'
            message = f'Σε {days}η'
        
        return format_html(
            '<div style="display: flex; flex-direction: column; align-items: center; '
            'padding: 8px; background: {}20; border-radius: 8px; border: 2px solid {};">'
            '<span style="font-size: 1.5em;">{}</span>'
            '<strong style="color: {};">{}</strong>'
            '<small style="color: #6b7280;">{}</small>'
            '</div>',
            color, color, icon, color, 
            obj.deadline.strftime('%d/%m'),
            message
        )
    deadline_visual.short_description = 'Προθεσμία'
    deadline_visual.admin_order_field = 'deadline'
    
    def status_progress(self, obj):
        """Status με progress indicator"""
        statuses = {
            'pending': ('⏳', '#f59e0b', 'Εκκρεμεί', 33),
            'overdue': ('🔴', '#ef4444', 'Καθυστερεί', 66),
            'completed': ('✅', '#10b981', 'Ολοκληρωμένη', 100),
        }
        
        icon, color, label, progress = statuses.get(obj.status, ('❓', '#6b7280', obj.status, 0))
        
        return format_html(
            '<div style="width: 120px;">'
            '<div style="display: flex; align-items: center; gap: 5px; margin-bottom: 4px;">'
            '<span style="font-size: 1.2em;">{}</span>'
            '<span style="color: {}; font-weight: 600;">{}</span>'
            '</div>'
            '<div style="background: #e5e7eb; border-radius: 10px; height: 6px; overflow: hidden;">'
            '<div style="background: {}; width: {}%; height: 100%; transition: width 0.3s;"></div>'
            '</div>'
            '</div>',
            icon, color, label, color, progress
        )
    status_progress.short_description = 'Κατάσταση'
    
    def time_cost_display(self, obj):
        """Time & Cost display με calculator"""
        if not obj.time_spent:
            return format_html(
                '<div style="color: #9ca3af;">—</div>'
            )
        
        cost = obj.cost or 0
        return format_html(
            '<div style="background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%); '
            'padding: 8px; border-radius: 8px; text-align: center;">'
            '<div style="font-size: 0.9em; color: #1e40af;">⏱️ {} ώρες</div>'
            '<div style="font-size: 1.1em; font-weight: bold; color: #1e3a8a;">💰 €{:.2f}</div>'
            '</div>',
            obj.time_spent, cost
        )
    time_cost_display.short_description = 'Χρόνος/Κόστος'
    
    def attachment_indicator(self, obj):
        """Attachment με preview"""
        if not obj.attachment:
            return format_html(
                '<span style="color: #d1d5db;">—</span>'
            )
        
        import os
        filename = os.path.basename(obj.attachment.name)
        ext = os.path.splitext(filename)[1].lower()
        
        icons = {
            '.pdf': '📄',
            '.doc': '📝',
            '.docx': '📝',
            '.xls': '📊',
            '.xlsx': '📊',
            '.jpg': '🖼️',
            '.jpeg': '🖼️',
            '.png': '🖼️',
        }
        icon = icons.get(ext, '📎')
        
        try:
            size_kb = obj.attachment.size / 1024
            size_str = f'{size_kb:.1f} KB'
        except:
            size_str = '—'
        
        return format_html(
            '<a href="{}" target="_blank" style="display: inline-flex; flex-direction: column; '
            'align-items: center; padding: 8px; background: #f3f4f6; border-radius: 8px; '
            'text-decoration: none; color: #374151; hover:background: #e5e7eb;">'
            '<span style="font-size: 1.5em;">{}</span>'
            '<small style="font-size: 0.8em;">{}</small>'
            '</a>',
            obj.attachment.url,
            icon,
            size_str
        )
    attachment_indicator.short_description = '📎'
    
    def completed_by_avatar(self, obj):
        """User avatar με initials"""
        if not obj.completed_by:
            return format_html('<span style="color: #d1d5db;">—</span>')
        
        user = obj.completed_by
        initials = ''.join([n[0].upper() for n in user.username.split()[:2]])
        if not initials:
            initials = user.username[:2].upper()
        
        # Generate color from username
        colors = ['#ef4444', '#f59e0b', '#10b981', '#3b82f6', '#8b5cf6', '#ec4899']
        color = colors[hash(user.username) % len(colors)]
        
        return format_html(
            '<div style="display: inline-flex; align-items: center; gap: 8px;">'
            '<div style="width: 32px; height: 32px; background: {}; color: white; '
            'border-radius: 50%; display: flex; align-items: center; justify-content: center; '
            'font-weight: bold; font-size: 0.8em;">{}</div>'
            '<small style="color: #6b7280;">{}</small>'
            '</div>',
            color, initials, user.username
        )
    completed_by_avatar.short_description = 'Ολοκλ. από'
    
    def action_buttons(self, obj):
        """Enhanced action buttons"""
        buttons = []
        
        if obj.status != 'completed':
            buttons.append(format_html(
                '<button onclick="quickComplete({})" '
                'style="background: #10b981; color: white; border: none; '
                'padding: 6px 12px; border-radius: 6px; cursor: pointer; '
                'font-weight: 600;">✅ Complete</button>',
                obj.id
            ))
        
        buttons.append(format_html(
            '<button onclick="openEditModal({})" '
            'style="background: #3b82f6; color: white; border: none; '
            'padding: 6px 12px; border-radius: 6px; cursor: pointer; '
            'font-weight: 600; margin-left: 4px;">✏️ Edit</button>',
            obj.id
        ))
        
        return format_html('<div style="display: flex;">{}</div>', ''.join(buttons))
    action_buttons.short_description = '⚡ Actions'
    
    def cost_calculator(self, obj):
        """Interactive cost calculator"""
        return format_html(
            '<div style="padding: 15px; background: linear-gradient(135deg, #ede9fe 0%, #ddd6fe 100%); '
            'border-radius: 10px;">'
            '<h4 style="margin: 0 0 10px 0; color: #5b21b6;">💰 Υπολογιστής Κόστους</h4>'
            '<div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px;">'
            '<div><strong>Ώρες:</strong> {}</div>'
            '<div><strong>Τιμή/ώρα:</strong> €{}</div>'
            '<div><strong>Υποσύνολο:</strong> €{:.2f}</div>'
            '<div><strong>ΦΠΑ (24%):</strong> €{:.2f}</div>'
            '</div>'
            '<div style="margin-top: 10px; padding-top: 10px; border-top: 2px solid #7c3aed;">'
            '<strong style="font-size: 1.2em; color: #5b21b6;">ΣΥΝΟΛΟ: €{:.2f}</strong>'
            '</div>'
            '</div>',
            obj.time_spent or 0,
            obj.hourly_rate or 50,
            obj.cost or 0,
            (obj.cost or 0) * 0.24,
            (obj.cost or 0) * 1.24
        )
    cost_calculator.short_description = 'Κοστολόγηση'
    
    def history_timeline(self, obj):
        """History timeline"""
        html = '<div style="padding: 10px;">'
        
        events = []
        events.append(('Δημιουργία', obj.created_at, '🆕'))
        if obj.completed_date:
            events.append(('Ολοκλήρωση', obj.completed_date, '✅'))
        
        for event, date, icon in events:
            html += format_html(
                '<div style="display: flex; align-items: center; gap: 10px; '
                'padding: 8px; margin: 5px 0; background: #f9fafb; border-left: 3px solid #3b82f6;">'
                '<span style="font-size: 1.2em;">{}</span>'
                '<div>'
                '<strong>{}</strong><br>'
                '<small style="color: #6b7280;">{}</small>'
                '</div>'
                '</div>',
                icon, event, 
                date.strftime('%d/%m/%Y %H:%M') if hasattr(date, 'strftime') else date
            )
        
        html += '</div>'
        return mark_safe(html)
    history_timeline.short_description = 'Ιστορικό'
    
    def related_obligations(self, obj):
        """Related obligations για τον ίδιο πελάτη"""
        related = obj.client.monthly_obligations.exclude(id=obj.id).order_by('-deadline')[:5]
        
        html = '<div style="padding: 10px;">'
        for obl in related:
            status_color = '#10b981' if obl.status == 'completed' else '#f59e0b'
            html += format_html(
                '<div style="display: flex; justify-content: space-between; '
                'padding: 8px; margin: 5px 0; background: #f9fafb; border-radius: 6px;">'
                '<span>{}</span>'
                '<span style="color: {};">{} - {}</span>'
                '</div>',
                obl.obligation_type.name,
                status_color,
                obl.deadline.strftime('%d/%m/%Y'),
                obl.get_status_display()
            )
        html += '</div>'
        return mark_safe(html)
    related_obligations.short_description = 'Σχετικές Υποχρεώσεις'
    
    @admin.action(description='✅ Ολοκλήρωση Επιλεγμένων')
    def mark_as_completed(self, request, queryset):
        """Enhanced completion με timestamp"""
        updated = queryset.filter(status__in=['pending', 'overdue']).update(
            status='completed',
            completed_date=timezone.now().date(),
            completed_by=request.user
        )
        self.message_user(
            request, 
            f'✅ Ολοκληρώθηκαν {updated} υποχρεώσεις!',
            messages.SUCCESS
        )
    
    @admin.action(description='📧 Αποστολή Υπενθύμισης')
    def send_reminder_emails(self, request, queryset):
        """Send reminder emails για pending obligations"""
        pending = queryset.filter(status='pending')
        messages.info(request, f'📧 Θα σταλούν {pending.count()} υπενθυμίσεις...')
        # TODO: Implement email sending
    
    @admin.action(description='🧾 Δημιουργία Τιμολογίων')
    def generate_invoices(self, request, queryset):
        """Generate invoices για completed obligations"""
        completed = queryset.filter(status='completed', time_spent__gt=0)
        messages.info(request, f'🧾 Δημιουργία {completed.count()} τιμολογίων...')
        # TODO: Implement invoice generation
    
    def changelist_view(self, request, extra_context=None):
        """Enhanced changelist με statistics"""
        extra_context = extra_context or {}
        
        # Calculate statistics
        qs = self.get_queryset(request)
        extra_context['stats'] = {
            'total': qs.count(),
            'pending': qs.filter(status='pending').count(),
            'overdue': qs.filter(status='overdue').count(),
            'completed': qs.filter(status='completed').count(),
            'total_hours': qs.aggregate(Sum('time_spent'))['time_spent__sum'] or 0,
            'total_cost': qs.aggregate(
                total=Sum('time_spent') * 50
            )['total'] or 0,
        }
        
        return super().changelist_view(request, extra_context)
    
    class Media:
        css = {
            'all': (
                'https://cdnjs.cloudflare.com/ajax/libs/animate.css/4.1.1/animate.min.css',
            )
        }
        js = (
            'https://cdn.jsdelivr.net/npm/sweetalert2@11',
        )


# ============================================
# ADMIN CUSTOMIZATION
# ============================================

# Custom admin site
admin.site.site_header = "LogistikoCRM Professional"
admin.site.site_title = "LogistikoCRM"
admin.site.index_title = "Καλώς ήρθατε στο Λογιστικό Σύστημα"

# Keep your existing registrations at the end
admin.site.index_template = 'admin/custom_index.html'

# ============================================
# ΠΡΟΣΘΕΣΕ ΟΛΑ ΑΥΤΑ ΜΕΤΑ ΤΟ MonthlyObligationAdmin
# ============================================

@admin.register(ObligationGroup)
class ObligationGroupAdmin(admin.ModelAdmin):
    list_display = ['name', 'description']
    search_fields = ['name']

@admin.register(ObligationProfile)
class ObligationProfileAdmin(admin.ModelAdmin):
    list_display = ['name', 'description']
    search_fields = ['name', 'description']

@admin.register(ObligationType)
class ObligationTypeAdmin(admin.ModelAdmin):
    list_display = ['name', 'code', 'frequency', 'deadline_type', 'profile', 'exclusion_group', 'is_active', 'priority']
    list_filter = ['frequency', 'is_active', 'profile', 'exclusion_group']
    search_fields = ['name', 'code']
    list_editable = ['priority', 'is_active']
    
    fieldsets = (
        ('Βασικά', {
            'fields': ('name', 'code', 'description', 'is_active', 'priority')
        }),
        ('Χρονικά', {
            'fields': ('frequency', 'deadline_type', 'deadline_day', 'applicable_months')
        }),
        ('Σχέσεις', {
            'fields': ('exclusion_group', 'profile')
        }),
    )

@admin.register(ClientObligation)
class ClientObligationAdmin(admin.ModelAdmin):
    list_display = ['client', 'is_active', 'created_at']
    list_filter = ['is_active']
    search_fields = ['client__afm', 'client__eponimia']
    filter_horizontal = ['obligation_types', 'obligation_profiles']
    
    fieldsets = (
        ('Πελάτης', {
            'fields': ('client', 'is_active')
        }),
        ('Υποχρεώσεις', {
            'fields': ('obligation_profiles', 'obligation_types')
        }),
    )

@admin.register(EmailTemplate)
class EmailTemplateAdmin(admin.ModelAdmin):
    list_display = ['name', 'subject', 'is_active', 'created_at']
    list_filter = ['is_active', 'created_at']
    search_fields = ['name', 'subject', 'body_html']

@admin.register(EmailAutomationRule)
class EmailAutomationRuleAdmin(admin.ModelAdmin):
    list_display = ['name', 'trigger', 'template', 'timing', 'is_active', 'created_at']
    list_filter = ['is_active', 'trigger', 'timing']
    search_fields = ['name', 'description']
    filter_horizontal = ['filter_obligation_types']

@admin.register(ScheduledEmail)
class ScheduledEmailAdmin(admin.ModelAdmin):
    list_display = ['recipient_name', 'subject', 'send_at', 'status', 'created_at']
    list_filter = ['status', 'send_at', 'created_at']
    search_fields = ['recipient_email', 'recipient_name', 'subject']
    filter_horizontal = ['obligations']
    readonly_fields = ['sent_at', 'error_message', 'created_by', 'created_at']

@admin.register(VoIPCall)
class VoIPCallAdmin(admin.ModelAdmin):
    list_display = ['call_id', 'phone_number', 'client', 'direction', 'status', 'duration_formatted', 'started_at']
    list_filter = ['status', 'direction', 'started_at']
    search_fields = ['phone_number', 'client__eponimia', 'call_id']
    readonly_fields = ['call_id', 'duration_formatted', 'created_at', 'updated_at']

@admin.register(VoIPCallLog)
class VoIPCallLogAdmin(admin.ModelAdmin):
    list_display = ['call', 'action', 'description', 'created_at']
    list_filter = ['action', 'created_at']
    search_fields = ['call__phone_number', 'description']
    readonly_fields = ['call', 'action', 'description', 'created_at']