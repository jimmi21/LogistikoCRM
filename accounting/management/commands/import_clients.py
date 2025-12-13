import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from accounting.models import ClientProfile
from datetime import datetime
import sys


class Command(BaseCommand):
    help = 'Î•Î¹ÏƒÎ±Î³Ï‰Î³Î® Ï€ÎµÎ»Î±Ï„ÏÎ½ Î±Ï€ÏŒ Excel Î¼Îµ ÎŸÎ›Î‘ Ï„Î± Ï€ÎµÎ´Î¯Î± ÎºÎ±Î¹ advanced features'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Î”Î¹Î±Î´ÏÎ¿Î¼Î® Ï„Î¿Ï… Excel Î±ÏÏ‡ÎµÎ¯Î¿Ï… (Ï€.Ï‡. clients.xlsx)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Test run - Ï€ÏÎ¿ÎµÏ€Î¹ÏƒÎºÏŒÏ€Î·ÏƒÎ· Ï‡Ï‰ÏÎ¯Ï‚ Î±Ï€Î¿Î¸Î®ÎºÎµÏ…ÏƒÎ· ÏƒÏ„Î· Î²Î¬ÏƒÎ·'
        )
        parser.add_argument(
            '--skip-errors',
            action='store_true',
            help='Î£Ï…Î½Î­Ï‡Î¹ÏƒÎµ Î±ÎºÏŒÎ¼Î± ÎºÎ±Î¹ Î±Î½ Ï…Ï€Î¬ÏÏ‡Î¿Ï…Î½ ÏƒÏ†Î¬Î»Î¼Î±Ï„Î±'
        )
        parser.add_argument(
            '--update-only',
            action='store_true',
            help='ÎœÏŒÎ½Î¿ update Ï…Ï€Î±ÏÏ‡ÏŒÎ½Ï„Ï‰Î½ (Î´ÎµÎ½ Î´Î·Î¼Î¹Î¿Ï…ÏÎ³ÎµÎ¯ Î½Î­Î¿Ï…Ï‚)'
        )
        parser.add_argument(
            '--create-only',
            action='store_true',
            help='ÎœÏŒÎ½Î¿ Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î¯Î± Î½Î­Ï‰Î½ (Î´ÎµÎ½ ÎµÎ½Î·Î¼ÎµÏÏÎ½ÎµÎ¹ Ï…Ï€Î¬ÏÏ‡Î¿Î½Ï„ÎµÏ‚)'
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Î•Î¼Ï†Î¬Î½Î¹ÏƒÎ· Î±Î½Î±Î»Ï…Ï„Î¹ÎºÏÎ½ Ï€Î»Î·ÏÎ¿Ï†Î¿ÏÎ¹ÏÎ½'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        dry_run = options['dry_run']
        skip_errors = options['skip_errors']
        update_only = options['update_only']
        create_only = options['create_only']
        verbose = options['verbose']
        
        # Banner
        self.print_banner()
        
        if dry_run:
            self.stdout.write(self.style.WARNING('ğŸ” DRY RUN MODE - Î”ÎµÎ½ Î¸Î± Î±Ï€Î¿Î¸Î·ÎºÎµÏ…Ï„Î¿ÏÎ½ Î±Î»Î»Î±Î³Î­Ï‚\n'))
        
        if update_only:
            self.stdout.write(self.style.NOTICE('ğŸ“ UPDATE ONLY - ÎœÏŒÎ½Î¿ ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ· Ï…Ï€Î±ÏÏ‡ÏŒÎ½Ï„Ï‰Î½\n'))
        
        if create_only:
            self.stdout.write(self.style.NOTICE('â• CREATE ONLY - ÎœÏŒÎ½Î¿ Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î¯Î± Î½Î­Ï‰Î½\n'))
        
        # Î†Î½Î¿Î¹Î³Î¼Î± Excel
        self.stdout.write(f'ğŸ“‚ Î†Î½Î¿Î¹Î³Î¼Î± Î±ÏÏ‡ÎµÎ¯Î¿Ï…: {excel_file}')
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'âŒ Î¤Î¿ Î±ÏÏ‡ÎµÎ¯Î¿ Î´ÎµÎ½ Î²ÏÎ­Î¸Î·ÎºÎµ: {excel_file}'))
            return
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'âŒ Î£Ï†Î¬Î»Î¼Î± Î±Î½Î¬Î³Î½Ï‰ÏƒÎ·Ï‚: {e}'))
            return
        
        # Field mapping - ÎŸÎ›Î‘ Ï„Î± Ï€ÎµÎ´Î¯Î±
        field_mapping = self.get_field_mapping()
        
        # Î”Î¹Î¬Î²Î±ÏƒÎ¼Î± headers
        headers = self.read_headers(ws, field_mapping)
        
        if not headers:
            self.stdout.write(self.style.ERROR('âŒ Î”ÎµÎ½ Î²ÏÎ­Î¸Î·ÎºÎ±Î½ Î­Î³ÎºÏ…ÏÎ± headers!'))
            return
        
        self.stdout.write(self.style.SUCCESS(f'âœ… Î’ÏÎ­Î¸Î·ÎºÎ±Î½ {len(headers)} Î­Î³ÎºÏ…ÏÎµÏ‚ ÏƒÏ„Î®Î»ÎµÏ‚\n'))
        
        if verbose:
            self.stdout.write('ğŸ“‹ Î£Ï„Î®Î»ÎµÏ‚ Ï€Î¿Ï… Î¸Î± ÎµÎ¹ÏƒÎ±Ï‡Î¸Î¿ÏÎ½:')
            for col_idx, field_name in sorted(headers.items()):
                self.stdout.write(f'  â€¢ Column {col_idx}: {field_name}')
            self.stdout.write('')
        
        # Counters
        stats = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
            'total_rows': 0
        }
        errors = []
        
        # Î ÏÎ¿ÏƒÎ´Î¹Î¿ÏÎ¹ÏƒÎ¼ÏŒÏ‚ start row (Ï€Î±ÏÎ¬Î»ÎµÎ¹ÏˆÎ· Ï€Î±ÏÎ±Î´ÎµÎ¹Î³Î¼Î¬Ï„Ï‰Î½)
        start_row = self.find_start_row(ws)
        total_rows = ws.max_row - start_row + 1
        
        self.stdout.write(f'ğŸ“Š Î•Ï€ÎµÎ¾ÎµÏÎ³Î±ÏƒÎ¯Î± {total_rows} Î³ÏÎ±Î¼Î¼ÏÎ½...\n')
        self.stdout.write('='*70 + '\n')
        
        # Process rows
        for row_num in range(start_row, ws.max_row + 1):
            stats['total_rows'] += 1
            
            # Progress indicator
            if stats['total_rows'] % 10 == 0:
                self.print_progress(stats['total_rows'], total_rows)
            
            try:
                # Parse row
                row_data = self.parse_row(ws, row_num, headers, verbose)
                
                if not row_data:
                    stats['skipped'] += 1
                    if verbose:
                        self.stdout.write(f'  â­ï¸  Î“ÏÎ±Î¼Î¼Î® {row_num}: Î Î±ÏÎ¬Î»ÎµÎ¹ÏˆÎ· (ÎºÎµÎ½Î® Î³ÏÎ±Î¼Î¼Î®)')
                    continue
                
                # Validate
                validation_errors = self.validate_row_data(row_data, row_num)
                if validation_errors:
                    stats['errors'] += 1
                    for error in validation_errors:
                        errors.append(f'Î“ÏÎ±Î¼Î¼Î® {row_num}: {error}')
                        self.stdout.write(self.style.ERROR(f'  âŒ {error}'))
                    if not skip_errors:
                        break
                    continue
                
                # Dry run - just show what would happen
                if dry_run:
                    exists = ClientProfile.objects.filter(afm=row_data['afm']).exists()
                    if exists:
                        stats['updated'] += 1
                        self.stdout.write(
                            self.style.WARNING(f'  ğŸ”„ [DRY RUN] Î˜Î± ÎµÎ½Î·Î¼ÎµÏÏ‰Î¸ÎµÎ¯: {row_data["afm"]} - {row_data["eponimia"]}')
                        )
                    else:
                        stats['created'] += 1
                        self.stdout.write(
                            self.style.SUCCESS(f'  âœ… [DRY RUN] Î˜Î± Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î·Î¸ÎµÎ¯: {row_data["afm"]} - {row_data["eponimia"]}')
                        )
                    continue
                
                # Real import with transaction
                with transaction.atomic():
                    afm = row_data.pop('afm')
                    
                    # Check if exists
                    exists = ClientProfile.objects.filter(afm=afm).exists()
                    
                    # Apply create/update filters
                    if update_only and not exists:
                        stats['skipped'] += 1
                        if verbose:
                            self.stdout.write(f'  â­ï¸  {afm}: Î Î±ÏÎ¬Î»ÎµÎ¹ÏˆÎ· (Î´ÎµÎ½ Ï…Ï€Î¬ÏÏ‡ÎµÎ¹ - update only mode)')
                        continue
                    
                    if create_only and exists:
                        stats['skipped'] += 1
                        if verbose:
                            self.stdout.write(f'  â­ï¸  {afm}: Î Î±ÏÎ¬Î»ÎµÎ¹ÏˆÎ· (Ï…Ï€Î¬ÏÏ‡ÎµÎ¹ Î®Î´Î· - create only mode)')
                        continue
                    
                    # Create or update
                    client, created = ClientProfile.objects.update_or_create(
                        afm=afm,
                        defaults=row_data
                    )
                    
                    if created:
                        stats['created'] += 1
                        self.stdout.write(
                            self.style.SUCCESS(f'  âœ… ÎÎ­Î¿Ï‚: {client.afm} - {client.eponimia}')
                        )
                    else:
                        stats['updated'] += 1
                        self.stdout.write(
                            self.style.WARNING(f'  ğŸ”„ Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ·: {client.afm} - {client.eponimia}')
                        )
            
            except Exception as e:
                stats['errors'] += 1
                error_msg = f'Î“ÏÎ±Î¼Î¼Î® {row_num}: {str(e)}'
                errors.append(error_msg)
                self.stdout.write(self.style.ERROR(f'  âŒ {error_msg}'))
                
                if not skip_errors:
                    self.stdout.write(self.style.ERROR(f'\nâ›” Î”Î¹Î±ÎºÏŒÏ€Î·ÎºÎµ Î»ÏŒÎ³Ï‰ ÏƒÏ†Î¬Î»Î¼Î±Ï„Î¿Ï‚. Î§ÏÎ·ÏƒÎ¹Î¼Î¿Ï€Î¿Î¯Î·ÏƒÎµ --skip-errors Î³Î¹Î± ÏƒÏ…Î½Î­Ï‡ÎµÎ¹Î±.'))
                    break
        
        # Final summary
        self.print_summary(stats, errors, dry_run)
    
    def get_field_mapping(self):
        """Î Î»Î®ÏÎµÏ‚ mapping Excel headers -> Model fields"""
        return {
            'Î‘.Î¦.Îœ.': 'afm',
            'Î”.ÎŸ.Î¥.': 'doy',
            'Î•Ï€Ï‰Î½Ï…Î¼Î¯Î±/Î•Ï€ÏÎ½Ï…Î¼Î¿': 'eponimia',
            'ÎŒÎ½Î¿Î¼Î±': 'onoma',
            'ÎŒÎ½Î¿Î¼Î± Î Î±Ï„ÏÏŒÏ‚': 'onoma_patros',
            'Î‘ÏÎ¹Î¸Î¼ÏŒÏ‚ Î¤Î±Ï…Ï„ÏŒÏ„Î·Ï„Î±Ï‚': 'arithmos_taftotitas',
            'Î•Î¯Î´Î¿Ï‚ Î¤Î±Ï…Ï„ÏŒÏ„Î·Ï„Î±Ï‚': 'eidos_taftotitas',
            'Î ÏÎ¿ÏƒÏ‰Ï€Î¹ÎºÏŒÏ‚ Î‘ÏÎ¹Î¸Î¼ÏŒÏ‚': 'prosopikos_arithmos',
            'Î‘.Îœ.Îš.Î‘.': 'amka',
            'Î‘.Îœ. Î™.Îš.Î‘.': 'am_ika',
            'Î‘ÏÎ¹Î¸Î¼ÏŒÏ‚ Î“.Î•.ÎœÎ—.': 'arithmos_gemi',
            'Î‘ÏÎ¹Î¸Î¼ÏŒÏ‚ Î”.Î¥Î .Î‘': 'arithmos_dypa',
            'Î—Î¼. Î“Î­Î½Î½Î·ÏƒÎ·Ï‚': 'imerominia_gennisis',
            'Î—Î¼. Î“Î¬Î¼Î¿Ï…': 'imerominia_gamou',
            'Î¦ÏÎ»Î¿': 'filo',
            'Î”Î¹ÎµÏÎ¸Ï…Î½ÏƒÎ· ÎšÎ±Ï„Î¿Î¹ÎºÎ¯Î±Ï‚': 'diefthinsi_katoikias',
            'Î‘ÏÎ¹Î¸Î¼ÏŒÏ‚': 'arithmos_katoikias',
            'Î ÏŒÎ»Î· ÎšÎ±Ï„Î¿Î¹ÎºÎ¯Î±Ï‚': 'poli_katoikias',
            'Î”Î®Î¼Î¿Ï‚ ÎšÎ±Ï„Î¿Î¹ÎºÎ¯Î±Ï‚': 'dimos_katoikias',
            'ÎÎ¿Î¼ÏŒÏ‚ ÎšÎ±Ï„Î¿Î¹ÎºÎ¯Î±Ï‚': 'nomos_katoikias',
            'T.K. ÎšÎ±Ï„Î¿Î¹ÎºÎ¯Î±Ï‚': 'tk_katoikias',
            'Î¤Î·Î»Î­Ï†Ï‰Î½Î¿ ÎŸÎ¹ÎºÎ¯Î±Ï‚ 1': 'tilefono_oikias_1',
            'Î¤Î·Î»Î­Ï†Ï‰Î½Î¿ ÎŸÎ¹ÎºÎ¯Î±Ï‚ 2': 'tilefono_oikias_2',
            'ÎšÎ¹Î½Î·Ï„ÏŒ Ï„Î·Î»Î­Ï†Ï‰Î½Î¿': 'kinito_tilefono',
            'Î”Î¹ÎµÏÎ¸Ï…Î½ÏƒÎ· Î•Ï€Î¹Ï‡ÎµÎ¯ÏÎ·ÏƒÎ·Ï‚': 'diefthinsi_epixeirisis',
            'Î‘ÏÎ¹Î¸Î¼ÏŒÏ‚ Î•Ï€Î¹Ï‡ÎµÎ¯ÏÎ·ÏƒÎ·Ï‚': 'arithmos_epixeirisis',
            'Î ÏŒÎ»Î· Î•Ï€Î¹Ï‡ÎµÎ¯ÏÎ·ÏƒÎ·Ï‚': 'poli_epixeirisis',
            'Î”Î®Î¼Î¿Ï‚ Î•Ï€Î¹Ï‡ÎµÎ¯ÏÎ·ÏƒÎ·Ï‚': 'dimos_epixeirisis',
            'ÎÎ¿Î¼ÏŒÏ‚ Î•Ï€Î¹Ï‡ÎµÎ¯ÏÎ·ÏƒÎ·Ï‚': 'nomos_epixeirisis',
            'Î¤.Îš. Î•Ï€Î¹Ï‡ÎµÎ¯ÏÎ·ÏƒÎ·Ï‚': 'tk_epixeirisis',
            'Î¤Î·Î»Î­Ï†Ï‰Î½Î¿ Î•Ï€Î¹Ï‡ÎµÎ¯ÏÎ·ÏƒÎ·Ï‚ 1': 'tilefono_epixeirisis_1',
            'Î¤Î·Î»Î­Ï†Ï‰Î½Î¿ Î•Ï€Î¹Ï‡ÎµÎ¯ÏÎ·ÏƒÎ·Ï‚ 2': 'tilefono_epixeirisis_2',
            'Email': 'email',
            'Î¤ÏÎ¬Ï€ÎµÎ¶Î±': 'trapeza',
            'IBAN': 'iban',
            'Î•Î¯Î´Î¿Ï‚ Î¥Ï€ÏŒÏ‡ÏÎµÎ¿Ï…': 'eidos_ipoxreou',
            'ÎšÎ±Ï„Î·Î³Î¿ÏÎ¯Î± Î’Î¹Î²Î»Î¯Ï‰Î½': 'katigoria_vivlion',
            'ÎÎ¿Î¼Î¹ÎºÎ® ÎœÎ¿ÏÏ†Î®': 'nomiki_morfi',
            'Î‘Î³ÏÏŒÏ„Î·Ï‚': 'agrotis',
            'Î—Î¼/Î½Î¯Î± ÎˆÎ½Î±ÏÎ¾Î·Ï‚ Î•ÏÎ³Î±ÏƒÎ¹ÏÎ½': 'imerominia_enarksis',
            'ÎŒÎ½Î¿Î¼Î± Î§ÏÎ®ÏƒÏ„Î· Taxis Net': 'onoma_xristi_taxisnet',
            'ÎšÏ‰Î´Î¹ÎºÏŒÏ‚ Taxis Net': 'kodikos_taxisnet',
            'ÎŒÎ½Î¿Î¼Î± Î§ÏÎ®ÏƒÏ„Î· Î™.Îš.Î‘. Î•ÏÎ³Î¿Î´ÏŒÏ„Î·': 'onoma_xristi_ika_ergodoti',
            'ÎšÏ‰Î´Î¹ÎºÏŒÏ‚ Î™.Îš.Î‘. Î•ÏÎ³Î¿Î´ÏŒÏ„Î·': 'kodikos_ika_ergodoti',
            'ÎŒÎ½Î¿Î¼Î± Î§ÏÎ®ÏƒÏ„Î· Î“.Î•.ÎœÎ—.': 'onoma_xristi_gemi',
            'ÎšÏ‰Î´Î¹ÎºÏŒÏ‚ Î“.Î•.ÎœÎ—.': 'kodikos_gemi',
            'Î‘.Î¦.Îœ Î£Ï…Î¶ÏÎ³Î¿Ï…/Îœ.Î£.Î£.': 'afm_sizigou',
            'Î‘.Î¦.Îœ. Î¦Î¿ÏÎ­Î±Ï‚': 'afm_foreas',
            'Î‘Îœ ÎšÎ›Î•Î™Î”Î™': 'am_klidi',
        }
    
    def read_headers(self, ws, field_mapping):
        """Î”Î¹Î±Î²Î¬Î¶ÎµÎ¹ Ï„Î± headers Î±Ï€ÏŒ Ï„Î·Î½ 1Î· Î³ÏÎ±Î¼Î¼Î®"""
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            header_name = str(cell.value).strip() if cell.value else None
            if header_name and header_name in field_mapping:
                headers[col_idx] = field_mapping[header_name]
        return headers
    
    def find_start_row(self, ws):
        """Î’ÏÎ¯ÏƒÎºÎµÎ¹ Î±Ï€ÏŒ Ï€Î¿Î¹Î± Î³ÏÎ±Î¼Î¼Î® Î½Î± Î¾ÎµÎºÎ¹Î½Î®ÏƒÎµÎ¹ (Ï€Î±ÏÎ¬Î»ÎµÎ¹ÏˆÎ· Ï€Î±ÏÎ±Î´ÎµÎ¹Î³Î¼Î¬Ï„Ï‰Î½)"""
        # ÎˆÎ»ÎµÎ³Ï‡Î¿Ï‚ Î±Î½ Î· 2Î· Î³ÏÎ±Î¼Î¼Î® ÎµÎ¯Î½Î±Î¹ Ï€Î±ÏÎ¬Î´ÎµÎ¹Î³Î¼Î±
        if ws.max_row >= 2:
            first_afm = ws.cell(2, 1).value
            if first_afm and str(first_afm).startswith('123456'):
                self.stdout.write(self.style.WARNING('âš ï¸  Î Î±ÏÎ¬Î»ÎµÎ¹ÏˆÎ· Ï€Î±ÏÎ¬Î´ÎµÎ¹Î³Î¼Î± Î³ÏÎ±Î¼Î¼Î®Ï‚'))
                return 3
        return 2
    
    def parse_row(self, ws, row_num, headers, verbose):
        """Parse Î¼Î¹Î± Î³ÏÎ±Î¼Î¼Î® Excel ÏƒÎµ dictionary"""
        row_data = {}
        has_data = False
        
        for col_idx, field_name in headers.items():
            cell = ws.cell(row_num, col_idx)
            value = cell.value
            
            # Skip Î±Î½ ÏŒÎ»Î± Ï„Î± ÎºÎµÎ»Î¹Î¬ ÎµÎ¯Î½Î±Î¹ ÎºÎµÎ½Î¬
            if value is not None and str(value).strip():
                has_data = True
            
            # Clean & convert value
            value = self.clean_value(value, field_name, row_num)
            
            if value is not None:
                row_data[field_name] = value
        
        return row_data if has_data else None
    
    def clean_value(self, value, field_name, row_num):
        """ÎšÎ±Î¸Î±ÏÎ¹ÏƒÎ¼ÏŒÏ‚ ÎºÎ±Î¹ Î¼ÎµÏ„Î±Ï„ÏÎ¿Ï€Î® Ï„Î¹Î¼ÏÎ½"""
        
        # Null/Empty check
        if value is None:
            return None
        
        # String cleaning
        if isinstance(value, str):
            value = value.strip()
            if value == '' or value.upper() in ['ÎšÎ•ÎÎŸ', 'EMPTY', '-', 'N/A']:
                return None
        
        # Î•Î¯Î´Î¿Ï‚ Î¥Ï€ÏŒÏ‡ÏÎµÎ¿Ï… - REQUIRED
        if field_name == 'eidos_ipoxreou':
            mapping = {
                'Î™Î”Î™Î©Î¤Î—Î£': 'individual',
                'Î•Î Î‘Î“Î“Î•Î›ÎœÎ‘Î¤Î™Î‘Î£': 'professional',
                'Î•Î¤Î‘Î™Î¡Î•Î™Î‘': 'company',
                'INDIVIDUAL': 'individual',
                'PROFESSIONAL': 'professional',
                'COMPANY': 'company',
            }
            value_upper = str(value).strip().upper()
            return mapping.get(value_upper, 'professional')  # Default
        
        # ÎšÎ±Ï„Î·Î³Î¿ÏÎ¯Î± Î’Î¹Î²Î»Î¯Ï‰Î½
        if field_name == 'katigoria_vivlion':
            mapping = {
                'Î‘': 'A',
                'Î’': 'B',
                'Î“': 'C',
                'Î§Î©Î¡Î™Î£': 'none',
                'A': 'A',
                'B': 'B',
                'C': 'C',
                'NONE': 'none',
            }
            if value:
                value_clean = str(value).strip().upper()
                return mapping.get(value_clean, '')
            return ''
        
        # Boolean fields
        if field_name == 'agrotis':
            if isinstance(value, bool):
                return value
            value_upper = str(value).upper()
            return value_upper in ['ÎÎ‘Î™', 'NAI', 'YES', 'TRUE', '1', 'Î']
        
        # Î¦ÏÎ»Î¿
        if field_name == 'filo':
            if value:
                value_upper = str(value).upper()
                if value_upper in ['Îœ', 'M', 'Î‘ÎÎ”Î¡Î‘Î£', 'MALE', 'MAN']:
                    return 'M'
                elif value_upper in ['Î“', 'F', 'Î“Î¥ÎÎ‘Î™ÎšÎ‘', 'FEMALE', 'WOMAN']:
                    return 'F'
            return ''
        
        # Dates
        if field_name in ('imerominia_gennisis', 'imerominia_gamou', 'imerominia_enarksis'):
            return self.parse_date(value, row_num, field_name)
        
        # Default: string
        return str(value) if value else ''
    
    def parse_date(self, value, row_num, field_name):
        """Parse Î·Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î± Î±Ï€ÏŒ Î´Î¹Î¬Ï†Î¿ÏÎµÏ‚ Î¼Î¿ÏÏ†Î­Ï‚"""
        if not value:
            return None
        
        # Î‘Î½ ÎµÎ¯Î½Î±Î¹ Î®Î´Î· datetime object
        if isinstance(value, datetime):
            return value.date()
        
        # Try Î´Î¹Î¬Ï†Î¿ÏÎµÏ‚ Î¼Î¿ÏÏ†Î­Ï‚
        date_formats = [
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%Y-%m-%d',
            '%d.%m.%Y',
            '%d/%m/%y',
        ]
        
        value_str = str(value).strip()
        
        for fmt in date_formats:
            try:
                return datetime.strptime(value_str, fmt).date()
            except ValueError:
                continue
        
        # Î‘Î½ Î´ÎµÎ½ Î¼Ï€ÏŒÏÎµÏƒÎµ Î½Î± parse, ÎµÏ€Î­ÏƒÏ„ÏÎµÏˆÎµ None
        return None
    
    def validate_row_data(self, row_data, row_num):
        """Validation rules"""
        errors = []
        
        # Required: AFM
        if not row_data.get('afm'):
            errors.append('Î‘Î¦Îœ ÎµÎ¯Î½Î±Î¹ Ï…Ï€Î¿Ï‡ÏÎµÏ‰Ï„Î¹ÎºÏŒ')
        else:
            afm = str(row_data['afm']).strip()
            if len(afm) != 9 or not afm.isdigit():
                errors.append(f'Î‘Î¦Îœ Ï€ÏÎ­Ï€ÎµÎ¹ Î½Î± ÎµÎ¯Î½Î±Î¹ 9 ÏˆÎ·Ï†Î¯Î± (Î´ÏŒÎ¸Î·ÎºÎµ: {afm})')
        
        # Required: Î•Ï€Ï‰Î½Ï…Î¼Î¯Î±
        if not row_data.get('eponimia'):
            errors.append('Î•Ï€Ï‰Î½Ï…Î¼Î¯Î± ÎµÎ¯Î½Î±Î¹ Ï…Ï€Î¿Ï‡ÏÎµÏ‰Ï„Î¹ÎºÎ®')
        
        # Required: Î•Î¯Î´Î¿Ï‚ Î¥Ï€ÏŒÏ‡ÏÎµÎ¿Ï…
        if not row_data.get('eidos_ipoxreou'):
            errors.append('Î•Î¯Î´Î¿Ï‚ Î¥Ï€ÏŒÏ‡ÏÎµÎ¿Ï… ÎµÎ¯Î½Î±Î¹ Ï…Ï€Î¿Ï‡ÏÎµÏ‰Ï„Î¹ÎºÏŒ')
        elif row_data['eidos_ipoxreou'] not in ['individual', 'professional', 'company']:
            errors.append(f'Î†ÎºÏ…ÏÎ¿ Î•Î¯Î´Î¿Ï‚ Î¥Ï€ÏŒÏ‡ÏÎµÎ¿Ï…: {row_data["eidos_ipoxreou"]}')
        
        # Email validation
        if row_data.get('email'):
            email = row_data['email']
            if '@' not in email or '.' not in email:
                errors.append(f'Î†ÎºÏ…ÏÎ¿ email: {email}')
        
        # IBAN validation (basic)
        if row_data.get('iban'):
            iban = str(row_data['iban']).replace(' ', '')
            if not iban.startswith('GR') or len(iban) != 27:
                errors.append(f'Î†ÎºÏ…ÏÎ¿ IBAN: {row_data["iban"]} (Ï€ÏÎ­Ï€ÎµÎ¹ Î½Î± Î¾ÎµÎºÎ¹Î½Î¬ Î¼Îµ GR ÎºÎ±Î¹ Î½Î± Î­Ï‡ÎµÎ¹ 27 Ï‡Î±ÏÎ±ÎºÏ„Î®ÏÎµÏ‚)')
        
        return errors
    
    def print_progress(self, current, total):
        """Progress bar"""
        percent = int((current / total) * 100)
        bar_length = 40
        filled = int(bar_length * current / total)
        bar = 'â–ˆ' * filled + 'â–‘' * (bar_length - filled)
        sys.stdout.write(f'\r  [{bar}] {percent}% ({current}/{total})')
        sys.stdout.flush()
        if current == total:
            sys.stdout.write('\n')
    
    def print_banner(self):
        """ASCII banner"""
        banner = """
â•”â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•—
â•‘                    ğŸ“¥ CLIENT IMPORT TOOL                          â•‘
â•‘                   D.P. Economy - v2.0                             â•‘
â•šâ•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        """
        self.stdout.write(self.style.SUCCESS(banner))
    
    def print_summary(self, stats, errors, dry_run):
        """Î¤ÎµÎ»Î¹ÎºÎ® Î±Î½Î±Ï†Î¿ÏÎ¬"""
        self.stdout.write('\n' + '='*70)
        
        if dry_run:
            self.stdout.write(self.style.SUCCESS('\nâœ… DRY RUN ÎŸÎ›ÎŸÎšÎ›Î—Î¡Î©Î˜Î—ÎšÎ•'))
        else:
            self.stdout.write(self.style.SUCCESS('\nâœ… IMPORT ÎŸÎ›ÎŸÎšÎ›Î—Î¡Î©Î˜Î—ÎšÎ•!'))
        
        self.stdout.write('\nğŸ“Š Î‘Î ÎŸÎ¤Î•Î›Î•Î£ÎœÎ‘Î¤Î‘:')
        self.stdout.write('â”€' * 70)
        self.stdout.write(f'  ğŸ“ Î£ÏÎ½Î¿Î»Î¿ Î³ÏÎ±Î¼Î¼ÏÎ½:        {stats["total_rows"]}')
        self.stdout.write(f'  âœ… ÎÎ­Î¿Î¹ Ï€ÎµÎ»Î¬Ï„ÎµÏ‚:          {stats["created"]}')
        self.stdout.write(f'  ğŸ”„ Î•Î½Î·Î¼ÎµÏÏ‰Î¼Î­Î½Î¿Î¹:          {stats["updated"]}')
        self.stdout.write(f'  â­ï¸  Î Î±ÏÎ±Î»ÎµÎ¯Ï†Î¸Î·ÎºÎ±Î½:        {stats["skipped"]}')
        self.stdout.write(f'  âŒ Î£Ï†Î¬Î»Î¼Î±Ï„Î±:              {stats["errors"]}')
        
        success_rate = 0
        if stats['total_rows'] > 0:
            successful = stats['created'] + stats['updated']
            success_rate = (successful / stats['total_rows']) * 100
        
        self.stdout.write(f'  ğŸ“ˆ Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±:              {success_rate:.1f}%')
        
        # Errors list
        if errors:
            self.stdout.write('\nâŒ Î›Î™Î£Î¤Î‘ Î£Î¦Î‘Î›ÎœÎ‘Î¤Î©Î:')
            self.stdout.write('â”€' * 70)
            for error in errors[:20]:  # Show max 20 errors
                self.stdout.write(f'  â€¢ {error}')
            
            if len(errors) > 20:
                self.stdout.write(f'  ... ÎºÎ±Î¹ {len(errors) - 20} Î±ÎºÏŒÎ¼Î± ÏƒÏ†Î¬Î»Î¼Î±Ï„Î±')
        
        self.stdout.write('\n' + '='*70)
        
        if dry_run:
            self.stdout.write(self.style.NOTICE('\nğŸ’¡ Î‘Ï…Ï„ÏŒ Î®Ï„Î±Î½ DRY RUN. Î¤ÏÎ­Î¾Îµ Ï‡Ï‰ÏÎ¯Ï‚ --dry-run Î³Î¹Î± Ï€ÏÎ±Î³Î¼Î±Ï„Î¹ÎºÏŒ import.\n'))
        elif stats['errors'] == 0:
            self.stdout.write(self.style.SUCCESS('\nğŸ‰ Î¤Î¿ import Î¿Î»Î¿ÎºÎ»Î·ÏÏÎ¸Î·ÎºÎµ Ï‡Ï‰ÏÎ¯Ï‚ ÏƒÏ†Î¬Î»Î¼Î±Ï„Î±!\n'))
        else:
            self.stdout.write(self.style.WARNING(f'\nâš ï¸  Î¤Î¿ import Î¿Î»Î¿ÎºÎ»Î·ÏÏÎ¸Î·ÎºÎµ Î¼Îµ {stats["errors"]} ÏƒÏ†Î¬Î»Î¼Î±Ï„Î±.\n'))