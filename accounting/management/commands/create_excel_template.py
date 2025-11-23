from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.styles import Font, PatternFill


class Command(BaseCommand):
    help = 'Δημιουργία Excel template για import πελατών'

    def add_arguments(self, parser):
        parser.add_argument('output_file', type=str, help='Όνομα αρχείου εξόδου (π.χ. template.xlsx)')

    def handle(self, *args, **kwargs):
        output_file = kwargs['output_file']
        
        # Δημιουργία workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Πελάτες"
        
        # Headers
        headers = [
            'Α.Φ.Μ.',
            'Επωνυμία/Επώνυμο',
            'Όνομα',
            'Όνομα Πατρός',
            'Αριθμός Ταυτότητας',
            'Είδος Ταυτότητας',
            'Προσωπικός Αριθμός',
            'Α.Μ.Κ.Α.',
            'Α.Μ. Ι.Κ.Α.',
            'Αριθμός Γ.Ε.ΜΗ.',
            'Αριθμός Δ.ΥΠ.Α',
            'Ημ. Γένησης',
            'Ημ. Γάμου',
            'Φύλο',
            'Διεύθυνση Κατοικίας',
            'Αριθμός',
            'Πόλη Κατοικίας',
            'Δήμος Κατοικίας',
            'Νομός Κατοικίας',
            'T.K. Κατοικίας',
            'Τηλέφωνο Οικίας 1',
            'Τηλέφωνο Οικίας 2',
            'Κινητό τηλέφωνο',
            'Διεύθυνση Επιχείρησης',
            'Αριθμός Επιχείρησης',
            'Πόλη Επιχείρησης',
            'Δήμος Επιχείρησης',
            'Νομός Επιχείρησης',
            'Τ.Κ. Επιχείρησης',
            'Τηλέφωνο Επιχείρησης 1',
            'Τηλέφωνο Επιχείρησης 2',
            'Email',
            'Τράπεζα',
            'IBAN',
            'Είδος Υπόχρεου',
            'Κατηγορία Βιβλίων',
            'Νομική Μορφή',
            'Αγρότης',
            'Ημ/νία Έναρξης Εργασιών',
            'Όνομα Χρήστη Taxis Net',
            'Κωδικός Taxis Net',
            'Όνομα Χρήστη Ι.Κ.Α. Εργοδότη',
            'Κωδικός Ι.Κ.Α. Εργοδότη',
            'Όνομα Χρήστη Γ.Ε.ΜΗ.',
            'Κωδικός Γ.Ε.ΜΗ.',
            'Α.Φ.Μ Συζύγου/Μ.Σ.Σ.',
            'Α.Φ.Μ. Φορέας',
            'ΑΜ ΚΛΕΙΔΙ',
        ]
        
        # Styling για headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Παραδείγματα (3 γραμμές)
        examples = [
            {
                'Α.Φ.Μ.': '123456789',
                'Επωνυμία/Επώνυμο': 'ΠΑΠΑΔΟΠΟΥΛΟΣ',
                'Όνομα': 'ΓΙΩΡΓΟΣ',
                'Όνομα Πατρός': 'ΔΗΜΗΤΡΙΟΣ',
                'Φύλο': 'Μ',
                'Κινητό τηλέφωνο': '6912345678',
                'Email': 'gpapadopoulos@example.com',
                'Πόλη Επιχείρησης': 'ΑΘΗΝΑ',
                'Είδος Υπόχρεου': 'ΕΠΑΓΓΕΛΜΑΤΙΑΣ',
                'Κατηγορία Βιβλίων': 'Β',
            },
            {
                'Α.Φ.Μ.': '987654321',
                'Επωνυμία/Επώνυμο': 'ΚΑΦΕ ΤΟΥ ΚΩΣΤΑ ΟΕ',
                'Όνομα': '',
                'Κινητό τηλέφωνο': '6987654321',
                'Email': 'info@kafetou.gr',
                'Πόλη Επιχείρησης': 'ΘΕΣΣΑΛΟΝΙΚΗ',
                'Είδος Υπόχρεου': 'ΕΤΑΙΡΕΙΑ',
                'Κατηγορία Βιβλίων': 'Γ',
                'Νομική Μορφή': 'ΟΕ',
            },
            {
                'Α.Φ.Μ.': '111222333',
                'Επωνυμία/Επώνυμο': 'ΝΙΚΟΛΑΟΥ',
                'Όνομα': 'ΜΑΡΙΑ',
                'Φύλο': 'Γ',
                'Κινητό τηλέφωνο': '6971234567',
                'Email': 'maria.nikolaou@gmail.com',
                'Πόλη Κατοικίας': 'ΤΡΙΚΑΛΑ',
                'Είδος Υπόχρεου': 'ΙΔΙΩΤΗΣ',
                'Κατηγορία Βιβλίων': 'ΧΩΡΙΣ',
            },
        ]
        
        # Γράψιμο παραδειγμάτων
        for row_idx, example in enumerate(examples, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = example.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Δημιουργία sheet με οδηγίες
        ws_instructions = wb.create_sheet("Οδηγίες")
        instructions = [
            ["ΟΔΗΓΙΕΣ ΣΥΜΠΛΗΡΩΣΗΣ EXCEL"],
            [""],
            ["ΥΠΟΧΡΕΩΤΙΚΑ ΠΕΔΙΑ:"],
            ["• Α.Φ.Μ. - Μοναδικό ανά πελάτη"],
            ["• Επωνυμία/Επώνυμο - Επωνυμία εταιρείας ή επώνυμο φυσικού προσώπου"],
            ["• Είδος Υπόχρεου - Επιλογές: ΙΔΙΩΤΗΣ, ΕΠΑΓΓΕΛΜΑΤΙΑΣ, ΕΤΑΙΡΕΙΑ"],
            [""],
            ["ΠΡΟΑΙΡΕΤΙΚΑ ΠΕΔΙΑ:"],
            ["• Όλα τα υπόλοιπα πεδία μπορούν να μείνουν κενά"],
            ["• Μπορείτε να τα συμπληρώσετε αργότερα από το σύστημα"],
            [""],
            ["ΕΙΔΙΚΕΣ ΤΙΜΕΣ:"],
            ["• Φύλο: Μ ή Γ"],
            ["• Κατηγορία Βιβλίων: Α, Β, Γ, ή ΧΩΡΙΣ"],
            ["• Αγρότης: ΝΑΙ ή κενό"],
            ["• Ημερομηνίες: Σε μορφή DD/MM/YYYY"],
            [""],
            ["ΣΗΜΕΙΩΣΕΙΣ:"],
            ["• Το σύστημα θα αγνοήσει γραμμές χωρίς ΑΦΜ"],
            ["• Αν υπάρχει ήδη ΑΦΜ, θα ενημερωθούν τα στοιχεία"],
            ["• Μην αλλάξετε τα ονόματα των στηλών!"],
        ]
        
        for row_idx, instruction in enumerate(instructions, start=1):
            ws_instructions.cell(row=row_idx, column=1, value=instruction[0])
            if row_idx == 1:
                ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        
        # Αποθήκευση
        wb.save(output_file)
        self.stdout.write(self.style.SUCCESS(f'✅ Template δημιουργήθηκε: {output_file}'))
        self.stdout.write(f'Περιέχει:')
        self.stdout.write(f'  • {len(headers)} στήλες')
        self.stdout.write(f'  • 3 παραδείγματα')
        self.stdout.write(f'  • Sheet με οδηγίες')