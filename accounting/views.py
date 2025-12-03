"""
Accounting Views - Complete Professional Implementation
Author: ddiplas
Version: 2.0
Description: Comprehensive views for accounting management system with advanced features
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count, Q, Sum, Avg
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import cache_page
from django.middleware.csrf import get_token
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.mail import send_mail
from django.conf import settings
from django.views.generic import ListView
from django.views.decorators.http import require_GET

from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response 

from .models import (
    ClientProfile, ClientObligation, MonthlyObligation,
    ObligationType, VoIPCall, VoIPCallLog, EmailTemplate, EmailLog,
    Ticket, ClientDocument, ScheduledEmail
)
from django.utils.html import strip_tags 
from .serializers import VoIPCallSerializer, ClientDocumentSerializer, VoIPCallLogSerializer 

from datetime import timedelta, datetime
from collections import defaultdict
import logging
import json
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv

# SECURITY: File upload validation
from common.utils.file_validation import validate_file_upload
from django.core.exceptions import ValidationError as DjangoValidationError

# ============================================
# LOGGER CONFIGURATION
# ============================================
logger = logging.getLogger(__name__)

# ============================================
# MAIN DASHBOARD
# ============================================

@staff_member_required
@login_required
def dashboard_view(request):
    """
    Enhanced Dashboard with comprehensive filtering and statistics
    Features: Advanced filters, real-time stats, bulk operations support
    """
    now = timezone.now()
    
    # ========== FILTER PARAMETERS ==========
    filter_params = {
        'status': request.GET.get('status', ''),
        'client': request.GET.get('client', ''),
        'type': request.GET.get('type', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'sort_by': request.GET.get('sort', 'deadline')
    }
    
    # Convert IDs to integers safely
    filter_client_id = _safe_int(filter_params['client'])
    filter_type_id = _safe_int(filter_params['type'])
    
    logger.info(f"Dashboard accessed by {request.user.username} with filters: {filter_params}")
    
    # ========== STATISTICS (UNFILTERED) ==========
    stats = _calculate_dashboard_stats()
    
    # ========== BUILD FILTERED QUERY ==========
    upcoming_query = _build_filtered_query(
        filter_params, filter_client_id, filter_type_id, now
    )
    
    # ========== APPLY SORTING ==========
    sort_mapping = {
        'deadline': 'deadline',
        'client': 'client__eponimia',
        'type': 'obligation_type__name',
        'status': 'status'
    }
    upcoming_query = upcoming_query.order_by(
        sort_mapping.get(filter_params['sort_by'], 'deadline')
    )
    
    # Get results with limit for performance
    upcoming = list(upcoming_query[:100])
    upcoming_count = upcoming_query.count()
    
    # ========== OVERDUE OBLIGATIONS ==========
    overdue_query = MonthlyObligation.objects.filter(
        deadline__lt=now.date(),
        status__in=['pending', 'overdue']
    ).select_related('client', 'obligation_type')
    
    # Apply client/type filters to overdue as well
    if filter_client_id:
        overdue_query = overdue_query.filter(client_id=filter_client_id)
    if filter_type_id:
        overdue_query = overdue_query.filter(obligation_type_id=filter_type_id)
    
    overdue_obligations = list(overdue_query.order_by('deadline')[:20])
    overdue_count = overdue_query.count()
    
    # ========== FILTER OPTIONS ==========
    all_clients = ClientProfile.objects.all().order_by('eponimia').values('id', 'eponimia', 'afm')
    all_types = ObligationType.objects.filter(is_active=True).order_by('name')
    
    # ========== PREPARE CONTEXT ==========
    context = {
        'title': 'Dashboard - Επισκόπηση',
        
        # Statistics
        'total_clients': stats['total_clients'],
        'pending': stats['pending'],
        'completed': stats['completed'],
        'overdue': stats['overdue'],
        
        # Main data
        'upcoming': upcoming,
        'upcoming_count': upcoming_count,
        'overdue_obligations': overdue_obligations,
        'overdue_count': overdue_count,
        
        # Filter options
        'all_clients': all_clients,
        'all_types': all_types,
        
        # Current filter values
        **{f'filter_{k}': v for k, v in filter_params.items()},
        
        # Additional metadata
        'current_date': now.date(),
        'user': request.user,
    }
    
    return render(request, 'accounting/dashboard.html', context)


# ============================================
# CLIENT DETAIL VIEW
# ============================================

@staff_member_required
def client_detail_view(request, client_id):
    """
    Comprehensive client view with all obligations and analytics
    """
    client = get_object_or_404(ClientProfile, id=client_id)
    now = timezone.now()
    
    # All obligations for this client
    all_obligations = MonthlyObligation.objects.filter(
        client=client
    ).select_related('obligation_type').order_by('-deadline')
    
    # Calculate statistics
    stats = {
        'total': all_obligations.count(),
        'pending': all_obligations.filter(status='pending').count(),
        'completed': all_obligations.filter(status='completed').count(),
        'overdue': all_obligations.filter(status='overdue').count(),
    }
    
    # Upcoming obligations (next 30 days)
    next_month = now.date() + timedelta(days=30)
    upcoming = all_obligations.filter(
        deadline__range=[now.date(), next_month],
        status='pending'
    ).order_by('deadline')[:10]
    
    # Overdue obligations
    overdue = all_obligations.filter(
        deadline__lt=now.date(),
        status__in=['pending', 'overdue']
    ).order_by('deadline')
    
    # Recent completed (last 30 days)
    past_month = now.date() - timedelta(days=30)
    recent_completed = all_obligations.filter(
        status='completed',
        completed_date__gte=past_month
    ).order_by('-completed_date')[:10]
    
    # Monthly breakdown (last 6 months)
    monthly_stats = _calculate_monthly_stats(all_obligations, now)
    
    # Active obligation types
    active_types = ClientObligation.objects.filter(
        client=client,
        is_active=True
    ).select_related('obligation_type')
    
    # Client documents
    documents = ClientDocument.objects.filter(
        client=client
    ).order_by('-uploaded_at')
    
    context = {
        'title': f'{client.eponimia} - Λεπτομέρειες',
        'client': client,
        'stats': stats,
        'upcoming': upcoming,
        'overdue': overdue,
        'recent_completed': recent_completed,
        'monthly_stats': monthly_stats,
        'active_types': active_types,
        'documents': documents,  # <-- ΠΡΟΣΤΕΘΗΚΕ
    }
    
    return render(request, 'accounting/client_detail.html', context)

# ============================================
# QUICK COMPLETE WITH FILE UPLOAD
# ============================================

@require_POST
@staff_member_required
def quick_complete_obligation(request, obligation_id):
    """
    Quick complete single obligation with optional file attachment
    """
    try:
        obligation = MonthlyObligation.objects.get(id=obligation_id)
        
        # Handle both FormData and JSON requests
        if 'multipart/form-data' in request.META.get('CONTENT_TYPE', ''):
            # Form data with file upload
            time_spent = request.POST.get('time_spent', 0)
            notes = request.POST.get('notes', '')
            attachment = request.FILES.get('attachment')
        else:
            # JSON data (no file)
            data = json.loads(request.body)
            time_spent = data.get('time_spent', 0)
            notes = data.get('notes', '')
            attachment = None
        
        # Update obligation
        obligation.status = 'completed'
        obligation.completed_date = timezone.now().date()
        obligation.completed_by = request.user
        
        # Handle time spent
        if time_spent:
            try:
                obligation.time_spent = float(time_spent)
            except (ValueError, TypeError):
                pass
        
        # Handle notes with timestamp
        if notes:
            timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
            new_note = f"[{timestamp}] {notes}"
            if obligation.notes:
                obligation.notes += f"\n{new_note}"
            else:
                obligation.notes = new_note
        
        # Handle file attachment with SECURITY validation
        if 'attachment' in request.FILES:
            uploaded_file = request.FILES['attachment']

            # SECURITY FIX: Validate uploaded file before processing
            try:
                validate_file_upload(uploaded_file)
            except DjangoValidationError as e:
                return JsonResponse({
                    'success': False,
                    'message': f'❌ Μη έγκυρο αρχείο: {str(e)}'
                }, status=400)

            # File is validated, proceed with archiving
            archive_path = obligation.archive_attachment(uploaded_file)
            logger.info(f'✅ Αρχειοθετήθηκε: {archive_path}')
        obligation.save()
        
        # Success response
        message = f'✅ {obligation.obligation_type.name} ολοκληρώθηκε!'
        if time_spent:
            message += f' (⏱️ {time_spent}h)'
        if attachment:
            message += ' 📎'
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except MonthlyObligation.DoesNotExist:
        return JsonResponse(
            {'success': False, 'message': '❌ Υποχρέωση δεν βρέθηκε'}, 
            status=404
        )
    except Exception as e:
        logger.error(f"Error in quick_complete: {str(e)}", exc_info=True)
        return JsonResponse(
            {'success': False, 'message': f'❌ Σφάλμα: {str(e)}'}, 
            status=500
        )

# accounting/models.py - ΜΕΣΑ στην class MonthlyObligation

def archive_attachment(self, uploaded_file, subfolder=None):
    """Αρχειοθέτηση αρχείου με σωστή δομή φακέλων"""
    import os
    from django.conf import settings
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    
    # Get or create configuration
    config, created = ArchiveConfiguration.objects.get_or_create(
        obligation_type=self.obligation_type,
        defaults={
            'filename_pattern': '{type_code}_{month}_{year}.pdf',
            'folder_pattern': 'clients/{client_afm}_{client_name}/{year}/{month}/',
        }
    )
    
    # Get the correct archive path
    archive_path = config.get_archive_path(self, uploaded_file.name)

    logger.debug(f"[ARCHIVE] Will save to: {archive_path}")

    # Delete old file if exists
    if self.attachment:
        try:
            if default_storage.exists(self.attachment.name):
                default_storage.delete(self.attachment.name)
                logger.debug(f"[ARCHIVE] Deleted old file: {self.attachment.name}")
        except Exception as e:
            logger.debug(f"[ARCHIVE] Could not delete old file: {e}")
    
    # Read file content
    if hasattr(uploaded_file, 'read'):
        file_content = uploaded_file.read()
        uploaded_file.seek(0)  # Reset pointer
    else:
        file_content = uploaded_file
    
    # Save to storage with correct path
    saved_path = default_storage.save(archive_path, ContentFile(file_content))
    logger.debug(f"[ARCHIVE] Saved to: {saved_path}")
    
    # Update model field
    self.attachment.name = saved_path
    
    # Update attachments list if multiple files enabled
    if config.allow_multiple_files:
        current = self.attachments or []
        if saved_path not in current:
            current.append(saved_path)
        self.attachments = current
    
    # Save the model
    self.save(update_fields=['attachment', 'attachments'])
    
    return saved_path
# ============================================
# BULK COMPLETE (SIMPLE VERSION)
# ============================================

@require_POST
@staff_member_required
def bulk_complete_view(request):
    """
    Simple bulk complete - all obligations get same treatment
    """
    try:
        obligation_ids = json.loads(request.POST.get('obligation_ids', '[]'))
        time_spent = request.POST.get('time_spent', '0')
        notes = request.POST.get('notes', '')
        attachments = request.FILES.getlist('attachments')

        # SECURITY FIX: Validate all uploaded files before processing
        for attachment in attachments:
            try:
                validate_file_upload(attachment)
            except DjangoValidationError as e:
                return JsonResponse({
                    'success': False,
                    'message': f'❌ Μη έγκυρο αρχείο "{attachment.name}": {str(e)}'
                }, status=400)

        if not obligation_ids:
            return JsonResponse({
                'success': False,
                'message': '❌ Δεν επιλέχθηκαν υποχρεώσεις'
            })

        obligations = MonthlyObligation.objects.filter(id__in=obligation_ids)
        completed_count = 0

        for idx, obl in enumerate(obligations):
            obl.status = 'completed'
            obl.completed_date = timezone.now().date()
            obl.completed_by = request.user

            if time_spent:
                obl.time_spent = float(time_spent)

            if notes:
                timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
                new_note = f"[{timestamp}] [BULK] {notes}"
                if obl.notes:
                    obl.notes += f"\n{new_note}"
                else:
                    obl.notes = new_note

            # Attach file if available (already validated above)
            if idx < len(attachments):
                obl.attachment = attachments[idx]

            obl.save()
            completed_count += 1

        return JsonResponse({
            'success': True,
            'completed_count': completed_count,
            'message': f'✅ Ολοκληρώθηκαν {completed_count} υποχρεώσεις!'
        })

    except Exception as e:
        logger.error(f"Error in bulk_complete: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'❌ Σφάλμα: {str(e)}'
        }, status=500)


# ============================================
# ADVANCED BULK COMPLETE WITH GROUPING
# ============================================

# accounting/views.py - Αντικατέστησε το advanced_bulk_complete με αυτό

@require_POST
@staff_member_required
def advanced_bulk_complete(request):
    """Advanced bulk complete με πλήρες debugging"""
    
    logger.info("="*50)
    logger.info("ADVANCED BULK COMPLETE - START")
    logger.info(f"Method: {request.method}")
    logger.info(f"POST data keys: {request.POST.keys()}")
    logger.info(f"FILES keys: {request.FILES.keys()}")
    
    try:
        # Get completion data
        completion_data_raw = request.POST.get('completion_data', '[]')
        logger.info(f"Raw completion_data: {completion_data_raw[:200]}...")  # First 200 chars
        
        completion_data = json.loads(completion_data_raw)
        notes = request.POST.get('notes', '')
        
        logger.info(f"Parsed completion_data: {len(completion_data)} groups")
        
        if not completion_data:
            logger.warning("No completion data found!")
            return JsonResponse({
                'success': False,
                'message': '❌ Δεν υπάρχουν δεδομένα ολοκλήρωσης',
                'completed_count': 0
            })
        
        completed_count = 0
        errors = []
        processed_details = []
        
        # Process each group
        for i, group_data in enumerate(completion_data):
            logger.info(f"\n--- Processing group {i+1}/{len(completion_data)} ---")
            
            client_afm = group_data.get('client_afm', 'NO_AFM')
            group_num = group_data.get('group', '0')
            obligation_ids = group_data.get('obligations', [])
            
            logger.info(f"AFM: {client_afm}, Group: {group_num}, Obligations: {obligation_ids}")
            
            # Get files for this group
            files_key = f"file_{client_afm}_{group_num}"
            files = request.FILES.getlist(files_key)
            
            logger.info(f"Looking for files with key: {files_key}")
            logger.info(f"Found {len(files)} files")
            
            if files:
                for f in files:
                    logger.info(f"  - File: {f.name} ({f.size} bytes)")
            
            # Process obligations
            for j, obl_id in enumerate(obligation_ids):
                try:
                    logger.info(f"  Processing obligation {obl_id}...")
                    
                    obligation = MonthlyObligation.objects.get(id=obl_id)
                    
                    # Update status
                    obligation.status = 'completed'
                    obligation.completed_date = timezone.now().date()
                    obligation.completed_by = request.user
                    
                    # Add notes
                    if notes:
                        timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
                        new_note = f"[{timestamp}] {notes}"
                        if obligation.notes:
                            obligation.notes += f"\n{new_note}"
                        else:
                            obligation.notes = new_note
                    
                    # Handle file
                    if group_num == '0':  # Individual files
                        if j < len(files):
                            logger.info(f"    Archiving individual file {j}: {files[j].name}")
                            archive_path = obligation.archive_attachment(files[j])
                            processed_details.append(f"✅ {obligation.obligation_type.name}: {archive_path}")
                            logger.info(f"    Archived to: {archive_path}")
                        else:
                            obligation.save()
                            processed_details.append(f"✅ {obligation.obligation_type.name} (χωρίς αρχείο)")
                            logger.info(f"    No file for this obligation")
                    else:  # Group file
                        if files:
                            file_to_use = files[0]
                            logger.info(f"    Using group file: {file_to_use.name}")
                            
                            # Create copy for each obligation
                            from django.core.files.base import ContentFile
                            file_content = file_to_use.read()
                            file_copy = ContentFile(file_content)
                            file_copy.name = file_to_use.name
                            file_to_use.seek(0)  # Reset for next use
                            
                            archive_path = obligation.archive_attachment(file_copy)
                            processed_details.append(f"✅ {obligation.obligation_type.name} (Group {group_num})")
                            logger.info(f"    Archived to: {archive_path}")
                        else:
                            obligation.save()
                            processed_details.append(f"✅ {obligation.obligation_type.name} (no file)")
                    
                    completed_count += 1
                    logger.info(f"    ✅ Completed successfully")
                    
                except MonthlyObligation.DoesNotExist:
                    error_msg = f"Obligation {obl_id} not found"
                    logger.error(f"    ❌ {error_msg}")
                    errors.append(error_msg)
                except Exception as e:
                    error_msg = f"Error with {obl_id}: {str(e)}"
                    logger.error(f"    ❌ {error_msg}", exc_info=True)
                    errors.append(error_msg)
        
        # Final summary
        logger.info(f"\n=== SUMMARY ===")
        logger.info(f"Completed: {completed_count}")
        logger.info(f"Errors: {len(errors)}")
        
        if completed_count > 0:
            message = f'✅ Ολοκληρώθηκαν {completed_count} υποχρεώσεις!'
            if errors:
                message += f' (⚠️ {len(errors)} σφάλματα)'
            success = True
        else:
            message = '❌ Καμία υποχρέωση δεν ολοκληρώθηκε'
            success = False
        
        response_data = {
            'success': success,
            'completed_count': completed_count,
            'message': message,
            'errors': errors[:5],
            'details': processed_details[:10]
        }
        
        logger.info(f"Response: {response_data}")
        logger.info("="*50)
        
        return JsonResponse(response_data)
        
    except json.JSONDecodeError as e:
        logger.error(f"JSON Decode Error: {e}")
        logger.error(f"Raw data was: {request.POST.get('completion_data', '')[:500]}")
        return JsonResponse({
            'success': False,
            'message': f'❌ JSON Error: {str(e)}',
            'completed_count': 0
        })
        
    except Exception as e:
        logger.error(f"Critical error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'❌ Κρίσιμο σφάλμα: {str(e)}',
            'completed_count': 0
        })

# Αφαίρεσε τις παλιές _process_individual_obligations και _process_grouped_obligations
# γιατί τώρα το χειρίζεται όλο μέσα στο advanced_bulk_complete

# ============================================
# EXPORT TO EXCEL
# ============================================

@staff_member_required
def export_filtered_excel(request):
    """
    Export filtered obligations to Excel with formatting
    """
    try:
        # Get filters from request
        filters = _get_filters_from_request(request)
        
        # Build query
        query = _build_export_query(filters)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Υποχρεώσεις"
        
        # Apply styling
        _apply_excel_styling(ws)
        
        # Add headers and data
        _write_excel_headers(ws)
        _write_excel_data(ws, query)
        
        # Auto-adjust columns
        _auto_adjust_excel_columns(ws)
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Ypohrewseis_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        logger.info(f"Excel exported by {request.user.username}: {query.count()} records")
        return response
        
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


# ============================================
# REPORTS & ANALYTICS
# ============================================

@staff_member_required
def reports_view(request):
    """
    Comprehensive analytics dashboard with charts and statistics
    """
    now = timezone.now()
    months_back = int(request.GET.get('months', 6))
    start_date = (now - timedelta(days=30*months_back)).date()
    
    # Monthly completion statistics
    monthly_stats = _calculate_monthly_completion_stats(start_date)
    
    # Client performance metrics
    client_stats = _calculate_client_performance(start_date)
    
    # Time tracking summary
    time_stats = _calculate_time_stats(start_date)
    
    # Revenue calculations
    revenue_data = _calculate_revenue(start_date)
    
    # Obligation type statistics
    type_stats = _calculate_type_stats(start_date)
    
    # Current month summary
    current_stats = _calculate_current_month_stats(now)

    # Format data for charts
    chart_data = _format_chart_data(monthly_stats)

    # Get clients for PDF export dropdown
    clients = ClientProfile.objects.filter(is_active=True).order_by('eponimia')

    # Year choices for monthly report
    current_year = now.year
    year_choices = list(range(current_year - 2, current_year + 2))

    context = {
        'title': 'Reports & Analytics',
        'months_back': months_back,
        **chart_data,
        'client_stats': client_stats,
        'time_stats': time_stats,
        'total_revenue': revenue_data['total'],
        'type_stats': type_stats,
        'current_stats': current_stats,
        # PDF Export context
        'clients': clients,
        'current_month': now.month,
        'current_year': current_year,
        'year_choices': year_choices,
    }

    return render(request, 'accounting/reports.html', context)


# ============================================
# VOIP DASHBOARD & MANAGEMENT
# ============================================

@staff_member_required
def voip_dashboard(request):
    """
    Modern VoIP Dashboard with real-time updates
    """
    # Get recent calls with optimized query
    calls = VoIPCall.objects.select_related('client').order_by('-started_at')[:50]
    
    # Calculate statistics
    today = timezone.now().date()
    stats = VoIPCall.objects.aggregate(
        total=Count('id'),
        missed=Count('id', filter=Q(status='missed')),
        completed=Count('id', filter=Q(status='completed')),
        today_total=Count('id', filter=Q(started_at__date=today)),
        pending_followup=Count('id', filter=Q(resolution='follow_up')),
    )
    
    # Calculate success rate
    stats['success_rate'] = _calculate_success_rate(
        stats['completed'], 
        stats['total']
    )
    
    context = {
        'calls': calls,
        'stats': stats,
        'csrf_token': get_token(request),
    }
    
    logger.info(f"VoIP Dashboard accessed by {request.user.username}")
    return render(request, 'admin/voip_dashboard.html', context)


@staff_member_required
@require_http_methods(["GET"])
@cache_page(5)  # Cache for 5 seconds
def voip_calls_api(request):
    """
    Real-time API for VoIP calls with AJAX support
    """
    try:
        # Get recent calls
        calls = VoIPCall.objects.select_related('client').order_by('-started_at')[:30]
        
        # Sort by priority (missed first)
        calls = sorted(calls, key=lambda x: (
            x.status != 'missed', 
            -x.started_at.timestamp()
        ))
        
        # Format data for JSON
        data = [_format_voip_call(call) for call in calls]
        
        return JsonResponse({
            'success': True,
            'calls': data,
            'timestamp': timezone.now().isoformat(),
            'total_missed': VoIPCall.objects.filter(status='missed').count(),
        })
        
    except Exception as e:
        logger.error(f"Error in voip_calls_api: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@staff_member_required
@require_POST
def voip_call_update(request, call_id):
    """Update VoIP call via AJAX"""
    try:
        call = VoIPCall.objects.select_related('client').get(id=call_id)
    except VoIPCall.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '❌ Κλήση δεν βρέθηκε!'
        }, status=404)

    # Parse JSON body
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': '❌ Invalid JSON data'
        }, status=400)

    old_values = {
        'notes': call.notes,
        'resolution': call.resolution,
        'status': call.status,
    }

    # Update fields safely
    if 'notes' in data:
        call.notes = data['notes'][:1000]

    if 'resolution' in data and data['resolution'] in ['pending', 'closed', 'follow_up', '']:
        call.resolution = data['resolution']

    old_status = call.status
    if 'status' in data and data['status'] in ['active', 'completed', 'missed', 'failed']:
        call.status = data['status']

    call.save()
    _log_voip_change(call, old_values, request.user)

    # Optional: trigger ticket when completed
    if old_status != 'completed' and call.status == 'completed':
        try:
            from accounting.tasks import trigger_answered_call_ticket
            trigger_answered_call_ticket(call, request.user)
            logger.info(f"Triggered answered call ticket for call #{call_id}")
        except Exception as e:
            logger.warning(f"Could not trigger answered call ticket: {e}")

    return JsonResponse({
        'success': True,
        'message': f'✅ Κλήση {call.phone_number} ενημερώθηκε!',
        'updated_call': _format_voip_call(call)
    })
# ============================================
# VOIP API VIEWSETS (REST Framework)
# ============================================

class VoIPCallViewSet(viewsets.ModelViewSet):
    """
    REST API ViewSet for VoIP calls
    """
    queryset = VoIPCall.objects.all()
    serializer_class = VoIPCallSerializer
    permission_classes = [permissions.IsAdminUser]  # SECURITY: Restrict to admin users only
    filterset_fields = ['direction', 'status', 'client', 'phone_number']
    search_fields = ['phone_number', 'client__eponimia', 'notes']
    ordering_fields = ['started_at', 'duration_seconds']
    ordering = ['-started_at']
    
    def perform_create(self, serializer):
        """Create call and auto-match client"""
        voip_call = serializer.save()
        client = self._match_client_by_phone(voip_call.phone_number)
        
        if client:
            voip_call.client = client
            voip_call.client_email = client.email
            voip_call.save()
            logger.info(f"Matched call to client: {client.eponimia}")
        
        # Log call creation
        VoIPCallLog.objects.create(
            call=voip_call,
            action='started',
            description=f'Call started from {voip_call.phone_number}'
        )
    
    def _match_client_by_phone(self, phone_number):
        """Match phone number to client"""
        clean_number = phone_number.replace(" ", "").replace("-", "")
        
        return ClientProfile.objects.filter(
            Q(tilefono_oikias_1__icontains=clean_number) |
            Q(tilefono_oikias_2__icontains=clean_number) |
            Q(kinito_tilefono__icontains=clean_number) |
            Q(tilefono_epixeirisis_1__icontains=clean_number) |
            Q(tilefono_epixeirisis_2__icontains=clean_number)
        ).first()
    
    @action(detail=True, methods=['post'])
    def end_call(self, request, pk=None):
        """End a call and process follow-up actions"""
        voip_call = self.get_object()
        ended_at = request.data.get('ended_at')
        
        if not ended_at:
            return Response(
                {'error': 'ended_at is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Parse and set end time
            from dateutil import parser as dateutil_parser
            if isinstance(ended_at, str):
                ended_at = dateutil_parser.parse(ended_at)
            
            if ended_at.tzinfo is None:
                ended_at = timezone.make_aware(ended_at)
            
            voip_call.ended_at = ended_at
            voip_call.status = 'completed' if voip_call.duration_seconds > 10 else 'missed'
            voip_call.save()
            
            # Log call end
            VoIPCallLog.objects.create(
                call=voip_call,
                action='ended',
                description=f'Call ended - Duration: {voip_call.duration_formatted}'
            )
            
            # Handle missed call actions
            if voip_call.status == 'missed':
                self._handle_missed_call(voip_call)
            
            return Response(
                VoIPCallSerializer(voip_call).data, 
                status=status.HTTP_200_OK
            )
            
        except Exception as e:
            logger.error(f"Error ending call: {e}", exc_info=True)
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _handle_missed_call(self, voip_call):
        """Process missed call - create ticket and send email"""
        try:
            voip_call.ticket_created = True
            voip_call.save()
            
            VoIPCallLog.objects.create(
                call=voip_call,
                action='ticket_created',
                description=f'Missed call ticket for {voip_call.phone_number}'
            )
            
            if voip_call.client_email:
                self._send_missed_call_email(voip_call)
                
        except Exception as e:
            logger.error(f"Error handling missed call: {e}")
    
    def _send_missed_call_email(self, voip_call):
        """Send email notification for missed call"""
        try:
            subject = f"Αναπάντητη κλήση - {voip_call.phone_number}"
            message = f"""
Καλησπέρα,

Είχατε μια αναπάντητη κλήση:

📞 Αριθμός: {voip_call.phone_number}
📅 Ώρα: {voip_call.started_at.strftime('%d/%m/%Y %H:%M')}
↔️ Κατεύθυνση: {voip_call.get_direction_display()}

Παρακαλώ επικοινωνήστε μαζί μας αν χρειάζεστε κάτι.

Ευχαριστούμε,
Λογιστικό Γραφείο
            """
            
            send_mail(
                subject, 
                message,
                settings.DEFAULT_FROM_EMAIL,
                [voip_call.client_email],
                fail_silently=False,
            )
            
            logger.info(f"Email sent to {voip_call.client_email}")
            
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
class VoIPCallsListView(ListView):
    """
    Django Class-Based View for VoIP calls listing
    Alternative to the function-based voip_dashboard
    """
    model = VoIPCall
    template_name = 'accounting/voip_calls_list.html'
    context_object_name = 'calls'
    paginate_by = 50
    
    def get_queryset(self):
        """Get filtered and sorted calls"""
        queryset = VoIPCall.objects.select_related('client').all()
        
        # Apply filters from GET parameters
        status = self.request.GET.get('status')
        direction = self.request.GET.get('direction')
        search = self.request.GET.get('search')
        
        if status:
            queryset = queryset.filter(status=status)
        if direction:
            queryset = queryset.filter(direction=direction)
        if search:
            queryset = queryset.filter(
                Q(phone_number__icontains=search) |
                Q(client__eponimia__icontains=search) |
                Q(notes__icontains=search)
            )
        
        return queryset.order_by('-started_at')
    
    def get_context_data(self, **kwargs):
        """Add statistics to context"""
        context = super().get_context_data(**kwargs)
        
        # Calculate stats
        context['total'] = VoIPCall.objects.count()
        context['missed'] = VoIPCall.objects.filter(status='missed').count()
        context['completed'] = VoIPCall.objects.filter(status='completed').count()
        
        if context['total'] > 0:
            context['success_rate'] = round(
                (context['completed'] / context['total']) * 100
            )
        else:
            context['success_rate'] = 0
        
        # Add filter values
        context['current_status'] = self.request.GET.get('status', '')
        context['current_direction'] = self.request.GET.get('direction', '')
        context['current_search'] = self.request.GET.get('search', '')
        
        return context


class VoIPCallLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only ViewSet for call logs
    """
    queryset = VoIPCallLog.objects.all()
    serializer_class = VoIPCallLogSerializer
    permission_classes = [permissions.IsAdminUser]  # SECURITY: Restrict to admin users only
    filterset_fields = ['call', 'action']
    ordering = ['-created_at']


# ============================================
# EMAIL AUTOMATION API
# ============================================

@require_http_methods(["GET"])
@staff_member_required
def api_email_templates(request):
    """
    API endpoint to get all active email templates
    """
    try:
        templates = EmailTemplate.objects.filter(is_active=True).order_by('name')
        
        result = [
            {
                'id': template.id,
                'name': template.name,
                'description': template.description or '',
                'subject': template.subject,
            }
            for template in templates
        ]
        
        logger.info(f"Email templates API: {len(result)} templates returned")
        return JsonResponse(result, safe=False)
        
    except Exception as e:
        logger.error(f"Error in api_email_templates: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@require_POST
@staff_member_required
def api_send_bulk_email(request):
    """
    Schedule bulk emails for obligations
    """
    try:
        data = json.loads(request.body)
        obligation_ids = data.get('obligation_ids', [])
        template_id = data.get('template_id')
        timing = data.get('timing', 'immediate')
        scheduled_time = data.get('scheduled_time')
        
        # Validate input
        if not obligation_ids or not template_id:
            return JsonResponse({
                'success': False,
                'error': 'Missing obligations or template'
            })
        
        # Get template
        try:
            template = EmailTemplate.objects.get(id=template_id, is_active=True)
        except EmailTemplate.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            })
        
        # Get obligations
        obligations = MonthlyObligation.objects.filter(
            id__in=obligation_ids
        ).select_related('client')
        
        if not obligations.exists():
            return JsonResponse({
                'success': False,
                'error': 'No obligations found'
            })
        
        # Calculate send time
        send_at = _calculate_send_time(timing, scheduled_time)
        
        # Group by client and create scheduled emails
        emails_created = _create_bulk_emails(
            obligations, template, send_at, request.user
        )
        
        return JsonResponse({
            'success': True,
            'emails_created': emails_created,
            'message': f'✅ Προγραμματίστηκαν {emails_created} emails'
        })
        
    except Exception as e:
        logger.error(f"Error in api_send_bulk_email: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


# ============================================
# NOTIFICATION SYSTEM
# ============================================

@staff_member_required
def get_notifications(request):
    """
    Get user notifications for dashboard
    """
    now = timezone.now()
    notifications = []
    
    # Overdue obligations
    overdue = MonthlyObligation.objects.filter(
        deadline__lt=now.date(),
        status__in=['pending', 'overdue']
    ).select_related('client', 'obligation_type').order_by('deadline')[:10]
    
    for obl in overdue:
        days_overdue = (now.date() - obl.deadline).days
        notifications.append({
            'id': obl.id,
            'type': 'overdue',
            'priority': 'high',
            'title': f'Καθυστερημένη: {obl.obligation_type.name}',
            'message': f'{obl.client.eponimia} - {days_overdue} μέρες καθυστέρηση',
            'deadline': obl.deadline.isoformat(),
            'client_id': obl.client.id,
            'icon': '🔴',
        })
    
    # Due today
    today_obligations = MonthlyObligation.objects.filter(
        deadline=now.date(),
        status='pending'
    ).select_related('client', 'obligation_type')
    
    for obl in today_obligations:
        notifications.append({
            'id': obl.id,
            'type': 'due_today',
            'priority': 'medium',
            'title': f'Λήγει Σήμερα: {obl.obligation_type.name}',
            'message': f'{obl.client.eponimia}',
            'deadline': obl.deadline.isoformat(),
            'client_id': obl.client.id,
            'icon': '⚠️',
        })
    
    # Upcoming (next 3 days)
    next_3_days = now.date() + timedelta(days=3)
    upcoming = MonthlyObligation.objects.filter(
        deadline__range=[now.date() + timedelta(days=1), next_3_days],
        status='pending'
    ).select_related('client', 'obligation_type').order_by('deadline')[:5]
    
    for obl in upcoming:
        days_until = (obl.deadline - now.date()).days
        notifications.append({
            'id': obl.id,
            'type': 'upcoming',
            'priority': 'low',
            'title': f'Προσεχώς: {obl.obligation_type.name}',
            'message': f'{obl.client.eponimia} - σε {days_until} μέρες',
            'deadline': obl.deadline.isoformat(),
            'client_id': obl.client.id,
            'icon': '📅',
        })
    
    return JsonResponse({
        'notifications': notifications,
        'count': len(notifications),
        'overdue_count': len([n for n in notifications if n['type'] == 'overdue']),
        'today_count': len([n for n in notifications if n['type'] == 'due_today']),
    })


# ============================================
# HELPER FUNCTIONS (Private)
# ============================================

def _safe_int(value):
    """Safely convert string to integer"""
    try:
        return int(value) if value else None
    except (ValueError, TypeError):
        return None


def _calculate_dashboard_stats():
    """Calculate unfiltered dashboard statistics"""
    return {
        'total_clients': ClientProfile.objects.count(),
        'pending': MonthlyObligation.objects.filter(status='pending').count(),
        'completed': MonthlyObligation.objects.filter(status='completed').count(),
        'overdue': MonthlyObligation.objects.filter(status='overdue').count(),
    }


def _build_filtered_query(filter_params, client_id, type_id, now):
    """Build filtered query for obligations"""
    # Date range
    if filter_params['date_from'] and filter_params['date_to']:
        query = MonthlyObligation.objects.filter(
            deadline__gte=filter_params['date_from'],
            deadline__lte=filter_params['date_to']
        )
    else:
        # Default: past 3 to next 3 months
        past_3_months = now.date() - timedelta(days=90)
        next_3_months = now.date() + timedelta(days=90)
        query = MonthlyObligation.objects.filter(
            deadline__range=[past_3_months, next_3_months]
        )
    
    # Apply filters
    if filter_params['status']:
        query = query.filter(status=filter_params['status'])
    
    if client_id:
        query = query.filter(client_id=client_id)
    
    if type_id:
        query = query.filter(obligation_type_id=type_id)
    
    return query.select_related('client', 'obligation_type')


def _calculate_monthly_stats(obligations, now):
    """Calculate monthly statistics for obligations"""
    stats = []
    for i in range(6):
        month_date = now.date().replace(day=1) - timedelta(days=30*i)
        month_obligations = obligations.filter(
            year=month_date.year,
            month=month_date.month
        )
        stats.append({
            'month': month_date.strftime('%B %Y'),
            'total': month_obligations.count(),
            'completed': month_obligations.filter(status='completed').count(),
            'pending': month_obligations.filter(status='pending').count(),
            'overdue': month_obligations.filter(status='overdue').count(),
        })
    return stats


def _process_individual_obligations(obligation_ids, files, notes, user):
    """Process obligations with individual files"""
    completed = 0
    errors = []
    details = []
    
    for i, obl_id in enumerate(obligation_ids):
        try:
            obligation = MonthlyObligation.objects.get(id=obl_id)
            obligation.status = 'completed'
            obligation.completed_date = timezone.now().date()
            obligation.completed_by = user
            
            # Add notes
            if notes:
                timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
                new_note = f"[{timestamp}] {notes}"
                if obligation.notes:
                    obligation.notes += f"\n{new_note}"
                else:
                    obligation.notes = new_note
            
            # Attach file if available
            if i < len(files):
                obligation.attachment = files[i]
                details.append(f"✅ {obligation.obligation_type.name} με αρχείο {files[i].name}")
            else:
                details.append(f"✅ {obligation.obligation_type.name}")
            
            obligation.save()
            completed += 1
            
        except MonthlyObligation.DoesNotExist:
            errors.append(f"Obligation {obl_id} not found")
        except Exception as e:
            errors.append(f"Error with obligation {obl_id}: {str(e)}")
    
    return {'completed': completed, 'errors': errors, 'details': details}


def _process_grouped_obligations(obligation_ids, files, notes, group_num, user):
    """Process obligations with shared group file"""
    completed = 0
    errors = []
    details = []
    
    file_to_use = files[0] if files else None
    
    for obl_id in obligation_ids:
        try:
            obligation = MonthlyObligation.objects.get(id=obl_id)
            obligation.status = 'completed'
            obligation.completed_date = timezone.now().date()
            obligation.completed_by = user
            
            # Add notes with group indicator
            timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
            group_note = f"[{timestamp}] [Ομάδα {group_num}] {notes}" if notes else f"[{timestamp}] [Ομάδα {group_num}]"
            
            if obligation.notes:
                obligation.notes += f"\n{group_note}"
            else:
                obligation.notes = group_note
            
            # Attach group file
            if file_to_use:
                # Create a copy for each obligation
                file_copy = ContentFile(file_to_use.read())
                file_copy.name = file_to_use.name
                obligation.attachment.save(file_to_use.name, file_copy, save=False)
                file_to_use.seek(0)  # Reset for next use
                details.append(f"✅ {obligation.obligation_type.name} (Ομάδα {group_num})")
            else:
                details.append(f"✅ {obligation.obligation_type.name} (Ομάδα {group_num} - χωρίς αρχείο)")
            
            obligation.save()
            completed += 1
            
        except MonthlyObligation.DoesNotExist:
            errors.append(f"Obligation {obl_id} not found")
        except Exception as e:
            errors.append(f"Error with obligation {obl_id}: {str(e)}")
    
    return {'completed': completed, 'errors': errors, 'details': details}


def _format_voip_call(call):
    """Format VoIP call for JSON response"""
    return {
        'id': call.id,
        'call_id': call.call_id,
        'phone_number': call.phone_number,
        'client_name': call.client.eponimia if call.client else 'Άγνωστος',
        'client_email': call.client_email or '—',
        'direction': call.get_direction_display(),
        'direction_icon': '📲' if call.direction == 'incoming' else '☎️',
        'status': call.get_status_display(),
        'status_value': call.status,
        'status_color': _get_status_color(call.status),
        'resolution': call.get_resolution_display() if call.resolution else '—',
        'resolution_value': call.resolution or '',
        'resolution_color': _get_resolution_color(call.resolution),
        'notes': call.notes[:50] + '...' if len(call.notes or '') > 50 else call.notes or '',
        'duration': call.duration_formatted,
        'started_at': call.started_at.strftime('%d/%m %H:%M:%S'),
        'is_missed': call.status == 'missed',
        'is_recent': (timezone.now() - call.started_at).total_seconds() < 300,
    }


def _get_status_color(status):
    """Get color for status badge"""
    colors = {
        'missed': 'dc2626',    # red
        'completed': '16a34a',  # green
        'active': '2563eb',     # blue
        'failed': 'ea580c',     # orange
    }
    return colors.get(status, '666')


def _get_resolution_color(resolution):
    """Get color for resolution badge"""
    colors = {
        'pending': 'f59e0b',    # amber
        'closed': '10b981',     # green
        'follow_up': '3b82f6',  # blue
        '': 'e5e7eb',          # gray
    }
    return colors.get(resolution, 'e5e7eb')


def _calculate_success_rate(completed, total):
    """Calculate success rate percentage"""
    if total > 0:
        return round((completed / total) * 100)
    return 0


def _log_voip_change(call, old_values, user):
    """Log VoIP call changes"""
    changes = []
    for field, old_value in old_values.items():
        new_value = getattr(call, field)
        if old_value != new_value:
            changes.append(f"{field}: {old_value} → {new_value}")
    
    if changes:
        VoIPCallLog.objects.create(
            call=call,
            action='updated',
            description=f"Updated by {user.username}: {', '.join(changes)}"
        )


def _calculate_send_time(timing, scheduled_time):
    """Calculate email send time based on timing option"""
    if timing == 'immediate':
        return timezone.now()
    elif timing == 'delay_1h':
        return timezone.now() + timedelta(hours=1)
    elif timing == 'delay_24h':
        return timezone.now() + timedelta(days=1)
    elif timing == 'scheduled' and scheduled_time:
        # Parse scheduled time
        hour, minute = scheduled_time.split(':')
        send_at = timezone.now().replace(
            hour=int(hour),
            minute=int(minute),
            second=0,
            microsecond=0
        )
        # If time has passed today, schedule for tomorrow
        if send_at < timezone.now():
            send_at += timedelta(days=1)
        return send_at
    else:
        return timezone.now()


def _create_bulk_emails(obligations, template, send_at, user):
    """Create scheduled emails grouped by client"""
    from collections import defaultdict
    
    client_obligations = defaultdict(list)
    for obl in obligations:
        client_obligations[obl.client.id].append(obl)
    
    emails_created = 0
    for client_id, client_obls in client_obligations.items():
        # Create scheduled email (assuming you have this model/function)
        # This would need to be implemented based on your email system
        # scheduled_email = create_scheduled_email(
        #     obligations=client_obls,
        #     template=template,
        #     send_at=send_at,
        #     user=user
        # )
        emails_created += 1
    
    return emails_created


# Excel helper functions
def _get_filters_from_request(request):
    """Extract filters from request"""
    return {
        'status': request.GET.get('status', ''),
        'client': request.GET.get('client', ''),
        'type': request.GET.get('type', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'sort_by': request.GET.get('sort', 'deadline'),
    }


def _build_export_query(filters):
    """Build query for export"""
    now = timezone.now()
    
    # Base query
    if filters['date_from'] and filters['date_to']:
        query = MonthlyObligation.objects.filter(
            deadline__gte=filters['date_from'],
            deadline__lte=filters['date_to']
        )
    else:
        next_month = now.date() + timedelta(days=30)
        query = MonthlyObligation.objects.filter(
            deadline__range=[now.date(), next_month]
        )
    
    # Apply filters
    if filters['status']:
        query = query.filter(status=filters['status'])
    
    if filters['client']:
        query = query.filter(client_id=filters['client'])
    
    if filters['type']:
        query = query.filter(obligation_type_id=filters['type'])
    
    # Sort
    sort_mapping = {
        'deadline': 'deadline',
        'client': 'client__eponimia',
        'type': 'obligation_type__name',
        'status': 'status'
    }
    
    return query.select_related('client', 'obligation_type').order_by(
        sort_mapping.get(filters['sort_by'], 'deadline')
    )


def _apply_excel_styling(ws):
    """Apply styling to Excel worksheet"""
    ws['A1'] = 'ΑΝΑΦΟΡΑ ΥΠΟΧΡΕΩΣΕΩΝ'
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A2'] = f'Ημερομηνία: {timezone.now().strftime("%d/%m/%Y %H:%M")}'


def _write_excel_headers(ws):
    """Write headers to Excel worksheet"""
    headers = ['#', 'Προθεσμία', 'Πελάτης', 'ΑΦΜ', 'Υποχρέωση', 'Κατάσταση', 'Ώρες', 'Σημειώσεις']
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _write_excel_data(ws, query):
    """Write data rows to Excel worksheet"""
    row = 6
    for idx, obl in enumerate(query, start=1):
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = obl.deadline.strftime('%d/%m/%Y')
        ws.cell(row=row, column=3).value = obl.client.eponimia
        ws.cell(row=row, column=4).value = obl.client.afm
        ws.cell(row=row, column=5).value = obl.obligation_type.name
        ws.cell(row=row, column=6).value = obl.get_status_display()
        ws.cell(row=row, column=7).value = float(obl.time_spent) if obl.time_spent else ''
        ws.cell(row=row, column=8).value = obl.notes[:100] if obl.notes else ''
        
        # Color coding for status
        status_cell = ws.cell(row=row, column=6)
        if obl.status == 'completed':
            status_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        elif obl.status == 'overdue':
            status_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        
        row += 1


def _auto_adjust_excel_columns(ws):
    """Auto-adjust column widths in Excel"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# Analytics helper functions
def _calculate_monthly_completion_stats(start_date):
    """Calculate monthly completion statistics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).annotate(
        month_label=TruncMonth('deadline')
    ).values('month_label').annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        pending=Count('id', filter=Q(status='pending')),
        overdue=Count('id', filter=Q(status='overdue')),
    ).order_by('month_label')


def _calculate_client_performance(start_date):
    """Calculate client performance metrics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).values(
        'client__id',
        'client__eponimia'
    ).annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        completion_rate=Count('id', filter=Q(status='completed')) * 100.0 / Count('id')
    ).order_by('-total')[:10]


def _calculate_time_stats(start_date):
    """Calculate time tracking statistics"""
    return MonthlyObligation.objects.filter(
        status='completed',
        completed_date__gte=start_date,
        time_spent__isnull=False
    ).aggregate(
        total_hours=Sum('time_spent'),
        avg_hours=Avg('time_spent'),
        total_tasks=Count('id')
    )


def _calculate_revenue(start_date):
    """Calculate revenue data"""
    obligations = MonthlyObligation.objects.filter(
        status='completed',
        completed_date__gte=start_date,
        time_spent__isnull=False
    ).select_related('client')
    
    total_revenue = 0
    for obl in obligations:
        # Assuming hourly rate from client or default
        hourly_rate = 50  # Default rate, adjust as needed
        if hasattr(obl, 'hourly_rate'):
            hourly_rate = obl.hourly_rate
        from decimal import Decimal
        total_revenue += Decimal(obl.time_spent or 0) * hourly_rate
    
    return {'total': total_revenue}


def _calculate_type_stats(start_date):
    """Calculate obligation type statistics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).values(
        'obligation_type__name'
    ).annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        avg_time=Avg('time_spent', filter=Q(status='completed'))
    ).order_by('-total')[:10]


def _calculate_current_month_stats(now):
    """Calculate current month statistics"""
    current_month = now.month
    current_year = now.year
    
    return {
        'total': MonthlyObligation.objects.filter(
            year=current_year, month=current_month
        ).count(),
        'completed': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='completed'
        ).count(),
        'pending': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='pending'
        ).count(),
        'overdue': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='overdue'
        ).count(),
    }


def _format_chart_data(monthly_stats):
    """Format data for chart display"""
    chart_labels = [stat['month_label'].strftime('%b %Y') for stat in monthly_stats]
    
    return {
        'chart_labels_json': json.dumps(chart_labels),
        'chart_completed_json': json.dumps([stat['completed'] for stat in monthly_stats]),
        'chart_pending_json': json.dumps([stat['pending'] for stat in monthly_stats]),
        'chart_overdue_json': json.dumps([stat['overdue'] for stat in monthly_stats]),
    }

@require_GET
@staff_member_required
def check_obligation_duplicate(request):
    """
    AJAX endpoint για έλεγχο duplicate obligation
    """
    client_id = request.GET.get('client')
    type_id = request.GET.get('type')
    year = request.GET.get('year')
    month = request.GET.get('month')
    
    if not all([client_id, type_id, year, month]):
        return JsonResponse({'exists': False})
    
    exists = MonthlyObligation.objects.filter(
        client_id=client_id,
        obligation_type_id=type_id,
        year=year,
        month=month
    ).exists()
    
    return JsonResponse({'exists': exists})

# ============================================
# END OF VIEWS
# ============================================

logger.info("Accounting views module loaded successfully")

# Πρόσθεσε αυτά στο τέλος του accounting/views.py

# ============================================
# MISSING VOIP FUNCTIONS - QUICK FIX
# ============================================

@staff_member_required
@cache_page(60)  # Cache 1 minute
def voip_statistics(request):
    """
    Advanced statistics για VoIP calls
    """
    try:
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        
        stats = {
            'today': {
                'total': VoIPCall.objects.filter(started_at__date=today).count(),
                'missed': VoIPCall.objects.filter(started_at__date=today, status='missed').count(),
                'completed': VoIPCall.objects.filter(started_at__date=today, status='completed').count(),
            },
            'week': {
                'total': VoIPCall.objects.filter(started_at__date__gte=week_ago).count(),
                'missed': VoIPCall.objects.filter(started_at__date__gte=week_ago, status='missed').count(),
                'completed': VoIPCall.objects.filter(started_at__date__gte=week_ago, status='completed').count(),
            },
            'by_client': list(
                VoIPCall.objects.filter(client__isnull=False)
                .values('client__eponimia')
                .annotate(count=Count('id'))
                .order_by('-count')[:10]
            ),
            'by_resolution': list(
                VoIPCall.objects.values('resolution')
                .annotate(count=Count('id'))
                .order_by('-count')
            ),
        }
        
        # Calculate average duration
        durations = VoIPCall.objects.filter(
            status='completed',
            duration_seconds__gt=0
        ).values_list('duration_seconds', flat=True)[:100]
        
        if durations:
            stats['average_duration'] = sum(durations) / len(durations)
        else:
            stats['average_duration'] = 0
        
        logger.info(f"Statistics generated for {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Error generating statistics: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@staff_member_required
@require_POST
def voip_bulk_action(request):
    """
    Bulk update multiple VoIP calls
    """
    try:
        data = json.loads(request.body)
        call_ids = data.get('call_ids', [])
        action = data.get('action')
        value = data.get('value')
        
        if not isinstance(call_ids, list) or len(call_ids) == 0:
            return JsonResponse({
                'success': False, 
                'message': 'No calls selected'
            }, status=400)
        
        # Update based on action
        updated = 0
        if action == 'resolution':
            if value in ['pending', 'closed', 'follow_up', '']:
                updated = VoIPCall.objects.filter(id__in=call_ids).update(resolution=value)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid resolution value'
                }, status=400)
                
        elif action == 'status':
            if value in ['active', 'completed', 'missed', 'failed']:
                updated = VoIPCall.objects.filter(id__in=call_ids).update(status=value)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid status value'
                }, status=400)
        
        elif action == 'delete':
            updated = VoIPCall.objects.filter(id__in=call_ids).delete()[0]
        
        else:
            return JsonResponse({
                'success': False,
                'message': f'Unknown action: {action}'
            }, status=400)
        
        logger.info(f"Bulk update: {updated} calls updated by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': f'✅ {updated} κλήσεις ενημερώθηκαν!',
            'updated_count': updated
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
        
    except Exception as e:
        logger.error(f"Error in bulk action: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@staff_member_required
def voip_export_csv(request):
    """
    Export VoIP calls to CSV
    """
    import csv
    from django.http import HttpResponse
    
    try:
        # Create response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="voip_calls_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Add BOM for Excel UTF-8 compatibility
        response.write('\ufeff')
        
        # Create CSV writer
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow([
            'ID',
            'Αριθμός',
            'Πελάτης',
            'Email',
            'Κατεύθυνση',
            'Κατάσταση',
            'Διάρκεια',
            'Ημερομηνία',
            'Ώρα',
            'Σημειώσεις',
            'Resolution'
        ])
        
        # Get filtered calls
        calls_query = VoIPCall.objects.select_related('client').all()
        
        # Apply filters if provided
        status = request.GET.get('status')
        if status:
            calls_query = calls_query.filter(status=status)
        
        direction = request.GET.get('direction')
        if direction:
            calls_query = calls_query.filter(direction=direction)
        
        date_from = request.GET.get('date_from')
        if date_from:
            calls_query = calls_query.filter(started_at__date__gte=date_from)
        
        date_to = request.GET.get('date_to')
        if date_to:
            calls_query = calls_query.filter(started_at__date__lte=date_to)
        
        # Order and limit
        calls_query = calls_query.order_by('-started_at')[:1000]
        
        # Write data rows
        for call in calls_query:
            writer.writerow([
                call.id,
                call.phone_number,
                call.client.eponimia if call.client else '—',
                call.client_email or '—',
                call.get_direction_display(),
                call.get_status_display(),
                call.duration_formatted,
                call.started_at.strftime('%d/%m/%Y'),
                call.started_at.strftime('%H:%M:%S'),
                (call.notes[:100] + '...') if call.notes and len(call.notes) > 100 else (call.notes or ''),
                call.get_resolution_display() if call.resolution else '—',
            ])
        
        logger.info(f"CSV exported by {request.user.username}: {calls_query.count()} records")
        return response
        
    except Exception as e:
        logger.error(f"Error exporting CSV: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# CLASS-BASED VIEW FOR VOIP (if needed)
# ============================================

from django.views.generic import ListView

class VoIPCallsListView(ListView):
    """
    Django Class-Based View for VoIP calls listing
    """
    model = VoIPCall
    template_name = 'accounting/voip_calls_list.html'
    context_object_name = 'calls'
    paginate_by = 50
    
    def get_queryset(self):
        """Get filtered and sorted calls"""
        queryset = VoIPCall.objects.select_related('client').all()
        
        # Apply filters from GET parameters
        status = self.request.GET.get('status')
        direction = self.request.GET.get('direction')
        search = self.request.GET.get('search')
        
        if status:
            queryset = queryset.filter(status=status)
        if direction:
            queryset = queryset.filter(direction=direction)
        if search:
            queryset = queryset.filter(
                Q(phone_number__icontains=search) |
                Q(client__eponimia__icontains=search) |
                Q(notes__icontains=search)
            )
        
        return queryset.order_by('-started_at')
    
    def get_context_data(self, **kwargs):
        """Add statistics to context"""
        context = super().get_context_data(**kwargs)
        
        # Calculate stats
        context['total'] = VoIPCall.objects.count()
        context['missed'] = VoIPCall.objects.filter(status='missed').count()
        context['completed'] = VoIPCall.objects.filter(status='completed').count()
        
        if context['total'] > 0:
            context['success_rate'] = round(
                (context['completed'] / context['total']) * 100
            )
        else:
            context['success_rate'] = 0
        
        # Add filter values
        context['current_status'] = self.request.GET.get('status', '')
        context['current_direction'] = self.request.GET.get('direction', '')
        context['current_search'] = self.request.GET.get('search', '')
        
        return context

        # ============================================
# TICKET MANAGEMENT VIEWS (NEW!)
# ============================================

@staff_member_required
@require_POST
def assign_ticket(request, ticket_id):
    """Assign ticket to current user"""
    try:
        ticket = Ticket.objects.get(id=ticket_id)
        
        # Mark as assigned
        ticket.mark_as_assigned(request.user)
        
        # Log
        VoIPCallLog.objects.create(
            call=ticket.call,
            action='ticket_assigned',
            description=f"Ticket #{ticket_id} assigned to {request.user.username}"
        )
        
        logger.info(f"Ticket #{ticket_id} assigned to {request.user.username}")
        
        return JsonResponse({
            'success': True, 
            'message': f'✅ Ticket ανατέθηκε σε {request.user.get_full_name() or request.user.username}'
        })
        
    except Ticket.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '❌ Ticket δεν βρέθηκε'
        }, status=404)
    except Exception as e:
        logger.error(f"Error assigning ticket: {e}")
        return JsonResponse({
            'success': False,
            'message': f'❌ Σφάλμα: {str(e)}'
        }, status=500)


@staff_member_required
@require_POST
def update_ticket_status(request, ticket_id):
    """Update ticket status and notes"""
    try:
        ticket = Ticket.objects.get(id=ticket_id)
        
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid JSON'
            }, status=400)
        
        # Update status
        if 'status' in data and data['status'] in ['open', 'assigned', 'in_progress', 'resolved', 'closed']:
            ticket.status = data['status']
        
        # Update notes
        if 'notes' in data:
            new_note = data['notes']
            timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
            if ticket.notes:
                ticket.notes += f"\n[{timestamp}] {new_note}"
            else:
                ticket.notes = f"[{timestamp}] {new_note}"
        
        # Auto-mark dates
        if ticket.status == 'resolved' and not ticket.resolved_at:
            ticket.mark_as_resolved()
        elif ticket.status == 'closed' and not ticket.closed_at:
            ticket.mark_as_closed()
        
        ticket.save()
        
        # Log
        VoIPCallLog.objects.create(
            call=ticket.call,
            action='ticket_updated',
            description=f"Ticket #{ticket_id} status → {ticket.get_status_display()}"
        )
        
        logger.info(f"Ticket #{ticket_id} updated to status {ticket.status}")
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Ticket ενημερώθηκε!',
            'ticket': {
                'id': ticket.id,
                'status': ticket.get_status_display(),
                'notes': ticket.notes,
            }
        })
        
    except Ticket.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '❌ Ticket δεν βρέθηκε'
        }, status=404)
    except Exception as e:
        logger.error(f"Error updating ticket: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'❌ Σφάλμα: {str(e)}'
        }, status=500)


@staff_member_required
@require_POST
def send_ticket_email(request):
    """Send email about ticket"""
    try:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid JSON'
            }, status=400)
        
        ticket_id = data.get('ticket_id')
        template_id = data.get('template_id')
        
        if not ticket_id or not template_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing ticket_id or template_id'
            }, status=400)
        
        # Get ticket
        ticket = Ticket.objects.get(id=ticket_id)
        
        # Get template
        template = EmailTemplate.objects.get(id=template_id, is_active=True)
        
        # Prepare context
        context = {
            'ticket_id': ticket.id,
            'phone': ticket.call.phone_number,
            'client_name': ticket.client.eponimia if ticket.client else 'Unknown',
            'ticket_title': ticket.title,
            'ticket_priority': ticket.get_priority_display(),
            'ticket_status': ticket.get_status_display(),
            'user_name': request.user.get_full_name() or request.user.username,
            'company_name': settings.COMPANY_NAME,
        }
        
        # Render template
        subject, body = template.render(context)
        
        # Get recipient
        recipient = ticket.call.client_email or ticket.client.email if ticket.client else settings.DEFAULT_FROM_EMAIL
        
        if not recipient:
            return JsonResponse({
                'success': False,
                'message': '❌ Δεν υπάρχει email για αποστολή'
            }, status=400)
        
        # Send email
        send_mail(
            subject=subject,
            message=strip_tags(body),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[recipient],
            html_message=body,
            fail_silently=False,
        )
        
        # Log email
        ScheduledEmail.objects.create(
            recipient_email=recipient,
            recipient_name=ticket.client.eponimia if ticket.client else 'Client',
            client=ticket.client,
            template=template,
            subject=subject,
            body_html=body,
            send_at=timezone.now(),
            created_by=request.user,
            status='sent',
            sent_at=timezone.now(),
        )
        
        # Update ticket
        ticket.email_sent = True
        ticket.save()
        
        logger.info(f"Email sent for ticket #{ticket_id} to {recipient}")
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Email στάλθηκε στο {recipient}'
        })
        
    except Ticket.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '❌ Ticket δεν βρέθηκε'
        }, status=404)
    except EmailTemplate.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '❌ Template δεν βρέθηκε'
        }, status=404)
    except Exception as e:
        logger.error(f"Error sending email: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'❌ Σφάλμα: {str(e)}'
        }, status=500)

# ============================================
# CALENDAR VIEWS
# ============================================
import calendar

@staff_member_required
@login_required
def calendar_view(request):
    """
    Προβολή ημερολογίου υποχρεώσεων με FullCalendar
    Features: Month/Week/List views, filters, color-coded status
    """
    today = timezone.now().date()

    # Get all clients for filter dropdown
    clients = ClientProfile.objects.filter(is_active=True).order_by('eponimia')

    # Get all obligation types for filter dropdown
    obligation_types = ObligationType.objects.filter(is_active=True).order_by('name')

    # Calculate statistics for the current month
    current_month_start = today.replace(day=1)
    if today.month == 12:
        next_month_start = today.replace(year=today.year + 1, month=1, day=1)
    else:
        next_month_start = today.replace(month=today.month + 1, day=1)

    month_obligations = MonthlyObligation.objects.filter(
        deadline__gte=current_month_start,
        deadline__lt=next_month_start
    )

    stats = {
        'pending_count': month_obligations.filter(status='pending').count(),
        'overdue_count': MonthlyObligation.objects.filter(
            status='pending',
            deadline__lt=today
        ).count(),
        'completed_this_month': month_obligations.filter(status='completed').count(),
    }

    context = {
        "title": "Ημερολόγιο Υποχρεώσεων",
        "clients": clients,
        "obligation_types": obligation_types,
        "today": today,
        "stats": stats,
    }
    return render(request, "accounting/calendar.html", context)


@staff_member_required
@login_required
@require_GET
def calendar_events_api(request):
    """
    API endpoint for FullCalendar events
    Returns JSON with obligations formatted for FullCalendar
    """
    # Get date range from FullCalendar
    start_str = request.GET.get('start', '')
    end_str = request.GET.get('end', '')

    # Get filter parameters
    client_id = request.GET.get('client', '')
    obligation_type_id = request.GET.get('type', '')
    status_filter = request.GET.get('status', '')

    # Parse dates
    try:
        if start_str:
            start_date = datetime.strptime(start_str[:10], '%Y-%m-%d').date()
        else:
            start_date = timezone.now().date().replace(day=1)

        if end_str:
            end_date = datetime.strptime(end_str[:10], '%Y-%m-%d').date()
        else:
            # Default to end of month
            if start_date.month == 12:
                end_date = start_date.replace(year=start_date.year + 1, month=1, day=1)
            else:
                end_date = start_date.replace(month=start_date.month + 1, day=1)
    except ValueError:
        start_date = timezone.now().date().replace(day=1)
        end_date = start_date + timedelta(days=31)

    # Build query
    queryset = MonthlyObligation.objects.filter(
        deadline__gte=start_date,
        deadline__lte=end_date
    ).select_related('client', 'obligation_type')

    # Apply filters
    if client_id:
        try:
            queryset = queryset.filter(client_id=int(client_id))
        except ValueError:
            pass

    if obligation_type_id:
        try:
            queryset = queryset.filter(obligation_type_id=int(obligation_type_id))
        except ValueError:
            pass

    if status_filter:
        if status_filter == 'overdue':
            queryset = queryset.filter(status='pending', deadline__lt=timezone.now().date())
        else:
            queryset = queryset.filter(status=status_filter)

    # Define colors based on status
    status_colors = {
        'pending': '#f59e0b',    # Yellow/amber
        'completed': '#22c55e',  # Green
        'overdue': '#ef4444',    # Red
    }

    today = timezone.now().date()

    # Build events list
    events = []
    for obligation in queryset:
        # Determine actual status (check for overdue)
        actual_status = obligation.status
        if actual_status == 'pending' and obligation.deadline and obligation.deadline < today:
            actual_status = 'overdue'

        color = status_colors.get(actual_status, '#6b7280')

        # Build event title
        client_name = obligation.client.eponimia if obligation.client else 'Άγνωστος'
        type_name = obligation.obligation_type.name if obligation.obligation_type else 'Άγνωστος'

        event = {
            'id': obligation.id,
            'title': f"{type_name} - {client_name}",
            'start': obligation.deadline.isoformat() if obligation.deadline else None,
            'end': obligation.deadline.isoformat() if obligation.deadline else None,
            'color': color,
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'status': actual_status,
                'client_id': obligation.client_id,
                'client_name': client_name,
                'client_afm': obligation.client.afm if obligation.client else '',
                'obligation_type': type_name,
                'obligation_type_id': obligation.obligation_type_id,
                'period': f"{obligation.month:02d}/{obligation.year}",
                'notes': obligation.notes or '',
                'edit_url': f"/admin/accounting/monthlyobligation/{obligation.id}/change/",
            }
        }
        events.append(event)

    return JsonResponse({'events': events})

class ClientDocumentViewSet(viewsets.ModelViewSet):
    queryset = ClientDocument.objects.all()
    serializer_class = ClientDocumentSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        client_id = self.request.query_params.get('client')
        if client_id:
            queryset = queryset.filter(client_id=client_id)
        return queryset

# accounting/views.py - Πρόσθεσε prints για debug

@require_POST
@staff_member_required
def advanced_bulk_complete(request):
    """SIMPLIFIED VERSION with logging"""

    logger.debug("=" * 60)
    logger.debug("ADVANCED BULK COMPLETE - START")

    try:
        # Parse data
        completion_data = json.loads(request.POST.get('completion_data', '[]'))
        logger.debug(f"Got {len(completion_data)} groups to process")

        completed_count = 0
        errors = []

        for group_data in completion_data:
            client_afm = group_data.get('client_afm')
            group_num = group_data.get('group', '0')
            obligation_ids = group_data.get('obligations', [])

            logger.debug(f"--- Processing AFM: {client_afm}, Group: {group_num} ---")
            logger.debug(f"    Obligations to process: {obligation_ids}")

            # Get files
            files_key = f"file_{client_afm}_{group_num}"
            files = request.FILES.getlist(files_key)
            logger.debug(f"    Files found with key '{files_key}': {len(files)}")

            # Process each obligation
            for idx, obl_id in enumerate(obligation_ids):
                logger.debug(f"    Processing obligation {obl_id}...")

                try:
                    obligation = MonthlyObligation.objects.get(id=obl_id)
                    logger.debug(f"      Found: {obligation}")

                    # Update status
                    obligation.status = 'completed'
                    obligation.completed_date = timezone.now().date()
                    obligation.completed_by = request.user

                    # Try to attach file if available
                    if idx < len(files) and group_num == '0':
                        # Individual file
                        file_to_use = files[idx]
                        logger.debug(f"      Attaching file: {file_to_use.name}")

                        try:
                            archive_path = obligation.archive_attachment(file_to_use)
                            logger.debug(f"      ✅ Archived to: {archive_path}")
                        except Exception as e:
                            logger.debug(f"      ❌ Archive failed: {e}")
                            obligation.save()  # Save without file
                    else:
                        # No file or group mode
                        obligation.save()
                        logger.debug(f"      Saved without file")

                    completed_count += 1
                    logger.debug(f"      ✅ Completed successfully")

                except MonthlyObligation.DoesNotExist:
                    logger.debug(f"      ❌ Obligation {obl_id} not found!")
                    errors.append(f"Not found: {obl_id}")
                except Exception as e:
                    logger.debug(f"      ❌ Error: {e}")
                    logger.debug(traceback.format_exc())
                    errors.append(str(e))

        logger.debug(f"=== FINAL RESULTS ===")
        logger.debug(f"Completed: {completed_count}")
        logger.debug(f"Errors: {len(errors)}")
        logger.debug("=" * 60)
        
        return JsonResponse({
            'success': completed_count > 0,
            'completed_count': completed_count,
            'message': f'Ολοκληρώθηκαν {completed_count} υποχρεώσεις',
            'errors': errors
        })
        
    except Exception as e:
        logger.error(f"CRITICAL ERROR: {e}")
        logger.error(traceback.format_exc())

        return JsonResponse({
            'success': False,
            'message': str(e),
            'completed_count': 0
        })


# accounting/views.py - Door control using Tasmota
# SECURITY: IP addresses loaded from Django settings

import requests
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import logging

logger = logging.getLogger(__name__)

# SECURITY FIX: Load IP from settings instead of hardcoding
TASMOTA_IP = settings.TASMOTA_IP
TASMOTA_PORT = settings.TASMOTA_PORT
TIMEOUT = 2  # 2 seconds


@require_http_methods(["GET"])
def door_status(request):
    """Check door status - ON or OFF"""
    try:
        url = f"http://{TASMOTA_IP}:{TASMOTA_PORT}/cm?cmnd=Power"
        
        logger.info(f"🔍 Checking status at {TASMOTA_IP}")
        response = requests.get(url, timeout=TIMEOUT)
        
        if response.status_code == 200:
            data = response.json()
            power = data.get("POWER", "OFF")
            
            logger.info(f"✅ Status: {power}")
            
            return JsonResponse({
                "success": True,
                "status": "open" if power == "ON" else "closed",
                "raw_power": power
            })
        else:
            logger.error(f"❌ HTTP {response.status_code}")
            return JsonResponse({
                "success": False,
                "error": f"HTTP {response.status_code}"
            }, status=500)
            
    except requests.exceptions.Timeout:
        logger.error(f"⏱️ Timeout connecting to {TASMOTA_IP}")
        return JsonResponse({
            "success": False,
            "error": "Timeout"
        }, status=504)
        
    except requests.exceptions.ConnectionError:
        logger.error(f"🔴 Cannot connect to {TASMOTA_IP}")
        return JsonResponse({
            "success": False,
            "error": f"Cannot connect to {TASMOTA_IP}"
        }, status=503)
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=500)


@require_http_methods(["POST"])
def open_door(request):
    """
    Toggle door - ON ↔ OFF
    ✅ SECURITY FIX: CSRF protection enabled to prevent unauthorized door control
    """
    try:
        # TOGGLE command
        url = f"http://{TASMOTA_IP}:{TASMOTA_PORT}/cm?cmnd=Power%20TOGGLE"
        
        logger.info(f"🔄 Toggling door at {TASMOTA_IP}")
        response = requests.get(url, timeout=TIMEOUT)
        
        if response.status_code == 200:
            data = response.json()
            new_state = data.get("POWER", "UNKNOWN")
            
            logger.info(f"✅ New state: {new_state}")
            
            return JsonResponse({
                "success": True,
                "new_state": new_state,
                "message": f"Πόρτα τώρα: {new_state}"
            })
        else:
            logger.error(f"❌ HTTP {response.status_code}")
            return JsonResponse({
                "success": False,
                "error": f"HTTP {response.status_code}"
            }, status=500)
        
    except requests.exceptions.Timeout:
        logger.error(f"⏱️ Timeout toggling door at {TASMOTA_IP}")
        return JsonResponse({
            "success": False,
            "error": "Timeout"
        }, status=504)
        
    except requests.exceptions.ConnectionError:
        logger.error(f"🔴 Cannot connect to {TASMOTA_IP}")
        return JsonResponse({
            "success": False,
            "error": f"Cannot connect to {TASMOTA_IP}"
        }, status=503)
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=500)


@require_http_methods(["GET", "POST"])
def door_control(request):
    """Unified door control endpoint"""
    if request.method == "POST":
        return open_door(request)
    else:
        return door_status(request)

# ============================================================================
# QUICK COMPLETE & FILE UPLOAD VIEWS
# ============================================================================

@require_POST
@staff_member_required
def quick_complete_obligation(request, obligation_id):
    """
    Quick complete obligation without file upload
    AJAX endpoint για γρήγορη ολοκλήρωση με προαιρετικό email notification
    """
    try:
        import json
        obligation = MonthlyObligation.objects.get(id=obligation_id)

        # Parse JSON body
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}

        # Get optional parameters
        send_email = data.get('send_email', False)
        time_spent = data.get('time_spent')

        # Update status
        old_status = obligation.status
        obligation.status = 'completed'
        obligation.completed_date = timezone.now().date()  # DateField, not datetime
        obligation.completed_by = request.user

        # Update time spent if provided
        if time_spent:
            try:
                obligation.time_spent = float(time_spent)
            except (ValueError, TypeError):
                pass

        obligation.save()

        # Log to audit trail
        from common.models import AuditLog
        changes = {
            'status': {'old': old_status, 'new': 'completed'}
        }
        if time_spent:
            changes['time_spent'] = {'old': None, 'new': time_spent}

        AuditLog.log(
            user=request.user,
            action='update',
            obj=obligation,
            changes=changes,
            description=f'Γρήγορη ολοκλήρωση υποχρέωσης: {obligation}',
            severity='medium',
            request=request
        )

        # Send email notification if requested
        if send_email:
            from accounting.services.email_service import trigger_automation_rules
            emails_created = trigger_automation_rules(obligation, trigger_type='on_complete')
            logger.info(f'📧 Created {len(emails_created)} email notifications for obligation {obligation_id}')

        return JsonResponse({
            'success': True,
            'message': 'Η υποχρέωση ολοκληρώθηκε επιτυχώς!' +
                      (' (Email στάλθηκε)' if send_email else '')
        })

    except MonthlyObligation.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Η υποχρέωση δεν βρέθηκε'
        }, status=404)
    except Exception as e:
        logger.error(f'Error completing obligation {obligation_id}: {e}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_POST
@staff_member_required
def complete_with_file(request, obligation_id):
    """
    Complete obligation WITH file upload
    Handles multipart/form-data για file upload + ολοκλήρωση
    """
    try:
        obligation = MonthlyObligation.objects.get(id=obligation_id)

        # Validate file
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'Δεν ανεβάσατε αρχείο'
            }, status=400)

        uploaded_file = request.FILES['file']

        # ✅ SECURITY: File validation
        from common.utils.file_validation import validate_file_upload
        try:
            validate_file_upload(uploaded_file)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Μη έγκυρο αρχείο: {str(e)}'
            }, status=400)

        # Create ClientDocument
        category = request.POST.get('category', 'general')
        description = request.POST.get('description', '')

        document = ClientDocument.objects.create(
            client=obligation.client,
            obligation=obligation,
            file=uploaded_file,
            filename=uploaded_file.name,
            file_type=uploaded_file.content_type,
            document_category=category,
            description=description
        )

        # Update obligation
        old_status = obligation.status
        obligation.status = 'completed'
        obligation.completed_date = timezone.now().date()  # DateField, not datetime
        obligation.completed_by = request.user
        obligation.attachment = uploaded_file  # Set primary attachment

        # Update time spent if provided
        time_spent = request.POST.get('time_spent')
        if time_spent:
            try:
                obligation.time_spent = float(time_spent)
            except ValueError:
                pass

        obligation.save()

        # Log to audit trail
        from common.models import AuditLog
        AuditLog.log(
            user=request.user,
            action='update',
            obj=obligation,
            changes={
                'status': {'old': old_status, 'new': 'completed'},
                'attachment': {'old': None, 'new': uploaded_file.name}
            },
            description=f'Ολοκλήρωση με αρχείο: {obligation}',
            severity='medium',
            request=request
        )

        # Send email notification if requested
        send_email = request.POST.get('send_email') == '1'
        if send_email:
            from accounting.services.email_service import trigger_automation_rules
            emails_created = trigger_automation_rules(obligation, trigger_type='on_complete')
            logger.info(f'📧 Created {len(emails_created)} email notifications for obligation {obligation_id}')

        return JsonResponse({
            'success': True,
            'message': 'Το αρχείο ανέβηκε και η υποχρέωση ολοκληρώθηκε!' +
                      (' (Email στάλθηκε)' if send_email else ''),
            'document_id': document.id,
            'obligation_id': obligation.id
        })

    except MonthlyObligation.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Η υποχρέωση δεν βρέθηκε'
        }, status=404)
    except Exception as e:
        logger.error(f'Error completing obligation {obligation_id} with file: {e}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_POST
@staff_member_required
def bulk_complete_obligations(request):
    """
    Bulk complete multiple obligations with optional file upload
    Handles μαζική ολοκλήρωση με προαιρετικό αρχείο
    """
    try:
        import json

        # Get obligation IDs
        obligation_ids_str = request.POST.get('obligation_ids')
        if not obligation_ids_str:
            return JsonResponse({
                'success': False,
                'error': 'Δεν επιλέξατε υποχρεώσεις'
            }, status=400)

        obligation_ids = json.loads(obligation_ids_str)

        if not obligation_ids:
            return JsonResponse({
                'success': False,
                'error': 'Δεν επιλέξατε υποχρεώσεις'
            }, status=400)

        # Get obligations
        obligations = MonthlyObligation.objects.filter(
            id__in=obligation_ids,
            status__in=['pending', 'overdue']
        )

        if not obligations.exists():
            return JsonResponse({
                'success': False,
                'error': 'Δεν βρέθηκαν έγκυρες υποχρεώσεις προς ολοκλήρωση'
            }, status=404)

        # Get optional parameters
        send_email = request.POST.get('send_email') == '1'
        category = request.POST.get('category', 'general')
        description = request.POST.get('description', '')
        uploaded_file = request.FILES.get('file')

        # Validate file if provided
        if uploaded_file:
            from common.utils.file_validation import validate_file_upload
            try:
                validate_file_upload(uploaded_file)
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Μη έγκυρο αρχείο: {str(e)}'
                }, status=400)

        completed_count = 0
        failed_count = 0
        errors = []

        # Complete each obligation
        for obligation in obligations:
            try:
                old_status = obligation.status
                obligation.status = 'completed'
                obligation.completed_date = timezone.now().date()
                obligation.completed_by = request.user

                # Attach file if provided
                if uploaded_file:
                    # Create document for this obligation
                    ClientDocument.objects.create(
                        client=obligation.client,
                        obligation=obligation,
                        file=uploaded_file,
                        filename=uploaded_file.name,
                        file_type=uploaded_file.content_type,
                        document_category=category,
                        description=description or f'Μαζική ολοκλήρωση - {timezone.now().date()}'
                    )
                    obligation.attachment = uploaded_file

                obligation.save()

                # Audit log
                from common.models import AuditLog
                AuditLog.log(
                    user=request.user,
                    action='update',
                    obj=obligation,
                    changes={
                        'status': {'old': old_status, 'new': 'completed'},
                        'bulk_completion': True
                    },
                    description=f'Μαζική ολοκλήρωση: {obligation}',
                    severity='medium',
                    request=request
                )

                # Send email if requested
                if send_email:
                    from accounting.services.email_service import trigger_automation_rules
                    trigger_automation_rules(obligation, trigger_type='on_complete')

                completed_count += 1

            except Exception as e:
                failed_count += 1
                errors.append(f'{obligation.client.eponimia} - {obligation.obligation_type.name}: {str(e)}')
                logger.error(f'Error bulk completing obligation {obligation.id}: {e}')

        # Build response message
        message = f'Ολοκληρώθηκαν {completed_count} υποχρεώσεις επιτυχώς'
        if failed_count > 0:
            message += f' ({failed_count} απέτυχαν)'

        return JsonResponse({
            'success': True,
            'message': message,
            'completed_count': completed_count,
            'failed_count': failed_count,
            'errors': errors if errors else None
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Μη έγκυρα δεδομένα υποχρεώσεων'
        }, status=400)
    except Exception as e:
        logger.error(f'Error in bulk completion: {e}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@staff_member_required
def obligation_detail_view(request, obligation_id):
    """
    Detail view για MonthlyObligation
    Επιτρέπει προβολή και επεξεργασία όλων των πεδίων + upload εγγράφων
    """
    try:
        obligation = MonthlyObligation.objects.select_related(
            'client', 'obligation_type', 'completed_by'
        ).prefetch_related(
            'client__documents'
        ).get(id=obligation_id)

        # Get all documents for this obligation
        documents = ClientDocument.objects.filter(
            Q(obligation=obligation) | Q(client=obligation.client)
        ).order_by('-uploaded_at')

        # Handle POST (edit obligation)
        if request.method == 'POST':
            # Update obligation fields
            obligation.notes = request.POST.get('notes', '')
            
            time_spent = request.POST.get('time_spent')
            if time_spent:
                try:
                    obligation.time_spent = float(time_spent)
                except ValueError:
                    pass

            hourly_rate = request.POST.get('hourly_rate')
            if hourly_rate:
                try:
                    obligation.hourly_rate = float(hourly_rate)
                except ValueError:
                    pass

            obligation.save()

            messages.success(request, '✅ Η υποχρέωση ενημερώθηκε επιτυχώς!')
            return redirect('accounting:obligation_detail', obligation_id=obligation.id)

        context = {
            'obligation': obligation,
            'documents': documents,
            'title': f'Υποχρέωση #{obligation.id} - {obligation.client.eponimia}',
        }

        return render(request, 'accounting/obligation_detail.html', context)

    except MonthlyObligation.DoesNotExist:
        messages.error(request, 'Η υποχρέωση δεν βρέθηκε')
        return redirect('accounting:dashboard')


# ============================================
# WIZARD API - Get Obligation Details for Wizard
# ============================================

@staff_member_required
@require_GET
def api_obligations_wizard(request):
    """
    API endpoint για το wizard μαζικής ολοκλήρωσης.
    Επιστρέφει λεπτομέρειες για τις επιλεγμένες υποχρεώσεις.
    """
    try:
        ids_param = request.GET.get('ids', '')
        if not ids_param:
            return JsonResponse({
                'success': False,
                'error': 'Δεν παρέχονται IDs υποχρεώσεων'
            }, status=400)

        # Parse IDs
        try:
            ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Μη έγκυρα IDs υποχρεώσεων'
            }, status=400)

        if not ids:
            return JsonResponse({
                'success': False,
                'error': 'Δεν παρέχονται IDs υποχρεώσεων'
            }, status=400)

        # Get obligations with related data
        obligations = MonthlyObligation.objects.filter(
            id__in=ids
        ).select_related(
            'client', 'obligation_type'
        ).order_by('deadline', 'client__eponimia')

        # Format data for wizard
        MONTH_NAMES = {
            1: 'Ιανουάριος', 2: 'Φεβρουάριος', 3: 'Μάρτιος', 4: 'Απρίλιος',
            5: 'Μάιος', 6: 'Ιούνιος', 7: 'Ιούλιος', 8: 'Αύγουστος',
            9: 'Σεπτέμβριος', 10: 'Οκτώβριος', 11: 'Νοέμβριος', 12: 'Δεκέμβριος'
        }

        obligations_data = []
        for ob in obligations:
            # Get existing documents for this obligation
            existing_docs = ClientDocument.objects.filter(
                obligation=ob
            ).values('id', 'filename', 'uploaded_at')

            obligations_data.append({
                'id': ob.id,
                'client_id': ob.client.id,
                'client_name': ob.client.eponimia,
                'client_afm': ob.client.afm,
                'client_email': ob.client.email or '',
                'obligation_type': ob.obligation_type.name,
                'obligation_code': ob.obligation_type.code,
                'period_month': ob.month,
                'period_year': ob.year,
                'period_display': f"{MONTH_NAMES.get(ob.month, ob.month)} {ob.year}",
                'due_date': ob.deadline.strftime('%d/%m/%Y') if ob.deadline else '',
                'due_date_iso': ob.deadline.isoformat() if ob.deadline else '',
                'status': ob.status,
                'notes': ob.notes or '',
                'has_attachment': bool(ob.attachment),
                'existing_documents': list(existing_docs),
            })

        return JsonResponse({
            'success': True,
            'count': len(obligations_data),
            'obligations': obligations_data
        })

    except Exception as e:
        logger.error(f"Error in api_obligations_wizard: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================
# WIZARD BULK PROCESS - Process Wizard Submissions
# ============================================

@require_POST
@staff_member_required
def wizard_bulk_process(request):
    """
    Επεξεργασία υποβολής wizard για μαζική ολοκλήρωση υποχρεώσεων.
    Κάθε υποχρέωση μπορεί να έχει το δικό της αρχείο.
    """
    try:
        # Parse results JSON
        results_str = request.POST.get('results', '{}')
        try:
            results = json.loads(results_str)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Μη έγκυρα δεδομένα αποτελεσμάτων'
            }, status=400)

        if not results:
            return JsonResponse({
                'success': False,
                'error': 'Δεν υπάρχουν υποχρεώσεις προς επεξεργασία'
            }, status=400)

        global_notes = request.POST.get('notes', '')

        completed_count = 0
        skipped_count = 0
        failed_count = 0
        errors = []
        processed_details = []

        # Process each obligation from results
        for ob_id_str, ob_data in results.items():
            try:
                ob_id = int(ob_id_str)

                # Skip if marked to skip
                if ob_data.get('skip', False):
                    skipped_count += 1
                    continue

                # Skip if not marked as complete
                if not ob_data.get('complete', False):
                    skipped_count += 1
                    continue

                obligation = MonthlyObligation.objects.select_related(
                    'client', 'obligation_type'
                ).get(id=ob_id)

                # Skip if already completed
                if obligation.status == 'completed':
                    skipped_count += 1
                    processed_details.append({
                        'id': ob_id,
                        'status': 'skipped',
                        'message': f'{obligation.client.eponimia} - {obligation.obligation_type.name}: Ήδη ολοκληρωμένη'
                    })
                    continue

                old_status = obligation.status

                # Update obligation
                obligation.status = 'completed'
                obligation.completed_date = timezone.now().date()
                obligation.completed_by = request.user

                # Handle time spent if provided
                time_spent = ob_data.get('time_spent')
                if time_spent:
                    try:
                        obligation.time_spent = float(time_spent)
                    except (ValueError, TypeError):
                        pass

                # Handle notes
                notes = ob_data.get('notes', '')
                combined_notes = f"{notes}\n{global_notes}".strip() if notes or global_notes else ''
                if combined_notes:
                    timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
                    new_note = f"[{timestamp}] [WIZARD] {combined_notes}"
                    if obligation.notes:
                        obligation.notes += f"\n{new_note}"
                    else:
                        obligation.notes = new_note

                # Handle file upload for this specific obligation
                file_key = f'file_{ob_id}'
                if file_key in request.FILES:
                    uploaded_file = request.FILES[file_key]

                    # Use archive_attachment method for proper file organization
                    try:
                        archive_path = obligation.archive_attachment(uploaded_file)
                        logger.info(f"Wizard: Archived file for obligation {ob_id}: {archive_path}")
                    except Exception as file_error:
                        logger.warning(f"Could not archive file for {ob_id}: {file_error}")
                        # Fallback: create ClientDocument
                        ClientDocument.objects.create(
                            client=obligation.client,
                            obligation=obligation,
                            file=uploaded_file,
                            filename=uploaded_file.name,
                            document_category=ob_data.get('category', 'general'),
                            description=f"Wizard upload - {timezone.now().strftime('%d/%m/%Y')}"
                        )

                obligation.save()

                # Audit log
                try:
                    from common.models import AuditLog
                    AuditLog.log(
                        user=request.user,
                        action='update',
                        obj=obligation,
                        changes={
                            'status': {'old': old_status, 'new': 'completed'},
                            'wizard_completion': True
                        },
                        description=f'Wizard ολοκλήρωση: {obligation}',
                        severity='medium',
                        request=request
                    )
                except Exception as audit_error:
                    logger.warning(f"Could not create audit log: {audit_error}")

                # Send email if requested for this specific obligation
                should_send_email = ob_data.get('send_email', False)
                email_sent = False
                email_error_msg = None

                if should_send_email and obligation.client.email:
                    try:
                        from accounting.services.email_service import EmailService
                        success, result = EmailService.send_obligation_completion_email(
                            obligation=obligation,
                            user=request.user,
                            include_attachment=True
                        )
                        if success:
                            email_sent = True
                            logger.info(f"📧 Email sent for obligation {ob_id} to {obligation.client.email}")
                        else:
                            email_error_msg = str(result)
                            logger.warning(f"Could not send email for {ob_id}: {result}")
                    except Exception as email_error:
                        email_error_msg = str(email_error)
                        logger.warning(f"Could not send email for {ob_id}: {email_error}")

                completed_count += 1
                message_parts = [f'{obligation.client.eponimia} - {obligation.obligation_type.name}: Ολοκληρώθηκε']
                if email_sent:
                    message_parts.append('📧')
                elif should_send_email and not email_sent:
                    message_parts.append(f'(email: {email_error_msg or "αποτυχία"})')

                processed_details.append({
                    'id': ob_id,
                    'status': 'completed',
                    'email_sent': email_sent,
                    'message': ' '.join(message_parts)
                })

            except MonthlyObligation.DoesNotExist:
                failed_count += 1
                errors.append(f'Υποχρέωση {ob_id_str} δεν βρέθηκε')
            except Exception as e:
                failed_count += 1
                errors.append(f'Σφάλμα με υποχρέωση {ob_id_str}: {str(e)}')
                logger.error(f"Error processing obligation {ob_id_str} in wizard: {e}", exc_info=True)

        # Build response message
        if completed_count > 0:
            message = f'✅ Ολοκληρώθηκαν {completed_count} υποχρεώσεις επιτυχώς!'
            if skipped_count > 0:
                message += f' ({skipped_count} παραλείφθηκαν)'
            if failed_count > 0:
                message += f' (⚠️ {failed_count} απέτυχαν)'
            success = True
        else:
            if skipped_count > 0:
                message = f'ℹ️ {skipped_count} υποχρεώσεις παραλείφθηκαν, καμία δεν ολοκληρώθηκε'
                success = True
            else:
                message = '❌ Καμία υποχρέωση δεν ολοκληρώθηκε'
                success = False

        logger.info(f"Wizard bulk process by {request.user.username}: "
                   f"completed={completed_count}, skipped={skipped_count}, failed={failed_count}")

        return JsonResponse({
            'success': success,
            'message': message,
            'completed_count': completed_count,
            'skipped_count': skipped_count,
            'failed_count': failed_count,
            'errors': errors[:10] if errors else [],
            'details': processed_details[:20] if processed_details else []
        })

    except Exception as e:
        logger.error(f"Critical error in wizard_bulk_process: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'Κρίσιμο σφάλμα: {str(e)}'
        }, status=500)


# ============================================
# GLOBAL SEARCH API
# ============================================

@require_GET
@staff_member_required
def global_search_api(request):
    """
    Global search API for clients and obligations.
    Searches by client name (eponimia), AFM, email, and obligation type.
    Returns max 5 results per category.
    """
    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'results': [], 'clients': [], 'obligations': []})

    try:
        # Search clients by eponimia, afm, email, or phone
        clients = ClientProfile.objects.filter(
            Q(eponimia__icontains=query) |
            Q(afm__icontains=query) |
            Q(email__icontains=query) |
            Q(kinito_tilefono__icontains=query) |
            Q(tilefono_epixeirisis_1__icontains=query)
        ).filter(is_active=True)[:5]

        # Search obligations by client name or obligation type name
        obligations = MonthlyObligation.objects.filter(
            Q(client__eponimia__icontains=query) |
            Q(obligation_type__name__icontains=query) |
            Q(client__afm__icontains=query)
        ).select_related('client', 'obligation_type').order_by('-deadline')[:5]

        # Format results
        clients_data = [{
            'id': c.id,
            'name': c.eponimia,
            'afm': c.afm,
            'email': c.email or '',
            'url': f'/admin/accounting/clientprofile/{c.id}/change/',
            'type': 'client'
        } for c in clients]

        obligations_data = [{
            'id': o.id,
            'name': str(o),
            'client': o.client.eponimia,
            'client_afm': o.client.afm,
            'obligation_type': o.obligation_type.name,
            'period': f'{o.month:02d}/{o.year}',
            'status': o.status,
            'url': f'/admin/accounting/monthlyobligation/{o.id}/change/',
            'type': 'obligation'
        } for o in obligations]

        return JsonResponse({
            'success': True,
            'query': query,
            'clients': clients_data,
            'obligations': obligations_data,
            'total': len(clients_data) + len(obligations_data)
        })

    except Exception as e:
        logger.error(f"Error in global_search_api: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e),
            'clients': [],
            'obligations': []
        }, status=500)


# ============================================
# PDF REPORT GENERATION
# ============================================

@staff_member_required
def client_report_pdf(request, client_id):
    """
    Generate PDF report for a specific client.
    Includes client info and obligation history.
    """
    from io import BytesIO
    from django.template.loader import render_to_string
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return HttpResponse(
            'xhtml2pdf δεν είναι εγκατεστημένο. Εγκαταστήστε το με: pip install xhtml2pdf',
            status=500
        )

    client = get_object_or_404(ClientProfile, id=client_id)

    # Get all obligations for stats calculation (before slicing)
    all_obligations = MonthlyObligation.objects.filter(client=client)

    # Calculate statistics from unsliced queryset
    stats = {
        'total': all_obligations.count(),
        'completed': all_obligations.filter(status='completed').count(),
        'pending': all_obligations.filter(status='pending').count(),
        'overdue': all_obligations.filter(status='overdue').count(),
    }

    # Get limited obligations for the table (after stats calculation)
    obligations = all_obligations.select_related('obligation_type').order_by('-deadline')[:50]

    # Completion rate
    if stats['total'] > 0:
        stats['completion_rate'] = round((stats['completed'] / stats['total']) * 100, 1)
    else:
        stats['completion_rate'] = 0

    html_content = render_to_string('reports/client_report.html', {
        'client': client,
        'obligations': obligations,
        'stats': stats,
        'generated_at': timezone.now()
    })

    try:
        result = BytesIO()
        pisa_status = pisa.CreatePDF(html_content, dest=result)
        if pisa_status.err:
            return HttpResponse("Error generating PDF", status=500)
        pdf = result.getvalue()
        response = HttpResponse(pdf, content_type='application/pdf')
        safe_name = client.afm.replace(' ', '_')
        response['Content-Disposition'] = f'filename="client_{safe_name}.pdf"'
        logger.info(f"PDF report generated for client {client.afm} by {request.user.username}")
        return response
    except Exception as e:
        logger.error(f"Error generating PDF for client {client_id}: {e}", exc_info=True)
        return HttpResponse(f'Σφάλμα δημιουργίας PDF: {str(e)}', status=500)


@staff_member_required
def monthly_report_pdf(request, year, month):
    """
    Generate PDF report for a specific month.
    Includes all obligations for that period.
    """
    from io import BytesIO
    from django.template.loader import render_to_string
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return HttpResponse(
            'xhtml2pdf δεν είναι εγκατεστημένο. Εγκαταστήστε το με: pip install xhtml2pdf',
            status=500
        )

    # Get all obligations for stats calculation (before any slicing)
    all_obligations = MonthlyObligation.objects.filter(
        year=year,
        month=month
    )

    # Calculate statistics from base queryset
    stats = {
        'total': all_obligations.count(),
        'completed': all_obligations.filter(status='completed').count(),
        'pending': all_obligations.filter(status='pending').count(),
        'overdue': all_obligations.filter(status='overdue').count(),
    }

    # Completion rate
    if stats['total'] > 0:
        stats['completion_rate'] = round((stats['completed'] / stats['total']) * 100, 1)
    else:
        stats['completion_rate'] = 0

    # Group by status (from base queryset)
    by_status = {
        'pending': all_obligations.filter(status='pending').select_related('client', 'obligation_type'),
        'completed': all_obligations.filter(status='completed').select_related('client', 'obligation_type'),
        'overdue': all_obligations.filter(status='overdue').select_related('client', 'obligation_type'),
    }

    # Get obligations for the table
    obligations = all_obligations.select_related('client', 'obligation_type').order_by('deadline', 'client__eponimia')

    # Month names in Greek
    month_names = {
        1: 'Ιανουάριος', 2: 'Φεβρουάριος', 3: 'Μάρτιος',
        4: 'Απρίλιος', 5: 'Μάιος', 6: 'Ιούνιος',
        7: 'Ιούλιος', 8: 'Αύγουστος', 9: 'Σεπτέμβριος',
        10: 'Οκτώβριος', 11: 'Νοέμβριος', 12: 'Δεκέμβριος'
    }
    month_name = month_names.get(month, str(month))

    html_content = render_to_string('reports/monthly_report.html', {
        'year': year,
        'month': month,
        'month_name': month_name,
        'obligations': obligations,
        'stats': stats,
        'by_status': by_status,
        'generated_at': timezone.now()
    })

    try:
        result = BytesIO()
        pisa_status = pisa.CreatePDF(html_content, dest=result)
        if pisa_status.err:
            return HttpResponse("Error generating PDF", status=500)
        pdf = result.getvalue()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'filename="report_{year}_{month:02d}.pdf"'
        logger.info(f"Monthly PDF report generated for {month}/{year} by {request.user.username}")
        return response
    except Exception as e:
        logger.error(f"Error generating monthly PDF for {month}/{year}: {e}", exc_info=True)
        return HttpResponse(f'Σφάλμα δημιουργίας PDF: {str(e)}', status=500)
