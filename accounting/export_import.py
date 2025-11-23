"""
Complete Export/Import utilities για ClientProfile
Centralized module για όλες τις export/import λειτουργίες
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from django.http import HttpResponse
from datetime import datetime
from .models import ClientProfile


def export_clients_to_excel(queryset=None):
    """
    Export πελατών σε Excel με ΟΛΑ τα πεδία (52 πεδία)
    
    Args:
        queryset: ClientProfile queryset (αν None, εξάγει όλους)
    
    Returns:
        HttpResponse με Excel file
    """
    
    if queryset is None:
        clients = ClientProfile.objects.all().order_by('eponimia')
    else:
        clients = queryset.order_by('eponimia')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Πελάτες'
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers - ΟΛΑ ΤΑ ΠΕΔΙΑ (52)
    headers = [
        ('afm', 'Α.Φ.Μ.'),
        ('doy', 'Δ.Ο.Υ.'),
        ('eponimia', 'Επωνυμία/Επώνυμο'),
        ('onoma', 'Όνομα'),
        ('onoma_patros', 'Όνομα Πατρός'),
        ('arithmos_taftotitas', 'Αριθμός Ταυτότητας'),
        ('eidos_taftotitas', 'Είδος Ταυτότητας'),
        ('prosopikos_arithmos', 'Προσωπικός Αριθμός'),
        ('amka', 'Α.Μ.Κ.Α.'),
        ('am_ika', 'Α.Μ. Ι.Κ.Α.'),
        ('arithmos_gemi', 'Αριθμός Γ.Ε.ΜΗ.'),
        ('arithmos_dypa', 'Αριθμός Δ.ΥΠ.Α'),
        ('imerominia_gennisis', 'Ημ. Γέννησης'),
        ('imerominia_gamou', 'Ημ. Γάμου'),
        ('filo', 'Φύλο'),
        ('diefthinsi_katoikias', 'Διεύθυνση Κατοικίας'),
        ('arithmos_katoikias', 'Αριθμός'),
        ('poli_katoikias', 'Πόλη Κατοικίας'),
        ('dimos_katoikias', 'Δήμος Κατοικίας'),
        ('nomos_katoikias', 'Νομός Κατοικίας'),
        ('tk_katoikias', 'T.K. Κατοικίας'),
        ('tilefono_oikias_1', 'Τηλέφωνο Οικίας 1'),
        ('tilefono_oikias_2', 'Τηλέφωνο Οικίας 2'),
        ('kinito_tilefono', 'Κινητό τηλέφωνο'),
        ('diefthinsi_epixeirisis', 'Διεύθυνση Επιχείρησης'),
        ('arithmos_epixeirisis', 'Αριθμός Επιχείρησης'),
        ('poli_epixeirisis', 'Πόλη Επιχείρησης'),
        ('dimos_epixeirisis', 'Δήμος Επιχείρησης'),
        ('nomos_epixeirisis', 'Νομός Επιχείρησης'),
        ('tk_epixeirisis', 'Τ.Κ. Επιχείρησης'),
        ('tilefono_epixeirisis_1', 'Τηλέφωνο Επιχείρησης 1'),
        ('tilefono_epixeirisis_2', 'Τηλέφωνο Επιχείρησης 2'),
        ('email', 'Email'),
        ('trapeza', 'Τράπεζα'),
        ('iban', 'IBAN'),
        ('eidos_ipoxreou', 'Είδος Υπόχρεου'),
        ('katigoria_vivlion', 'Κατηγορία Βιβλίων'),
        ('nomiki_morfi', 'Νομική Μορφή'),
        ('agrotis', 'Αγρότης'),
        ('imerominia_enarksis', 'Ημ/νία Έναρξης Εργασιών'),
        ('onoma_xristi_taxisnet', 'Όνομα Χρήστη Taxis Net'),
        ('kodikos_taxisnet', 'Κωδικός Taxis Net'),
        ('onoma_xristi_ika_ergodoti', 'Όνομα Χρήστη Ι.Κ.Α. Εργοδότη'),
        ('kodikos_ika_ergodoti', 'Κωδικός Ι.Κ.Α. Εργοδότη'),
        ('onoma_xristi_gemi', 'Όνομα Χρήστη Γ.Ε.ΜΗ.'),
        ('kodikos_gemi', 'Κωδικός Γ.Ε.ΜΗ.'),
        ('afm_sizigou', 'Α.Φ.Μ Συζύγου'),
        ('afm_foreas', 'Α.Φ.Μ. Φορέας'),
        ('am_klidi', 'ΑΜ ΚΛΕΙΔΙ'),
        ('is_active', 'Ενεργός'),
        ('created_at', 'Ημ. Δημιουργίας'),
        ('updated_at', 'Ημ. Ενημέρωσης'),
    ]
    
    # Title row
    ws.merge_cells('A1:AX1')
    title_cell = ws['A1']
    title_cell.value = f'ΕΞΑΓΩΓΗ ΠΕΛΑΤΩΝ - {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    title_cell.font = Font(bold=True, size=14, color="667EEA")
    title_cell.alignment = Alignment(horizontal='center')
    
    ws.row_dimensions[2].height = 5
    
    # Headers (Row 3)
    for col_idx, (field, header) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row_idx, client in enumerate(clients, 4):
        for col_idx, (field, header) in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = getattr(client, field, '')
            
            if field == 'eidos_ipoxreou' and value:
                value = client.get_eidos_ipoxreou_display()
            elif field == 'katigoria_vivlion' and value:
                value = client.get_katigoria_vivlion_display()
            elif field == 'filo' and value:
                value = client.get_filo_display()
            elif field in ('agrotis', 'is_active'):
                value = 'ΝΑΙ' if value else 'ΟΧΙ'
            elif field in ('imerominia_gennisis', 'imerominia_gamou', 'imerominia_enarksis') and value:
                value = value.strftime('%d/%m/%Y')
            elif field in ('created_at', 'updated_at') and value:
                value = value.strftime('%d/%m/%Y %H:%M')
            
            cell.value = value or ''
            cell.border = border
            
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Auto-adjust column widths (skip merged cells)
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Summary row
    summary_row = len(clients) + 5
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = f'Σύνολο: {clients.count()} πελάτες'
    summary_cell.font = Font(bold=True, italic=True, size=12, color="667EEA")
    summary_cell.fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    summary_cell.alignment = Alignment(horizontal='left')
    
    ws.freeze_panes = 'A4'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if queryset and queryset.count() < ClientProfile.objects.count():
        filename = f'Clients_Selected_{timestamp}.xlsx'
    else:
        filename = f'Clients_All_{timestamp}.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


def export_clients_summary_to_excel(queryset=None):
    """
    Export συνοπτικής λίστας πελατών (11 basic fields)
    Για γρήγορο export χωρίς όλα τα πεδία
    
    Args:
        queryset: ClientProfile queryset (αν None, εξάγει όλους)
    
    Returns:
        HttpResponse με Excel file
    """
    
    if queryset is None:
        clients = ClientProfile.objects.all().order_by('eponimia')
    else:
        clients = queryset.order_by('eponimia')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Πελάτες - Σύνοψη'
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Basic headers
    headers = [
        ('afm', 'Α.Φ.Μ.'),
        ('eponimia', 'Επωνυμία/Επώνυμο'),
        ('onoma', 'Όνομα'),
        ('doy', 'Δ.Ο.Υ.'),
        ('eidos_ipoxreou', 'Είδος Υπόχρεου'),
        ('katigoria_vivlion', 'Κατηγορία Βιβλίων'),
        ('email', 'Email'),
        ('kinito_tilefono', 'Κινητό'),
        ('tilefono_epixeirisis_1', 'Τηλ. Επιχείρησης'),
        ('is_active', 'Ενεργός'),
        ('created_at', 'Δημιουργήθηκε'),
    ]
    
    # Headers (Row 1)
    for col_idx, (field, header) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row_idx, client in enumerate(clients, 2):
        for col_idx, (field, header) in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = getattr(client, field, '')
            
            if field == 'eidos_ipoxreou' and value:
                value = client.get_eidos_ipoxreou_display()
            elif field == 'katigoria_vivlion' and value:
                value = client.get_katigoria_vivlion_display()
            elif field == 'is_active':
                value = 'ΝΑΙ' if value else 'ΟΧΙ'
            elif field == 'created_at' and value:
                value = value.strftime('%d/%m/%Y')
            
            cell.value = value or ''
            cell.border = border
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Summary row
    summary_row = len(clients) + 3
    ws.cell(row=summary_row, column=1).value = f'Σύνολο: {clients.count()} πελάτες'
    ws.cell(row=summary_row, column=1).font = Font(bold=True, italic=True, color="667EEA")
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'Clients_Summary_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response