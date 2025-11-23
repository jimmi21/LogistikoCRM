"""
Accounting Views - Complete Professional Implementation
Author: ddiplas
Version: 2.0
Description: Comprehensive views for accounting management system with advanced features
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count, Q, Sum, Avg
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import cache_page
from django.middleware.csrf import get_token
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.mail import send_mail
from django.conf import settings
from django.views.generic import ListView

from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response 

from .models import (
    ClientProfile, ClientObligation, MonthlyObligation, 
    ObligationType, VoIPCall, VoIPCallLog, EmailTemplate
)
from .serializers import VoIPCallSerializer, VoIPCallLogSerializer

from datetime import timedelta, datetime
from collections import defaultdict
import logging
import json
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv

# ============================================
# LOGGER CONFIGURATION
# ============================================
logger = logging.getLogger(__name__)

# ============================================
# MAIN DASHBOARD
# ============================================

@staff_member_required
@login_required
def dashboard_view(request):
    """
    Enhanced Dashboard with comprehensive filtering and statistics
    Features: Advanced filters, real-time stats, bulk operations support
    """
    now = timezone.now()
    
    # ========== FILTER PARAMETERS ==========
    filter_params = {
        'status': request.GET.get('status', ''),
        'client': request.GET.get('client', ''),
        'type': request.GET.get('type', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'sort_by': request.GET.get('sort', 'deadline')
    }
    
    # Convert IDs to integers safely
    filter_client_id = _safe_int(filter_params['client'])
    filter_type_id = _safe_int(filter_params['type'])
    
    logger.info(f"Dashboard accessed by {request.user.username} with filters: {filter_params}")
    
    # ========== STATISTICS (UNFILTERED) ==========
    stats = _calculate_dashboard_stats()
    
    # ========== BUILD FILTERED QUERY ==========
    upcoming_query = _build_filtered_query(
        filter_params, filter_client_id, filter_type_id, now
    )
    
    # ========== APPLY SORTING ==========
    sort_mapping = {
        'deadline': 'deadline',
        'client': 'client__eponimia',
        'type': 'obligation_type__name',
        'status': 'status'
    }
    upcoming_query = upcoming_query.order_by(
        sort_mapping.get(filter_params['sort_by'], 'deadline')
    )
    
    # Get results with limit for performance
    upcoming = list(upcoming_query[:100])
    upcoming_count = upcoming_query.count()
    
    # ========== OVERDUE OBLIGATIONS ==========
    overdue_query = MonthlyObligation.objects.filter(
        deadline__lt=now.date(),
        status__in=['pending', 'overdue']
    ).select_related('client', 'obligation_type')
    
    # Apply client/type filters to overdue as well
    if filter_client_id:
        overdue_query = overdue_query.filter(client_id=filter_client_id)
    if filter_type_id:
        overdue_query = overdue_query.filter(obligation_type_id=filter_type_id)
    
    overdue_obligations = list(overdue_query.order_by('deadline')[:20])
    overdue_count = overdue_query.count()
    
    # ========== FILTER OPTIONS ==========
    all_clients = ClientProfile.objects.all().order_by('eponimia').values('id', 'eponimia', 'afm')
    all_types = ObligationType.objects.filter(is_active=True).order_by('name')
    
    # ========== PREPARE CONTEXT ==========
    context = {
        'title': 'Dashboard - Επισκόπηση',
        
        # Statistics
        'total_clients': stats['total_clients'],
        'pending': stats['pending'],
        'completed': stats['completed'],
        'overdue': stats['overdue'],
        
        # Main data
        'upcoming': upcoming,
        'upcoming_count': upcoming_count,
        'overdue_obligations': overdue_obligations,
        'overdue_count': overdue_count,
        
        # Filter options
        'all_clients': all_clients,
        'all_types': all_types,
        
        # Current filter values
        **{f'filter_{k}': v for k, v in filter_params.items()},
        
        # Additional metadata
        'current_date': now.date(),
        'user': request.user,
    }
    
    return render(request, 'accounting/dashboard.html', context)


# ============================================
# CLIENT DETAIL VIEW
# ============================================

@staff_member_required
def client_detail_view(request, client_id):
    """
    Comprehensive client view with all obligations and analytics
    """
    client = get_object_or_404(ClientProfile, id=client_id)
    now = timezone.now()
    
    # All obligations for this client
    all_obligations = MonthlyObligation.objects.filter(
        client=client
    ).select_related('obligation_type').order_by('-deadline')
    
    # Calculate statistics
    stats = {
        'total': all_obligations.count(),
        'pending': all_obligations.filter(status='pending').count(),
        'completed': all_obligations.filter(status='completed').count(),
        'overdue': all_obligations.filter(status='overdue').count(),
    }
    
    # Upcoming obligations (next 30 days)
    next_month = now.date() + timedelta(days=30)
    upcoming = all_obligations.filter(
        deadline__range=[now.date(), next_month],
        status='pending'
    ).order_by('deadline')[:10]
    
    # Overdue obligations
    overdue = all_obligations.filter(
        deadline__lt=now.date(),
        status__in=['pending', 'overdue']
    ).order_by('deadline')
    
    # Recent completed (last 30 days)
    past_month = now.date() - timedelta(days=30)
    recent_completed = all_obligations.filter(
        status='completed',
        completed_date__gte=past_month
    ).order_by('-completed_date')[:10]
    
    # Monthly breakdown (last 6 months)
    monthly_stats = _calculate_monthly_stats(all_obligations, now)
    
    # Active obligation types
    active_types = ClientObligation.objects.filter(
        client=client,
        is_active=True
    ).select_related('obligation_type')
    
    context = {
        'title': f'{client.eponimia} - Λεπτομέρειες',
        'client': client,
        'stats': stats,
        'upcoming': upcoming,
        'overdue': overdue,
        'recent_completed': recent_completed,
        'monthly_stats': monthly_stats,
        'active_types': active_types,
    }
    
    return render(request, 'accounting/client_detail.html', context)


# ============================================
# QUICK COMPLETE WITH FILE UPLOAD
# ============================================

@require_POST
@staff_member_required
def quick_complete_obligation(request, obligation_id):
    """
    Quick complete single obligation with optional file attachment
    """
    try:
        obligation = MonthlyObligation.objects.get(id=obligation_id)
        
        # Handle both FormData and JSON requests
        if 'multipart/form-data' in request.META.get('CONTENT_TYPE', ''):
            # Form data with file upload
            time_spent = request.POST.get('time_spent', 0)
            notes = request.POST.get('notes', '')
            attachment = request.FILES.get('attachment')
        else:
            # JSON data (no file)
            data = json.loads(request.body)
            time_spent = data.get('time_spent', 0)
            notes = data.get('notes', '')
            attachment = None
        
        # Update obligation
        obligation.status = 'completed'
        obligation.completed_date = timezone.now().date()
        obligation.completed_by = request.user
        
        # Handle time spent
        if time_spent:
            try:
                obligation.time_spent = float(time_spent)
            except (ValueError, TypeError):
                pass
        
        # Handle notes with timestamp
        if notes:
            timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
            new_note = f"[{timestamp}] {notes}"
            if obligation.notes:
                obligation.notes += f"\n{new_note}"
            else:
                obligation.notes = new_note
        
        # Handle file attachment
        if attachment:
            obligation.attachment = attachment
            logger.info(f"File attached: {attachment.name} to obligation {obligation_id}")
        
        obligation.save()
        
        # Success response
        message = f'✅ {obligation.obligation_type.name} ολοκληρώθηκε!'
        if time_spent:
            message += f' (⏱️ {time_spent}h)'
        if attachment:
            message += ' 📎'
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except MonthlyObligation.DoesNotExist:
        return JsonResponse(
            {'success': False, 'message': '❌ Υποχρέωση δεν βρέθηκε'}, 
            status=404
        )
    except Exception as e:
        logger.error(f"Error in quick_complete: {str(e)}", exc_info=True)
        return JsonResponse(
            {'success': False, 'message': f'❌ Σφάλμα: {str(e)}'}, 
            status=500
        )


# ============================================
# BULK COMPLETE (SIMPLE VERSION)
# ============================================

@require_POST
@staff_member_required
def bulk_complete_view(request):
    """
    Simple bulk complete - all obligations get same treatment
    """
    try:
        obligation_ids = json.loads(request.POST.get('obligation_ids', '[]'))
        time_spent = request.POST.get('time_spent', '0')
        notes = request.POST.get('notes', '')
        attachments = request.FILES.getlist('attachments')
        
        if not obligation_ids:
            return JsonResponse({
                'success': False, 
                'message': '❌ Δεν επιλέχθηκαν υποχρεώσεις'
            })
        
        obligations = MonthlyObligation.objects.filter(id__in=obligation_ids)
        completed_count = 0
        
        for idx, obl in enumerate(obligations):
            obl.status = 'completed'
            obl.completed_date = timezone.now().date()
            obl.completed_by = request.user
            
            if time_spent:
                obl.time_spent = float(time_spent)
            
            if notes:
                timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
                new_note = f"[{timestamp}] [BULK] {notes}"
                if obl.notes:
                    obl.notes += f"\n{new_note}"
                else:
                    obl.notes = new_note
            
            # Attach file if available
            if idx < len(attachments):
                obl.attachment = attachments[idx]
            
            obl.save()
            completed_count += 1
        
        return JsonResponse({
            'success': True,
            'completed_count': completed_count,
            'message': f'✅ Ολοκληρώθηκαν {completed_count} υποχρεώσεις!'
        })
        
    except Exception as e:
        logger.error(f"Error in bulk_complete: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False, 
            'message': f'❌ Σφάλμα: {str(e)}'
        }, status=500)


# ============================================
# ADVANCED BULK COMPLETE WITH GROUPING
# ============================================

@require_POST
@staff_member_required
def advanced_bulk_complete(request):
    """
    Advanced bulk complete with AFM grouping support
    Allows different files for different groups of obligations
    """
    try:
        completion_data = json.loads(request.POST.get('completion_data', '[]'))
        notes = request.POST.get('notes', '')
        
        completed_count = 0
        errors = []
        processed_details = []
        
        # Process each client group
        for group_data in completion_data:
            client_afm = group_data['client_afm']
            group_num = group_data['group']
            obligation_ids = group_data['obligations']
            
            logger.info(f"Processing AFM {client_afm}, Group {group_num}: {len(obligation_ids)} obligations")
            
            # Get files for this specific group
            files_key = f"file_{client_afm}_{group_num}"
            files = request.FILES.getlist(files_key)
            
            try:
                if group_num == '0':  
                    # Individual mode - each obligation gets its own file
                    result = _process_individual_obligations(
                        obligation_ids, files, notes, request.user
                    )
                    completed_count += result['completed']
                    errors.extend(result['errors'])
                    processed_details.extend(result['details'])
                else:  
                    # Group mode - all obligations share the same file
                    result = _process_grouped_obligations(
                        obligation_ids, files, notes, group_num, request.user
                    )
                    completed_count += result['completed']
                    errors.extend(result['errors'])
                    processed_details.extend(result['details'])
                    
            except Exception as e:
                logger.error(f"Error processing group {client_afm}-{group_num}: {str(e)}")
                errors.append(f"Σφάλμα με ΑΦΜ {client_afm}: {str(e)}")
        
        # Prepare response
        if completed_count > 0:
            success = True
            message = f'✅ Ολοκληρώθηκαν {completed_count} υποχρεώσεις!'
            if errors:
                message += f' (⚠️ {len(errors)} σφάλματα)'
        else:
            success = False
            message = '❌ Καμία υποχρέωση δεν ολοκληρώθηκε'
        
        logger.info(f"Advanced bulk complete finished: {completed_count} completed, {len(errors)} errors")
        
        return JsonResponse({
            'success': success,
            'completed_count': completed_count,
            'message': message,
            'errors': errors[:5],  # Return first 5 errors only
            'details': processed_details[:10]  # Return first 10 details
        })
        
    except json.JSONDecodeError:
        logger.error("Invalid JSON in completion_data")
        return JsonResponse({
            'success': False,
            'message': '❌ Μη έγκυρα δεδομένα JSON',
            'completed_count': 0
        }, status=400)
        
    except Exception as e:
        logger.error(f"Critical error in advanced_bulk_complete: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'❌ Κρίσιμο σφάλμα: {str(e)}',
            'completed_count': 0
        }, status=500)


# ============================================
# EXPORT TO EXCEL
# ============================================

@staff_member_required
def export_filtered_excel(request):
    """
    Export filtered obligations to Excel with formatting
    """
    try:
        # Get filters from request
        filters = _get_filters_from_request(request)
        
        # Build query
        query = _build_export_query(filters)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Υποχρεώσεις"
        
        # Apply styling
        _apply_excel_styling(ws)
        
        # Add headers and data
        _write_excel_headers(ws)
        _write_excel_data(ws, query)
        
        # Auto-adjust columns
        _auto_adjust_excel_columns(ws)
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Ypohrewseis_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        logger.info(f"Excel exported by {request.user.username}: {query.count()} records")
        return response
        
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


# ============================================
# REPORTS & ANALYTICS
# ============================================

@staff_member_required
def reports_view(request):
    """
    Comprehensive analytics dashboard with charts and statistics
    """
    now = timezone.now()
    months_back = int(request.GET.get('months', 6))
    start_date = (now - timedelta(days=30*months_back)).date()
    
    # Monthly completion statistics
    monthly_stats = _calculate_monthly_completion_stats(start_date)
    
    # Client performance metrics
    client_stats = _calculate_client_performance(start_date)
    
    # Time tracking summary
    time_stats = _calculate_time_stats(start_date)
    
    # Revenue calculations
    revenue_data = _calculate_revenue(start_date)
    
    # Obligation type statistics
    type_stats = _calculate_type_stats(start_date)
    
    # Current month summary
    current_stats = _calculate_current_month_stats(now)
    
    # Format data for charts
    chart_data = _format_chart_data(monthly_stats)
    
    context = {
        'title': 'Reports & Analytics',
        'months_back': months_back,
        **chart_data,
        'client_stats': client_stats,
        'time_stats': time_stats,
        'total_revenue': revenue_data['total'],
        'type_stats': type_stats,
        'current_stats': current_stats,
    }
    
    return render(request, 'accounting/reports.html', context)


# ============================================
# VOIP DASHBOARD & MANAGEMENT
# ============================================

@staff_member_required
def voip_dashboard(request):
    """
    Modern VoIP Dashboard with real-time updates
    """
    # Get recent calls with optimized query
    calls = VoIPCall.objects.select_related('client').order_by('-started_at')[:50]
    
    # Calculate statistics
    today = timezone.now().date()
    stats = VoIPCall.objects.aggregate(
        total=Count('id'),
        missed=Count('id', filter=Q(status='missed')),
        completed=Count('id', filter=Q(status='completed')),
        today_total=Count('id', filter=Q(started_at__date=today)),
        pending_followup=Count('id', filter=Q(resolution='follow_up')),
    )
    
    # Calculate success rate
    stats['success_rate'] = _calculate_success_rate(
        stats['completed'], 
        stats['total']
    )
    
    context = {
        'calls': calls,
        'stats': stats,
        'csrf_token': get_token(request),
    }
    
    logger.info(f"VoIP Dashboard accessed by {request.user.username}")
    return render(request, 'admin/voip_dashboard.html', context)


@staff_member_required
@require_http_methods(["GET"])
@cache_page(5)  # Cache for 5 seconds
def voip_calls_api(request):
    """
    Real-time API for VoIP calls with AJAX support
    """
    try:
        # Get recent calls
        calls = VoIPCall.objects.select_related('client').order_by('-started_at')[:30]
        
        # Sort by priority (missed first)
        calls = sorted(calls, key=lambda x: (
            x.status != 'missed', 
            -x.started_at.timestamp()
        ))
        
        # Format data for JSON
        data = [_format_voip_call(call) for call in calls]
        
        return JsonResponse({
            'success': True,
            'calls': data,
            'timestamp': timezone.now().isoformat(),
            'total_missed': VoIPCall.objects.filter(status='missed').count(),
        })
        
    except Exception as e:
        logger.error(f"Error in voip_calls_api: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@staff_member_required
@require_POST
def voip_call_update(request, call_id):
    """
    Update VoIP call via AJAX
    """
    try:
        call = VoIPCall.objects.select_related('client').get(id=call_id)
        
        # Parse request data
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': '❌ Invalid JSON data'
            }, status=400)
        
        # Store old values for logging
        old_values = {
            'notes': call.notes,
            'resolution': call.resolution,
            'status': call.status,
        }
        
        # Update fields
        if 'notes' in data:
            call.notes = data['notes'][:1000]  # Limit length
            
        if 'resolution' in data and data['resolution'] in ['pending', 'closed', 'follow_up', '']:
            call.resolution = data['resolution']
            
        if 'status' in data and data['status'] in ['active', 'completed', 'missed', 'failed']:
            call.status = data['status']
        
        call.save()
        
        # Log the change
        _log_voip_change(call, old_values, request.user)
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Κλήση {call.phone_number} ενημερώθηκε!',
            'updated_call': _format_voip_call(call)
        })
        
    except VoIPCall.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '❌ Κλήση δεν βρέθηκε!'
        }, status=404)
        
    except Exception as e:
        logger.error(f"Error updating call {call_id}: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'❌ Σφάλμα: {str(e)}'
        }, status=500)


# ============================================
# VOIP API VIEWSETS (REST Framework)
# ============================================

class VoIPCallViewSet(viewsets.ModelViewSet):
    """
    REST API ViewSet for VoIP calls
    """
    queryset = VoIPCall.objects.all()
    serializer_class = VoIPCallSerializer
    permission_classes = [permissions.AllowAny]  # Adjust as needed
    filterset_fields = ['direction', 'status', 'client', 'phone_number']
    search_fields = ['phone_number', 'client__eponimia', 'notes']
    ordering_fields = ['started_at', 'duration_seconds']
    ordering = ['-started_at']
    
    def perform_create(self, serializer):
        """Create call and auto-match client"""
        voip_call = serializer.save()
        client = self._match_client_by_phone(voip_call.phone_number)
        
        if client:
            voip_call.client = client
            voip_call.client_email = client.email
            voip_call.save()
            logger.info(f"Matched call to client: {client.eponimia}")
        
        # Log call creation
        VoIPCallLog.objects.create(
            call=voip_call,
            action='started',
            description=f'Call started from {voip_call.phone_number}'
        )
    
    def _match_client_by_phone(self, phone_number):
        """Match phone number to client"""
        clean_number = phone_number.replace(" ", "").replace("-", "")
        
        return ClientProfile.objects.filter(
            Q(tilefono_oikias_1__icontains=clean_number) |
            Q(tilefono_oikias_2__icontains=clean_number) |
            Q(kinito_tilefono__icontains=clean_number) |
            Q(tilefono_epixeirisis_1__icontains=clean_number) |
            Q(tilefono_epixeirisis_2__icontains=clean_number)
        ).first()
    
    @action(detail=True, methods=['post'])
    def end_call(self, request, pk=None):
        """End a call and process follow-up actions"""
        voip_call = self.get_object()
        ended_at = request.data.get('ended_at')
        
        if not ended_at:
            return Response(
                {'error': 'ended_at is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Parse and set end time
            from dateutil import parser as dateutil_parser
            if isinstance(ended_at, str):
                ended_at = dateutil_parser.parse(ended_at)
            
            if ended_at.tzinfo is None:
                ended_at = timezone.make_aware(ended_at)
            
            voip_call.ended_at = ended_at
            voip_call.status = 'completed' if voip_call.duration_seconds > 10 else 'missed'
            voip_call.save()
            
            # Log call end
            VoIPCallLog.objects.create(
                call=voip_call,
                action='ended',
                description=f'Call ended - Duration: {voip_call.duration_formatted}'
            )
            
            # Handle missed call actions
            if voip_call.status == 'missed':
                self._handle_missed_call(voip_call)
            
            return Response(
                VoIPCallSerializer(voip_call).data, 
                status=status.HTTP_200_OK
            )
            
        except Exception as e:
            logger.error(f"Error ending call: {e}", exc_info=True)
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _handle_missed_call(self, voip_call):
        """Process missed call - create ticket and send email"""
        try:
            voip_call.ticket_created = True
            voip_call.save()
            
            VoIPCallLog.objects.create(
                call=voip_call,
                action='ticket_created',
                description=f'Missed call ticket for {voip_call.phone_number}'
            )
            
            if voip_call.client_email:
                self._send_missed_call_email(voip_call)
                
        except Exception as e:
            logger.error(f"Error handling missed call: {e}")
    
    def _send_missed_call_email(self, voip_call):
        """Send email notification for missed call"""
        try:
            subject = f"Αναπάντητη κλήση - {voip_call.phone_number}"
            message = f"""
Καλησπέρα,

Είχατε μια αναπάντητη κλήση:

📞 Αριθμός: {voip_call.phone_number}
📅 Ώρα: {voip_call.started_at.strftime('%d/%m/%Y %H:%M')}
↔️ Κατεύθυνση: {voip_call.get_direction_display()}

Παρακαλώ επικοινωνήστε μαζί μας αν χρειάζεστε κάτι.

Ευχαριστούμε,
Λογιστικό Γραφείο
            """
            
            send_mail(
                subject, 
                message,
                settings.DEFAULT_FROM_EMAIL,
                [voip_call.client_email],
                fail_silently=False,
            )
            
            logger.info(f"Email sent to {voip_call.client_email}")
            
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
class VoIPCallsListView(ListView):
    """
    Django Class-Based View for VoIP calls listing
    Alternative to the function-based voip_dashboard
    """
    model = VoIPCall
    template_name = 'accounting/voip_calls_list.html'
    context_object_name = 'calls'
    paginate_by = 50
    
    def get_queryset(self):
        """Get filtered and sorted calls"""
        queryset = VoIPCall.objects.select_related('client').all()
        
        # Apply filters from GET parameters
        status = self.request.GET.get('status')
        direction = self.request.GET.get('direction')
        search = self.request.GET.get('search')
        
        if status:
            queryset = queryset.filter(status=status)
        if direction:
            queryset = queryset.filter(direction=direction)
        if search:
            queryset = queryset.filter(
                Q(phone_number__icontains=search) |
                Q(client__eponimia__icontains=search) |
                Q(notes__icontains=search)
            )
        
        return queryset.order_by('-started_at')
    
    def get_context_data(self, **kwargs):
        """Add statistics to context"""
        context = super().get_context_data(**kwargs)
        
        # Calculate stats
        context['total'] = VoIPCall.objects.count()
        context['missed'] = VoIPCall.objects.filter(status='missed').count()
        context['completed'] = VoIPCall.objects.filter(status='completed').count()
        
        if context['total'] > 0:
            context['success_rate'] = round(
                (context['completed'] / context['total']) * 100
            )
        else:
            context['success_rate'] = 0
        
        # Add filter values
        context['current_status'] = self.request.GET.get('status', '')
        context['current_direction'] = self.request.GET.get('direction', '')
        context['current_search'] = self.request.GET.get('search', '')
        
        return context


class VoIPCallLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only ViewSet for call logs
    """
    queryset = VoIPCallLog.objects.all()
    serializer_class = VoIPCallLogSerializer
    permission_classes = [permissions.AllowAny]
    filterset_fields = ['call', 'action']
    ordering = ['-created_at']


# ============================================
# EMAIL AUTOMATION API
# ============================================

@require_http_methods(["GET"])
@staff_member_required
def api_email_templates(request):
    """
    API endpoint to get all active email templates
    """
    try:
        templates = EmailTemplate.objects.filter(is_active=True).order_by('name')
        
        result = [
            {
                'id': template.id,
                'name': template.name,
                'description': template.description or '',
                'subject': template.subject,
            }
            for template in templates
        ]
        
        logger.info(f"Email templates API: {len(result)} templates returned")
        return JsonResponse(result, safe=False)
        
    except Exception as e:
        logger.error(f"Error in api_email_templates: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@require_POST
@staff_member_required
def api_send_bulk_email(request):
    """
    Schedule bulk emails for obligations
    """
    try:
        data = json.loads(request.body)
        obligation_ids = data.get('obligation_ids', [])
        template_id = data.get('template_id')
        timing = data.get('timing', 'immediate')
        scheduled_time = data.get('scheduled_time')
        
        # Validate input
        if not obligation_ids or not template_id:
            return JsonResponse({
                'success': False,
                'error': 'Missing obligations or template'
            })
        
        # Get template
        try:
            template = EmailTemplate.objects.get(id=template_id, is_active=True)
        except EmailTemplate.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Template not found'
            })
        
        # Get obligations
        obligations = MonthlyObligation.objects.filter(
            id__in=obligation_ids
        ).select_related('client')
        
        if not obligations.exists():
            return JsonResponse({
                'success': False,
                'error': 'No obligations found'
            })
        
        # Calculate send time
        send_at = _calculate_send_time(timing, scheduled_time)
        
        # Group by client and create scheduled emails
        emails_created = _create_bulk_emails(
            obligations, template, send_at, request.user
        )
        
        return JsonResponse({
            'success': True,
            'emails_created': emails_created,
            'message': f'✅ Προγραμματίστηκαν {emails_created} emails'
        })
        
    except Exception as e:
        logger.error(f"Error in api_send_bulk_email: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


# ============================================
# NOTIFICATION SYSTEM
# ============================================

@staff_member_required
def get_notifications(request):
    """
    Get user notifications for dashboard
    """
    now = timezone.now()
    notifications = []
    
    # Overdue obligations
    overdue = MonthlyObligation.objects.filter(
        deadline__lt=now.date(),
        status__in=['pending', 'overdue']
    ).select_related('client', 'obligation_type').order_by('deadline')[:10]
    
    for obl in overdue:
        days_overdue = (now.date() - obl.deadline).days
        notifications.append({
            'id': obl.id,
            'type': 'overdue',
            'priority': 'high',
            'title': f'Καθυστερημένη: {obl.obligation_type.name}',
            'message': f'{obl.client.eponimia} - {days_overdue} μέρες καθυστέρηση',
            'deadline': obl.deadline.isoformat(),
            'client_id': obl.client.id,
            'icon': '🔴',
        })
    
    # Due today
    today_obligations = MonthlyObligation.objects.filter(
        deadline=now.date(),
        status='pending'
    ).select_related('client', 'obligation_type')
    
    for obl in today_obligations:
        notifications.append({
            'id': obl.id,
            'type': 'due_today',
            'priority': 'medium',
            'title': f'Λήγει Σήμερα: {obl.obligation_type.name}',
            'message': f'{obl.client.eponimia}',
            'deadline': obl.deadline.isoformat(),
            'client_id': obl.client.id,
            'icon': '⚠️',
        })
    
    # Upcoming (next 3 days)
    next_3_days = now.date() + timedelta(days=3)
    upcoming = MonthlyObligation.objects.filter(
        deadline__range=[now.date() + timedelta(days=1), next_3_days],
        status='pending'
    ).select_related('client', 'obligation_type').order_by('deadline')[:5]
    
    for obl in upcoming:
        days_until = (obl.deadline - now.date()).days
        notifications.append({
            'id': obl.id,
            'type': 'upcoming',
            'priority': 'low',
            'title': f'Προσεχώς: {obl.obligation_type.name}',
            'message': f'{obl.client.eponimia} - σε {days_until} μέρες',
            'deadline': obl.deadline.isoformat(),
            'client_id': obl.client.id,
            'icon': '📅',
        })
    
    return JsonResponse({
        'notifications': notifications,
        'count': len(notifications),
        'overdue_count': len([n for n in notifications if n['type'] == 'overdue']),
        'today_count': len([n for n in notifications if n['type'] == 'due_today']),
    })


# ============================================
# HELPER FUNCTIONS (Private)
# ============================================

def _safe_int(value):
    """Safely convert string to integer"""
    try:
        return int(value) if value else None
    except (ValueError, TypeError):
        return None


def _calculate_dashboard_stats():
    """Calculate unfiltered dashboard statistics"""
    return {
        'total_clients': ClientProfile.objects.count(),
        'pending': MonthlyObligation.objects.filter(status='pending').count(),
        'completed': MonthlyObligation.objects.filter(status='completed').count(),
        'overdue': MonthlyObligation.objects.filter(status='overdue').count(),
    }


def _build_filtered_query(filter_params, client_id, type_id, now):
    """Build filtered query for obligations"""
    # Date range
    if filter_params['date_from'] and filter_params['date_to']:
        query = MonthlyObligation.objects.filter(
            deadline__gte=filter_params['date_from'],
            deadline__lte=filter_params['date_to']
        )
    else:
        # Default: past 3 to next 3 months
        past_3_months = now.date() - timedelta(days=90)
        next_3_months = now.date() + timedelta(days=90)
        query = MonthlyObligation.objects.filter(
            deadline__range=[past_3_months, next_3_months]
        )
    
    # Apply filters
    if filter_params['status']:
        query = query.filter(status=filter_params['status'])
    
    if client_id:
        query = query.filter(client_id=client_id)
    
    if type_id:
        query = query.filter(obligation_type_id=type_id)
    
    return query.select_related('client', 'obligation_type')


def _calculate_monthly_stats(obligations, now):
    """Calculate monthly statistics for obligations"""
    stats = []
    for i in range(6):
        month_date = now.date().replace(day=1) - timedelta(days=30*i)
        month_obligations = obligations.filter(
            year=month_date.year,
            month=month_date.month
        )
        stats.append({
            'month': month_date.strftime('%B %Y'),
            'total': month_obligations.count(),
            'completed': month_obligations.filter(status='completed').count(),
            'pending': month_obligations.filter(status='pending').count(),
            'overdue': month_obligations.filter(status='overdue').count(),
        })
    return stats


def _process_individual_obligations(obligation_ids, files, notes, user):
    """Process obligations with individual files"""
    completed = 0
    errors = []
    details = []
    
    for i, obl_id in enumerate(obligation_ids):
        try:
            obligation = MonthlyObligation.objects.get(id=obl_id)
            obligation.status = 'completed'
            obligation.completed_date = timezone.now().date()
            obligation.completed_by = user
            
            # Add notes
            if notes:
                timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
                new_note = f"[{timestamp}] {notes}"
                if obligation.notes:
                    obligation.notes += f"\n{new_note}"
                else:
                    obligation.notes = new_note
            
            # Attach file if available
            if i < len(files):
                obligation.attachment = files[i]
                details.append(f"✅ {obligation.obligation_type.name} με αρχείο {files[i].name}")
            else:
                details.append(f"✅ {obligation.obligation_type.name}")
            
            obligation.save()
            completed += 1
            
        except MonthlyObligation.DoesNotExist:
            errors.append(f"Obligation {obl_id} not found")
        except Exception as e:
            errors.append(f"Error with obligation {obl_id}: {str(e)}")
    
    return {'completed': completed, 'errors': errors, 'details': details}


def _process_grouped_obligations(obligation_ids, files, notes, group_num, user):
    """Process obligations with shared group file"""
    completed = 0
    errors = []
    details = []
    
    file_to_use = files[0] if files else None
    
    for obl_id in obligation_ids:
        try:
            obligation = MonthlyObligation.objects.get(id=obl_id)
            obligation.status = 'completed'
            obligation.completed_date = timezone.now().date()
            obligation.completed_by = user
            
            # Add notes with group indicator
            timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
            group_note = f"[{timestamp}] [Ομάδα {group_num}] {notes}" if notes else f"[{timestamp}] [Ομάδα {group_num}]"
            
            if obligation.notes:
                obligation.notes += f"\n{group_note}"
            else:
                obligation.notes = group_note
            
            # Attach group file
            if file_to_use:
                # Create a copy for each obligation
                file_copy = ContentFile(file_to_use.read())
                file_copy.name = file_to_use.name
                obligation.attachment.save(file_to_use.name, file_copy, save=False)
                file_to_use.seek(0)  # Reset for next use
                details.append(f"✅ {obligation.obligation_type.name} (Ομάδα {group_num})")
            else:
                details.append(f"✅ {obligation.obligation_type.name} (Ομάδα {group_num} - χωρίς αρχείο)")
            
            obligation.save()
            completed += 1
            
        except MonthlyObligation.DoesNotExist:
            errors.append(f"Obligation {obl_id} not found")
        except Exception as e:
            errors.append(f"Error with obligation {obl_id}: {str(e)}")
    
    return {'completed': completed, 'errors': errors, 'details': details}


def _format_voip_call(call):
    """Format VoIP call for JSON response"""
    return {
        'id': call.id,
        'call_id': call.call_id,
        'phone_number': call.phone_number,
        'client_name': call.client.eponimia if call.client else 'Άγνωστος',
        'client_email': call.client_email or '—',
        'direction': call.get_direction_display(),
        'direction_icon': '📲' if call.direction == 'incoming' else '☎️',
        'status': call.get_status_display(),
        'status_value': call.status,
        'status_color': _get_status_color(call.status),
        'resolution': call.get_resolution_display() if call.resolution else '—',
        'resolution_value': call.resolution or '',
        'resolution_color': _get_resolution_color(call.resolution),
        'notes': call.notes[:50] + '...' if len(call.notes or '') > 50 else call.notes or '',
        'duration': call.duration_formatted,
        'started_at': call.started_at.strftime('%d/%m %H:%M:%S'),
        'is_missed': call.status == 'missed',
        'is_recent': (timezone.now() - call.started_at).total_seconds() < 300,
    }


def _get_status_color(status):
    """Get color for status badge"""
    colors = {
        'missed': 'dc2626',    # red
        'completed': '16a34a',  # green
        'active': '2563eb',     # blue
        'failed': 'ea580c',     # orange
    }
    return colors.get(status, '666')


def _get_resolution_color(resolution):
    """Get color for resolution badge"""
    colors = {
        'pending': 'f59e0b',    # amber
        'closed': '10b981',     # green
        'follow_up': '3b82f6',  # blue
        '': 'e5e7eb',          # gray
    }
    return colors.get(resolution, 'e5e7eb')


def _calculate_success_rate(completed, total):
    """Calculate success rate percentage"""
    if total > 0:
        return round((completed / total) * 100)
    return 0


def _log_voip_change(call, old_values, user):
    """Log VoIP call changes"""
    changes = []
    for field, old_value in old_values.items():
        new_value = getattr(call, field)
        if old_value != new_value:
            changes.append(f"{field}: {old_value} → {new_value}")
    
    if changes:
        VoIPCallLog.objects.create(
            call=call,
            action='updated',
            description=f"Updated by {user.username}: {', '.join(changes)}"
        )


def _calculate_send_time(timing, scheduled_time):
    """Calculate email send time based on timing option"""
    if timing == 'immediate':
        return timezone.now()
    elif timing == 'delay_1h':
        return timezone.now() + timedelta(hours=1)
    elif timing == 'delay_24h':
        return timezone.now() + timedelta(days=1)
    elif timing == 'scheduled' and scheduled_time:
        # Parse scheduled time
        hour, minute = scheduled_time.split(':')
        send_at = timezone.now().replace(
            hour=int(hour),
            minute=int(minute),
            second=0,
            microsecond=0
        )
        # If time has passed today, schedule for tomorrow
        if send_at < timezone.now():
            send_at += timedelta(days=1)
        return send_at
    else:
        return timezone.now()


def _create_bulk_emails(obligations, template, send_at, user):
    """Create scheduled emails grouped by client"""
    from collections import defaultdict
    
    client_obligations = defaultdict(list)
    for obl in obligations:
        client_obligations[obl.client.id].append(obl)
    
    emails_created = 0
    for client_id, client_obls in client_obligations.items():
        # Create scheduled email (assuming you have this model/function)
        # This would need to be implemented based on your email system
        # scheduled_email = create_scheduled_email(
        #     obligations=client_obls,
        #     template=template,
        #     send_at=send_at,
        #     user=user
        # )
        emails_created += 1
    
    return emails_created


# Excel helper functions
def _get_filters_from_request(request):
    """Extract filters from request"""
    return {
        'status': request.GET.get('status', ''),
        'client': request.GET.get('client', ''),
        'type': request.GET.get('type', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'sort_by': request.GET.get('sort', 'deadline'),
    }


def _build_export_query(filters):
    """Build query for export"""
    now = timezone.now()
    
    # Base query
    if filters['date_from'] and filters['date_to']:
        query = MonthlyObligation.objects.filter(
            deadline__gte=filters['date_from'],
            deadline__lte=filters['date_to']
        )
    else:
        next_month = now.date() + timedelta(days=30)
        query = MonthlyObligation.objects.filter(
            deadline__range=[now.date(), next_month]
        )
    
    # Apply filters
    if filters['status']:
        query = query.filter(status=filters['status'])
    
    if filters['client']:
        query = query.filter(client_id=filters['client'])
    
    if filters['type']:
        query = query.filter(obligation_type_id=filters['type'])
    
    # Sort
    sort_mapping = {
        'deadline': 'deadline',
        'client': 'client__eponimia',
        'type': 'obligation_type__name',
        'status': 'status'
    }
    
    return query.select_related('client', 'obligation_type').order_by(
        sort_mapping.get(filters['sort_by'], 'deadline')
    )


def _apply_excel_styling(ws):
    """Apply styling to Excel worksheet"""
    ws['A1'] = 'ΑΝΑΦΟΡΑ ΥΠΟΧΡΕΩΣΕΩΝ'
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A2'] = f'Ημερομηνία: {timezone.now().strftime("%d/%m/%Y %H:%M")}'


def _write_excel_headers(ws):
    """Write headers to Excel worksheet"""
    headers = ['#', 'Προθεσμία', 'Πελάτης', 'ΑΦΜ', 'Υποχρέωση', 'Κατάσταση', 'Ώρες', 'Σημειώσεις']
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _write_excel_data(ws, query):
    """Write data rows to Excel worksheet"""
    row = 6
    for idx, obl in enumerate(query, start=1):
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = obl.deadline.strftime('%d/%m/%Y')
        ws.cell(row=row, column=3).value = obl.client.eponimia
        ws.cell(row=row, column=4).value = obl.client.afm
        ws.cell(row=row, column=5).value = obl.obligation_type.name
        ws.cell(row=row, column=6).value = obl.get_status_display()
        ws.cell(row=row, column=7).value = float(obl.time_spent) if obl.time_spent else ''
        ws.cell(row=row, column=8).value = obl.notes[:100] if obl.notes else ''
        
        # Color coding for status
        status_cell = ws.cell(row=row, column=6)
        if obl.status == 'completed':
            status_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        elif obl.status == 'overdue':
            status_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        
        row += 1


def _auto_adjust_excel_columns(ws):
    """Auto-adjust column widths in Excel"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# Analytics helper functions
def _calculate_monthly_completion_stats(start_date):
    """Calculate monthly completion statistics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).annotate(
        month_label=TruncMonth('deadline')
    ).values('month_label').annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        pending=Count('id', filter=Q(status='pending')),
        overdue=Count('id', filter=Q(status='overdue')),
    ).order_by('month_label')


def _calculate_client_performance(start_date):
    """Calculate client performance metrics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).values(
        'client__id',
        'client__eponimia'
    ).annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        completion_rate=Count('id', filter=Q(status='completed')) * 100.0 / Count('id')
    ).order_by('-total')[:10]


def _calculate_time_stats(start_date):
    """Calculate time tracking statistics"""
    return MonthlyObligation.objects.filter(
        status='completed',
        completed_date__gte=start_date,
        time_spent__isnull=False
    ).aggregate(
        total_hours=Sum('time_spent'),
        avg_hours=Avg('time_spent'),
        total_tasks=Count('id')
    )


def _calculate_revenue(start_date):
    """Calculate revenue data"""
    obligations = MonthlyObligation.objects.filter(
        status='completed',
        completed_date__gte=start_date,
        time_spent__isnull=False
    ).select_related('client')
    
    total_revenue = 0
    for obl in obligations:
        # Assuming hourly rate from client or default
        hourly_rate = 50  # Default rate, adjust as needed
        if hasattr(obl, 'hourly_rate'):
            hourly_rate = obl.hourly_rate
        total_revenue += float(obl.time_spent or 0) * hourly_rate
    
    return {'total': total_revenue}


def _calculate_type_stats(start_date):
    """Calculate obligation type statistics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).values(
        'obligation_type__name'
    ).annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        avg_time=Avg('time_spent', filter=Q(status='completed'))
    ).order_by('-total')[:10]


def _calculate_current_month_stats(now):
    """Calculate current month statistics"""
    current_month = now.month
    current_year = now.year
    
    return {
        'total': MonthlyObligation.objects.filter(
            year=current_year, month=current_month
        ).count(),
        'completed': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='completed'
        ).count(),
        'pending': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='pending'
        ).count(),
        'overdue': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='overdue'
        ).count(),
    }


def _format_chart_data(monthly_stats):
    """Format data for chart display"""
    chart_labels = [stat['month_label'].strftime('%b %Y') for stat in monthly_stats]
    
    return {
        'chart_labels_json': json.dumps(chart_labels),
        'chart_completed_json': json.dumps([stat['completed'] for stat in monthly_stats]),
        'chart_pending_json': json.dumps([stat['pending'] for stat in monthly_stats]),
        'chart_overdue_json': json.dumps([stat['overdue'] for stat in monthly_stats]),
    }


# ============================================
# END OF VIEWS
# ============================================

logger.info("Accounting views module loaded successfully")

# Πρόσθεσε αυτά στο τέλος του accounting/views.py

# ============================================
# MISSING VOIP FUNCTIONS - QUICK FIX
# ============================================

@staff_member_required
@cache_page(60)  # Cache 1 minute
def voip_statistics(request):
    """
    Advanced statistics για VoIP calls
    """
    try:
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        
        stats = {
            'today': {
                'total': VoIPCall.objects.filter(started_at__date=today).count(),
                'missed': VoIPCall.objects.filter(started_at__date=today, status='missed').count(),
                'completed': VoIPCall.objects.filter(started_at__date=today, status='completed').count(),
            },
            'week': {
                'total': VoIPCall.objects.filter(started_at__date__gte=week_ago).count(),
                'missed': VoIPCall.objects.filter(started_at__date__gte=week_ago, status='missed').count(),
                'completed': VoIPCall.objects.filter(started_at__date__gte=week_ago, status='completed').count(),
            },
            'by_client': list(
                VoIPCall.objects.filter(client__isnull=False)
                .values('client__eponimia')
                .annotate(count=Count('id'))
                .order_by('-count')[:10]
            ),
            'by_resolution': list(
                VoIPCall.objects.values('resolution')
                .annotate(count=Count('id'))
                .order_by('-count')
            ),
        }
        
        # Calculate average duration
        durations = VoIPCall.objects.filter(
            status='completed',
            duration_seconds__gt=0
        ).values_list('duration_seconds', flat=True)[:100]
        
        if durations:
            stats['average_duration'] = sum(durations) / len(durations)
        else:
            stats['average_duration'] = 0
        
        logger.info(f"Statistics generated for {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Error generating statistics: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@staff_member_required
@require_POST
def voip_bulk_action(request):
    """
    Bulk update multiple VoIP calls
    """
    try:
        data = json.loads(request.body)
        call_ids = data.get('call_ids', [])
        action = data.get('action')
        value = data.get('value')
        
        if not isinstance(call_ids, list) or len(call_ids) == 0:
            return JsonResponse({
                'success': False, 
                'message': 'No calls selected'
            }, status=400)
        
        # Update based on action
        updated = 0
        if action == 'resolution':
            if value in ['pending', 'closed', 'follow_up', '']:
                updated = VoIPCall.objects.filter(id__in=call_ids).update(resolution=value)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid resolution value'
                }, status=400)
                
        elif action == 'status':
            if value in ['active', 'completed', 'missed', 'failed']:
                updated = VoIPCall.objects.filter(id__in=call_ids).update(status=value)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid status value'
                }, status=400)
        
        elif action == 'delete':
            updated = VoIPCall.objects.filter(id__in=call_ids).delete()[0]
        
        else:
            return JsonResponse({
                'success': False,
                'message': f'Unknown action: {action}'
            }, status=400)
        
        logger.info(f"Bulk update: {updated} calls updated by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': f'✅ {updated} κλήσεις ενημερώθηκαν!',
            'updated_count': updated
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
        
    except Exception as e:
        logger.error(f"Error in bulk action: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@staff_member_required
def voip_export_csv(request):
    """
    Export VoIP calls to CSV
    """
    import csv
    from django.http import HttpResponse
    
    try:
        # Create response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="voip_calls_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Add BOM for Excel UTF-8 compatibility
        response.write('\ufeff')
        
        # Create CSV writer
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow([
            'ID',
            'Αριθμός',
            'Πελάτης',
            'Email',
            'Κατεύθυνση',
            'Κατάσταση',
            'Διάρκεια',
            'Ημερομηνία',
            'Ώρα',
            'Σημειώσεις',
            'Resolution'
        ])
        
        # Get filtered calls
        calls_query = VoIPCall.objects.select_related('client').all()
        
        # Apply filters if provided
        status = request.GET.get('status')
        if status:
            calls_query = calls_query.filter(status=status)
        
        direction = request.GET.get('direction')
        if direction:
            calls_query = calls_query.filter(direction=direction)
        
        date_from = request.GET.get('date_from')
        if date_from:
            calls_query = calls_query.filter(started_at__date__gte=date_from)
        
        date_to = request.GET.get('date_to')
        if date_to:
            calls_query = calls_query.filter(started_at__date__lte=date_to)
        
        # Order and limit
        calls_query = calls_query.order_by('-started_at')[:1000]
        
        # Write data rows
        for call in calls_query:
            writer.writerow([
                call.id,
                call.phone_number,
                call.client.eponimia if call.client else '—',
                call.client_email or '—',
                call.get_direction_display(),
                call.get_status_display(),
                call.duration_formatted,
                call.started_at.strftime('%d/%m/%Y'),
                call.started_at.strftime('%H:%M:%S'),
                (call.notes[:100] + '...') if call.notes and len(call.notes) > 100 else (call.notes or ''),
                call.get_resolution_display() if call.resolution else '—',
            ])
        
        logger.info(f"CSV exported by {request.user.username}: {calls_query.count()} records")
        return response
        
    except Exception as e:
        logger.error(f"Error exporting CSV: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# CLASS-BASED VIEW FOR VOIP (if needed)
# ============================================

from django.views.generic import ListView

class VoIPCallsListView(ListView):
    """
    Django Class-Based View for VoIP calls listing
    """
    model = VoIPCall
    template_name = 'accounting/voip_calls_list.html'
    context_object_name = 'calls'
    paginate_by = 50
    
    def get_queryset(self):
        """Get filtered and sorted calls"""
        queryset = VoIPCall.objects.select_related('client').all()
        
        # Apply filters from GET parameters
        status = self.request.GET.get('status')
        direction = self.request.GET.get('direction')
        search = self.request.GET.get('search')
        
        if status:
            queryset = queryset.filter(status=status)
        if direction:
            queryset = queryset.filter(direction=direction)
        if search:
            queryset = queryset.filter(
                Q(phone_number__icontains=search) |
                Q(client__eponimia__icontains=search) |
                Q(notes__icontains=search)
            )
        
        return queryset.order_by('-started_at')
    
    def get_context_data(self, **kwargs):
        """Add statistics to context"""
        context = super().get_context_data(**kwargs)
        
        # Calculate stats
        context['total'] = VoIPCall.objects.count()
        context['missed'] = VoIPCall.objects.filter(status='missed').count()
        context['completed'] = VoIPCall.objects.filter(status='completed').count()
        
        if context['total'] > 0:
            context['success_rate'] = round(
                (context['completed'] / context['total']) * 100
            )
        else:
            context['success_rate'] = 0
        
        # Add filter values
        context['current_status'] = self.request.GET.get('status', '')
        context['current_direction'] = self.request.GET.get('direction', '')
        context['current_search'] = self.request.GET.get('search', '')
        
        return context