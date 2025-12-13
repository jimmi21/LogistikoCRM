# -*- coding: utf-8 -*-
"""
Export/Import API for D.P. Economy
===================================
Handles Excel export and import for clients matching the admin template.
"""
import io
import os
import tempfile
from datetime import datetime

from django.http import HttpResponse
from django.db import transaction
from django.core.management import call_command
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from .models import ClientProfile, ObligationType, ObligationProfile, ClientObligation

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ==============================================================================
# EXCEL TEMPLATE DOWNLOAD
# ==============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_clients_template(request):
    """
    GET /api/v1/export/clients/template/
    Download Excel template for client import (using existing management command).
    """
    if not HAS_OPENPYXL:
        return Response(
            {'error': 'Η βιβλιοθήκη openpyxl δεν είναι εγκατεστημένη.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    try:
        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp_path = tmp.name

        # Use the existing management command
        call_command('create_excel_template', tmp_path)

        # Read and return
        with open(tmp_path, 'rb') as f:
            response = HttpResponse(
                f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            timestamp = datetime.now().strftime('%Y%m%d')
            response['Content-Disposition'] = f'attachment; filename="Template_Pelaton_{timestamp}.xlsx"'

        # Cleanup
        os.unlink(tmp_path)
        return response

    except Exception as e:
        return Response(
            {'error': f'Σφάλμα δημιουργίας template: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==============================================================================
# CLIENT EXPORT TO EXCEL
# ==============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_clients_csv(request):
    """
    GET /api/v1/export/clients/csv/
    Export all clients to Excel file (same format as template).
    """
    if not HAS_OPENPYXL:
        return Response(
            {'error': 'Η βιβλιοθήκη openpyxl δεν είναι εγκατεστημένη.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Πελάτες"

        # Headers matching the template
        headers = [
            'Α.Φ.Μ.',
            'Επωνυμία/Επώνυμο',
            'Όνομα',
            'Όνομα Πατρός',
            'Κινητό τηλέφωνο',
            'Email',
            'Διεύθυνση Επιχείρησης',
            'Πόλη Επιχείρησης',
            'Τ.Κ. Επιχείρησης',
            'Τηλέφωνο Επιχείρησης 1',
            'Δ.Ο.Υ.',
            'Είδος Υπόχρεου',
            'Κατηγορία Βιβλίων',
            'Νομική Μορφή',
            'Α.Μ.Κ.Α.',
            'Αριθμός Γ.Ε.ΜΗ.',
            'IBAN',
            'Ενεργός',
        ]

        # Styling for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        clients = ClientProfile.objects.filter(is_active=True).order_by('eponimia')

        # Map eidos_ipoxreou to Greek
        eidos_map = {
            'individual': 'ΙΔΙΩΤΗΣ',
            'professional': 'ΕΠΑΓΓΕΛΜΑΤΙΑΣ',
            'company': 'ΕΤΑΙΡΕΙΑ',
        }

        # Map katigoria to Greek
        katigoria_map = {
            'A': 'Α',
            'B': 'Β',
            'C': 'Γ',
            'none': 'ΧΩΡΙΣ',
        }

        for row_idx, client in enumerate(clients, start=2):
            ws.cell(row=row_idx, column=1, value=client.afm or '')
            ws.cell(row=row_idx, column=2, value=client.eponimia or '')
            ws.cell(row=row_idx, column=3, value=client.onoma or '')
            ws.cell(row=row_idx, column=4, value=client.onoma_patros or '')
            ws.cell(row=row_idx, column=5, value=client.kinito_tilefono or '')
            ws.cell(row=row_idx, column=6, value=client.email or '')
            ws.cell(row=row_idx, column=7, value=client.diefthinsi_epixeirisis or '')
            ws.cell(row=row_idx, column=8, value=client.poli_epixeirisis or '')
            ws.cell(row=row_idx, column=9, value=client.tk_epixeirisis or '')
            ws.cell(row=row_idx, column=10, value=client.tilefono_epixeirisis_1 or '')
            ws.cell(row=row_idx, column=11, value=client.doy or '')
            ws.cell(row=row_idx, column=12, value=eidos_map.get(client.eidos_ipoxreou, ''))
            ws.cell(row=row_idx, column=13, value=katigoria_map.get(client.katigoria_vivlion, ''))
            ws.cell(row=row_idx, column=14, value=client.nomiki_morfi or '')
            ws.cell(row=row_idx, column=15, value=client.amka or '')
            ws.cell(row=row_idx, column=16, value=client.arithmos_gemi or '')
            ws.cell(row=row_idx, column=17, value=client.iban or '')
            ws.cell(row=row_idx, column=18, value='ΝΑΙ' if client.is_active else 'ΟΧΙ')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value or ''))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Pelates_{timestamp}.xlsx"'

        return response

    except Exception as e:
        return Response(
            {'error': f'Σφάλμα εξαγωγής: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==============================================================================
# CLIENT IMPORT FROM EXCEL
# ==============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_clients_csv(request):
    """
    POST /api/v1/import/clients/csv/
    Import clients from Excel file (same format as template).
    """
    if not HAS_OPENPYXL:
        return Response(
            {'error': 'Η βιβλιοθήκη openpyxl δεν είναι εγκατεστημένη.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    if 'file' not in request.FILES:
        return Response(
            {'error': 'Δεν επιλέχθηκε αρχείο.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    excel_file = request.FILES['file']
    mode = request.data.get('mode', 'skip')  # 'skip' or 'update'

    # Check file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response(
            {'error': 'Μόνο αρχεία Excel (.xlsx, .xls) επιτρέπονται.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Read Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Field mapping - Greek headers to model fields
        field_mapping = {
            'Α.Φ.Μ.': 'afm',
            'Δ.Ο.Υ.': 'doy',
            'Επωνυμία/Επώνυμο': 'eponimia',
            'Όνομα': 'onoma',
            'Όνομα Πατρός': 'onoma_patros',
            'Αριθμός Ταυτότητας': 'arithmos_taftotitas',
            'Είδος Ταυτότητας': 'eidos_taftotitas',
            'Προσωπικός Αριθμός': 'prosopikos_arithmos',
            'Α.Μ.Κ.Α.': 'amka',
            'Α.Μ. Ι.Κ.Α.': 'am_ika',
            'Αριθμός Γ.Ε.ΜΗ.': 'arithmos_gemi',
            'Αριθμός Δ.ΥΠ.Α': 'arithmos_dypa',
            'Φύλο': 'filo',
            'Διεύθυνση Κατοικίας': 'diefthinsi_katoikias',
            'Αριθμός': 'arithmos_katoikias',
            'Πόλη Κατοικίας': 'poli_katoikias',
            'Δήμος Κατοικίας': 'dimos_katoikias',
            'Νομός Κατοικίας': 'nomos_katoikias',
            'T.K. Κατοικίας': 'tk_katoikias',
            'Τηλέφωνο Οικίας 1': 'tilefono_oikias_1',
            'Τηλέφωνο Οικίας 2': 'tilefono_oikias_2',
            'Κινητό τηλέφωνο': 'kinito_tilefono',
            'Διεύθυνση Επιχείρησης': 'diefthinsi_epixeirisis',
            'Αριθμός Επιχείρησης': 'arithmos_epixeirisis',
            'Πόλη Επιχείρησης': 'poli_epixeirisis',
            'Δήμος Επιχείρησης': 'dimos_epixeirisis',
            'Νομός Επιχείρησης': 'nomos_epixeirisis',
            'Τ.Κ. Επιχείρησης': 'tk_epixeirisis',
            'Τηλέφωνο Επιχείρησης 1': 'tilefono_epixeirisis_1',
            'Τηλέφωνο Επιχείρησης 2': 'tilefono_epixeirisis_2',
            'Email': 'email',
            'Τράπεζα': 'trapeza',
            'IBAN': 'iban',
            'Είδος Υπόχρεου': 'eidos_ipoxreou',
            'Κατηγορία Βιβλίων': 'katigoria_vivlion',
            'Νομική Μορφή': 'nomiki_morfi',
            'Αγρότης': 'agrotis',
            'Όνομα Χρήστη Taxis Net': 'onoma_xristi_taxisnet',
            'Κωδικός Taxis Net': 'kodikos_taxisnet',
            'Όνομα Χρήστη Ι.Κ.Α. Εργοδότη': 'onoma_xristi_ika_ergodoti',
            'Κωδικός Ι.Κ.Α. Εργοδότη': 'kodikos_ika_ergodoti',
            'Όνομα Χρήστη Γ.Ε.ΜΗ.': 'onoma_xristi_gemi',
            'Κωδικός Γ.Ε.ΜΗ.': 'kodikos_gemi',
            'Α.Φ.Μ Συζύγου/Μ.Σ.Σ.': 'afm_sizigou',
            'Α.Φ.Μ. Φορέας': 'afm_foreas',
            'ΑΜ ΚΛΕΙΔΙ': 'am_klidi',
        }

        # Read headers from row 1
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            header_name = str(cell.value).strip() if cell.value else None
            if header_name and header_name in field_mapping:
                headers[col_idx] = field_mapping[header_name]

        if not headers:
            return Response(
                {'error': 'Δεν βρέθηκαν έγκυρα headers. Χρησιμοποιήστε το template.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        # Determine start row (skip example rows with AFM starting with 123456)
        start_row = 2
        first_afm_cell = ws.cell(2, 1).value
        if first_afm_cell and str(first_afm_cell).startswith('123456'):
            start_row = 3

        with transaction.atomic():
            for row_num in range(start_row, ws.max_row + 1):
                # Parse row
                row_data = {}
                has_data = False

                for col_idx, field_name in headers.items():
                    value = ws.cell(row_num, col_idx).value
                    if value is not None and str(value).strip():
                        has_data = True

                    # Clean and convert value
                    value = clean_value(value, field_name)
                    if value is not None:
                        row_data[field_name] = value

                if not has_data:
                    continue

                # Validate required fields
                afm = row_data.get('afm', '').strip()
                eponimia = row_data.get('eponimia', '').strip()

                if not afm:
                    errors.append(f'Γραμμή {row_num}: Λείπει το ΑΦΜ')
                    continue

                if len(afm) != 9 or not afm.isdigit():
                    errors.append(f'Γραμμή {row_num}: Μη έγκυρο ΑΦΜ "{afm}"')
                    continue

                if not eponimia:
                    errors.append(f'Γραμμή {row_num}: Λείπει η επωνυμία')
                    continue

                # Default eidos_ipoxreou if not provided
                if 'eidos_ipoxreou' not in row_data or not row_data['eidos_ipoxreou']:
                    row_data['eidos_ipoxreou'] = 'professional'

                # Check if client exists
                existing = ClientProfile.objects.filter(afm=afm).first()

                if existing:
                    if mode == 'update':
                        # Update existing client
                        for field, value in row_data.items():
                            if field != 'afm' and value:
                                setattr(existing, field, value)
                        existing.save()
                        updated_count += 1
                    else:
                        skipped_count += 1
                else:
                    # Create new client
                    row_data['is_active'] = True
                    ClientProfile.objects.create(**row_data)
                    created_count += 1

        return Response({
            'success': True,
            'created_count': created_count,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'errors': errors[:20],
            'message': f'Δημιουργήθηκαν {created_count} πελάτες. Ενημερώθηκαν {updated_count}. Παραλείφθηκαν {skipped_count}.'
        })

    except Exception as e:
        return Response(
            {'error': f'Σφάλμα κατά την ανάγνωση του αρχείου: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


def clean_value(value, field_name):
    """Clean and convert values based on field type."""
    if value is None:
        return None

    # String cleaning
    if isinstance(value, str):
        value = value.strip()
        if value == '' or value.upper() in ['ΚΕΝΟ', 'EMPTY', '-', 'N/A']:
            return None

    # Είδος Υπόχρεου mapping
    if field_name == 'eidos_ipoxreou':
        mapping = {
            'ΙΔΙΩΤΗΣ': 'individual',
            'ΕΠΑΓΓΕΛΜΑΤΙΑΣ': 'professional',
            'ΕΤΑΙΡΕΙΑ': 'company',
            'INDIVIDUAL': 'individual',
            'PROFESSIONAL': 'professional',
            'COMPANY': 'company',
        }
        value_upper = str(value).strip().upper()
        return mapping.get(value_upper, 'professional')

    # Κατηγορία Βιβλίων mapping
    if field_name == 'katigoria_vivlion':
        mapping = {
            'Α': 'A',
            'Β': 'B',
            'Γ': 'C',
            'ΧΩΡΙΣ': 'none',
            'A': 'A',
            'B': 'B',
            'C': 'C',
            'NONE': 'none',
        }
        if value:
            value_clean = str(value).strip().upper()
            return mapping.get(value_clean, '')
        return ''

    # Boolean fields
    if field_name == 'agrotis':
        if isinstance(value, bool):
            return value
        value_upper = str(value).upper()
        return value_upper in ['ΝΑΙ', 'NAI', 'YES', 'TRUE', '1', 'Ν']

    # Φύλο mapping
    if field_name == 'filo':
        if value:
            value_upper = str(value).upper()
            if value_upper in ['Μ', 'M', 'ΑΝΔΡΑΣ', 'MALE', 'MAN']:
                return 'M'
            elif value_upper in ['Γ', 'F', 'ΓΥΝΑΙΚΑ', 'FEMALE', 'WOMAN']:
                return 'F'
        return ''

    # Default: string
    return str(value) if value else ''


# ==============================================================================
# OBLIGATION PROFILES EXPORT (keeping CSV for simplicity)
# ==============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_obligation_profiles_csv(request):
    """Export obligation profiles - kept for compatibility."""
    from django.http import HttpResponse
    import csv

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="obligation_profiles_{timestamp}.csv"'
    response.write('\ufeff')  # BOM

    writer = csv.writer(response)
    writer.writerow(['ID', 'Όνομα', 'Περιγραφή', 'Τύποι Υποχρεώσεων'])

    profiles = ObligationProfile.objects.all().prefetch_related('obligations')
    for profile in profiles:
        type_codes = ', '.join([ot.code for ot in profile.obligations.all()])
        writer.writerow([profile.id, profile.name or '', profile.description or '', type_codes])

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_obligation_types_csv(request):
    """Export obligation types - kept for compatibility."""
    from django.http import HttpResponse
    import csv

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="obligation_types_{timestamp}.csv"'
    response.write('\ufeff')  # BOM

    writer = csv.writer(response)
    writer.writerow(['ID', 'Κωδικός', 'Όνομα', 'Περιγραφή', 'Συχνότητα', 'Ενεργός'])

    types = ObligationType.objects.all().order_by('code')
    for ot in types:
        writer.writerow([
            ot.id, ot.code or '', ot.name or '', ot.description or '',
            ot.frequency or '', 'Ναι' if ot.is_active else 'Όχι'
        ])

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_client_obligations_csv(request):
    """Export client obligation assignments."""
    from django.http import HttpResponse
    import csv

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="client_obligations_{timestamp}.csv"'
    response.write('\ufeff')  # BOM

    writer = csv.writer(response)
    writer.writerow(['ΑΦΜ', 'Επωνυμία', 'Profiles', 'Τύποι'])

    cos = ClientObligation.objects.all().select_related('client').prefetch_related(
        'obligation_profiles', 'obligation_types'
    )
    for co in cos:
        profiles = ', '.join([p.name for p in co.obligation_profiles.all()])
        types = ', '.join([t.code for t in co.obligation_types.all()])
        writer.writerow([co.client.afm or '', co.client.eponimia or '', profiles, types])

    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_client_obligations_csv(request):
    """Import client obligation assignments from CSV."""
    import csv

    if 'file' not in request.FILES:
        return Response({'error': 'Δεν επιλέχθηκε αρχείο.'}, status=status.HTTP_400_BAD_REQUEST)

    csv_file = request.FILES['file']
    mode = request.data.get('mode', 'add')

    if not csv_file.name.endswith('.csv'):
        return Response({'error': 'Μόνο αρχεία CSV επιτρέπονται.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        decoded_file = csv_file.read().decode('utf-8-sig')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)

        profiles_by_name = {p.name.lower(): p for p in ObligationProfile.objects.all()}
        types_by_code = {t.code.lower(): t for t in ObligationType.objects.filter(is_active=True)}

        updated_count = 0
        created_count = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(reader, start=2):
                afm = row.get('ΑΦΜ', '').strip()
                if not afm:
                    continue

                client = ClientProfile.objects.filter(afm=afm).first()
                if not client:
                    errors.append(f'Γραμμή {row_num}: ΑΦΜ {afm} δεν βρέθηκε')
                    continue

                profile_names = [p.strip().lower() for p in row.get('Profiles', '').split(',') if p.strip()]
                type_codes = [t.strip().lower() for t in row.get('Τύποι', '').split(',') if t.strip()]

                client_obligation, created = ClientObligation.objects.get_or_create(
                    client=client, defaults={'is_active': True}
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

                profiles_to_assign = [profiles_by_name[n] for n in profile_names if n in profiles_by_name]
                types_to_assign = [types_by_code[c] for c in type_codes if c in types_by_code]

                if mode == 'replace':
                    client_obligation.obligation_profiles.set(profiles_to_assign)
                    client_obligation.obligation_types.set(types_to_assign)
                else:
                    for p in profiles_to_assign:
                        client_obligation.obligation_profiles.add(p)
                    for t in types_to_assign:
                        client_obligation.obligation_types.add(t)

        return Response({
            'success': True,
            'created_count': created_count,
            'updated_count': updated_count,
            'errors': errors[:20],
            'message': f'Ενημερώθηκαν {updated_count}, δημιουργήθηκαν {created_count}.'
        })

    except Exception as e:
        return Response({'error': f'Σφάλμα: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
