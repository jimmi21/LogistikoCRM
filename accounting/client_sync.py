# accounting/client_sync.py
"""
Unified Export/Import System για ClientProfile
Ίδιο format για export και import!
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime
from .models import ClientProfile
from django.contrib import messages
import os

class ClientExcelHandler:
    """Handler για Export/Import με ΙΔΙΟ format"""
    
    # ΣΤΑΘΕΡΑ headers - ίδια για export & import!
    HEADERS = [
        ('afm', 'Α.Φ.Μ.'),
        ('doy', 'Δ.Ο.Υ.'),
        ('eponimia', 'Επωνυμία/Επώνυμο'),
        ('onoma', 'Όνομα'),
        ('onoma_patros', 'Όνομα Πατρός'),
        ('eidos_ipoxreou', 'Είδος Υπόχρεου'),
        ('katigoria_vivlion', 'Κατηγορία Βιβλίων'),
        ('email', 'Email'),
        ('kinito_tilefono', 'Κινητό τηλέφωνο'),
        ('diefthinsi_epixeirisis', 'Διεύθυνση Επιχείρησης'),
        ('poli_epixeirisis', 'Πόλη Επιχείρησης'),
        ('tk_epixeirisis', 'Τ.Κ. Επιχείρησης'),
        ('iban', 'IBAN'),
        ('is_active', 'Ενεργός'),
        ('tilefono_epixeirisis_1', 'Τηλέφωνο Επιχείρησης'),
        ('nomiki_morfi', 'Νομική Μορφή'),
        ('agrotis', 'Αγρότης'),
    ]
    
    @staticmethod
    def export_all_clients():
        """Export ΟΛΩΝ των πελατών με format για re-import"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Πελάτες"
        
        # Style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        for col_idx, (field, header) in enumerate(ClientExcelHandler.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 20
        
        # Data
        clients = ClientProfile.objects.all().order_by('eponimia')
        
        for row_idx, client in enumerate(clients, 2):
            for col_idx, (field, header) in enumerate(ClientExcelHandler.HEADERS, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = getattr(client, field, '')
                
                # Format για re-import compatibility
                if field == 'eidos_ipoxreou':
                    # Κρατάμε το internal value για εύκολο re-import
                    value = value if value else 'professional'
                elif field == 'katigoria_vivlion':
                    value = value if value else ''
                elif field == 'agrotis':
                    value = 'ΝΑΙ' if value else ''
                elif field == 'is_active':
                    value = 'ΝΑΙ' if value else 'ΟΧΙ'
                
                cell.value = value or ''
                cell.border = border
                
                # Alternate colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Summary
        summary_row = len(clients) + 3
        ws.cell(row=summary_row, column=1).value = f"Σύνολο: {clients.count()} πελάτες"
        ws.cell(row=summary_row, column=1).font = Font(bold=True, italic=True)
        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Clients_Full_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
    
    @staticmethod
    def import_clients(file_path, replace_mode=False, request=None):
        """Import με επιλογή replace ή update"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Create reverse mapping
            header_to_field = {header: field for field, header in ClientExcelHandler.HEADERS}
            
            # Read headers
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                header = str(cell.value).strip() if cell.value else None
                if header in header_to_field:
                    headers[col_idx] = header_to_field[header]
            
            if not headers:
                raise ValueError("Δεν βρέθηκαν έγκυρα headers!")
            
            created = 0
            updated = 0
            replaced = 0
            skipped = 0
            
            # Value mappings
            taxpayer_map = {
                'individual': 'individual',
                'professional': 'professional',
                'company': 'company',
                'ΙΔΙΩΤΗΣ': 'individual',
                'ΕΠΑΓΓΕΛΜΑΤΙΑΣ': 'professional',
                'ΕΤΑΙΡΕΙΑ': 'company',
            }
            
            book_map = {
                'A': 'A', 'B': 'B', 'C': 'C', 
                'none': 'none', '': 'none',
                'Α': 'A', 'Β': 'B', 'Γ': 'C',
                'ΧΩΡΙΣ': 'none',
            }
            
            # Process rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                data = {}
                
                for col_idx, cell in enumerate(row, 1):
                    if col_idx in headers:
                        field = headers[col_idx]
                        value = cell.value
                        
                        if value is not None:
                            value = str(value).strip()
                            
                            # Convert values
                            if field == 'eidos_ipoxreou':
                                value = taxpayer_map.get(value, 'professional')
                            elif field == 'katigoria_vivlion':
                                value = book_map.get(value, '')
                            elif field == 'is_active':
                                value = value.upper() in ['ΝΑΙ', 'YES', '1', 'TRUE']
                            elif field == 'agrotis':
                                value = value.upper() in ['ΝΑΙ', 'YES', '1', 'TRUE']
                            
                            data[field] = value
                
                # Skip if no AFM
                if not data.get('afm'):
                    skipped += 1
                    continue
                
                # Defaults
                if 'eidos_ipoxreou' not in data:
                    data['eidos_ipoxreou'] = 'professional'
                if 'is_active' not in data:
                    data['is_active'] = True
                
                afm = data.pop('afm')
                
                if replace_mode:
                    # REPLACE mode - διαγραφή και νέα εγγραφή
                    ClientProfile.objects.filter(afm=afm).delete()
                    ClientProfile.objects.create(afm=afm, **data)
                    replaced += 1
                else:
                    # UPDATE mode - ενημέρωση ή δημιουργία
                    client, created_flag = ClientProfile.objects.update_or_create(
                        afm=afm,
                        defaults=data
                    )
                    if created_flag:
                        created += 1
                    else:
                        updated += 1
            
            return {
                'success': True,
                'created': created,
                'updated': updated,
                'replaced': replaced,
                'skipped': skipped,
                'total': row_idx - 1,
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }