# -*- coding: utf-8 -*-
"""
accounting/api_reports.py
Author: Claude
Description: REST API views for Reports statistics
"""

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth
from datetime import timedelta
from calendar import monthrange

from .models import ClientProfile, MonthlyObligation
from .utils.report_constants import (
    get_date_range,
    get_previous_period_range,
    GREEK_MONTHS_SHORT,
)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_stats(request):
    """
    GET /api/reports/stats/

    Returns comprehensive statistics for the Reports page.
    Parameters:
    - period: today, week, month, quarter, year, all (default: month)

    Returns:
    - total_clients: Total active clients
    - completed_obligations: Completed obligations in period
    - pending_obligations: Currently pending obligations
    - overdue_obligations: Currently overdue obligations
    - obligations_by_type: Breakdown by obligation type
    - monthly_activity: Monthly completion counts (last 12 months)
    - completion_rate: Percentage of completed vs total
    - comparison: Trend compared to previous period
    """
    period = request.query_params.get('period', 'month')
    today = timezone.now().date()

    # Get date range based on period
    start_date, end_date = get_date_range(period)

    # Total active clients
    total_clients = ClientProfile.objects.filter(is_active=True).count()

    # Base querysets
    all_obligations = MonthlyObligation.objects.all()

    # Completed in period
    completed_qs = all_obligations.filter(status='completed')
    if start_date and end_date:
        completed_qs = completed_qs.filter(
            completed_date__gte=start_date,
            completed_date__lte=end_date
        )
    completed_obligations = completed_qs.count()

    # Pending (current)
    pending_obligations = all_obligations.filter(status='pending').count()

    # Overdue (current - includes pending past deadline)
    overdue_obligations = all_obligations.filter(
        Q(status='overdue') | Q(status='pending', deadline__lt=today)
    ).count()

    # Obligations by type (all time or in period)
    if start_date and end_date:
        type_qs = all_obligations.filter(
            Q(created_at__date__gte=start_date) |
            Q(deadline__gte=start_date, deadline__lte=end_date)
        )
    else:
        type_qs = all_obligations

    obligations_by_type = list(
        type_qs.values('obligation_type__name', 'obligation_type__code')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Monthly activity (last 12 months)
    twelve_months_ago = today.replace(day=1) - timedelta(days=365)
    monthly_activity = list(
        all_obligations.filter(
            status='completed',
            completed_date__gte=twelve_months_ago
        ).annotate(
            month=TruncMonth('completed_date')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')
    )

    # Format monthly activity for frontend
    monthly_data = []

    for item in monthly_activity:
        if item['month']:
            month_idx = item['month'].month - 1
            monthly_data.append({
                'month': GREEK_MONTHS_SHORT[month_idx],
                'month_num': item['month'].month,
                'year': item['month'].year,
                'count': item['count']
            })

    # Fill in missing months with zero
    current_month = today.replace(day=1)
    all_months = []
    for i in range(12):
        month_date = current_month - timedelta(days=30*i)
        month_date = month_date.replace(day=1)
        month_idx = month_date.month - 1

        existing = next(
            (m for m in monthly_data
             if m['month_num'] == month_date.month and m['year'] == month_date.year),
            None
        )

        all_months.append({
            'month': GREEK_MONTHS_SHORT[month_idx],
            'month_num': month_date.month,
            'year': month_date.year,
            'count': existing['count'] if existing else 0
        })

    all_months.reverse()

    # Completion rate
    total_in_period = completed_obligations + pending_obligations + overdue_obligations
    completion_rate = round(
        (completed_obligations / total_in_period * 100) if total_in_period > 0 else 0,
        1
    )

    # Comparison with previous period (for trend indicators)
    comparison = calculate_comparison(period, start_date, end_date)

    return Response({
        'period': period,
        'total_clients': total_clients,
        'completed_obligations': completed_obligations,
        'pending_obligations': pending_obligations,
        'overdue_obligations': overdue_obligations,
        'obligations_by_type': obligations_by_type,
        'monthly_activity': all_months,
        'completion_rate': completion_rate,
        'comparison': comparison,
        'generated_at': timezone.now().isoformat()
    })


def calculate_comparison(period: str, current_start, current_end):
    """
    Calculate comparison with previous period for trend indicators.
    Uses centralized date range utilities.
    """
    # Use centralized utility for previous period calculation
    prev_start, prev_end = get_previous_period_range(period, current_start, current_end)

    if prev_start is None or prev_end is None:
        return {'clients_change': 0, 'completed_change': 0}

    # Previous period stats
    prev_completed = MonthlyObligation.objects.filter(
        status='completed',
        completed_date__gte=prev_start,
        completed_date__lte=prev_end
    ).count()

    # Current period stats (for comparison calculation)
    curr_completed = MonthlyObligation.objects.filter(
        status='completed',
        completed_date__gte=current_start,
        completed_date__lte=current_end
    ).count() if current_start and current_end else 0

    # Calculate percentage changes
    def calc_change(current, previous):
        if previous == 0:
            return 100 if current > 0 else 0
        return round((current - previous) / previous * 100, 1)

    # Client comparison - new clients in period
    new_clients_current = ClientProfile.objects.filter(
        created_at__date__gte=current_start,
        created_at__date__lte=current_end
    ).count() if current_start and current_end else 0

    new_clients_prev = ClientProfile.objects.filter(
        created_at__date__gte=prev_start,
        created_at__date__lte=prev_end
    ).count()

    return {
        'clients_change': calc_change(new_clients_current, new_clients_prev),
        'completed_change': calc_change(curr_completed, prev_completed),
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_export(request):
    """
    GET /api/reports/export/

    Returns metadata about available exports.
    For actual file downloads, use /api/reports/export/download/
    """
    return Response({
        'available_exports': [
            {
                'name': 'Αναφορά πελατών',
                'type': 'clients',
                'description': 'Πλήρης λίστα πελατών με στοιχεία επικοινωνίας',
                'formats': ['xlsx']
            },
            {
                'name': 'Αναφορά υποχρεώσεων',
                'type': 'obligations',
                'description': 'Κατάσταση υποχρεώσεων ανά μήνα',
                'formats': ['xlsx']
            },
            {
                'name': 'Οικονομική αναφορά',
                'type': 'financial',
                'description': 'Έσοδα και στατιστικά χρεώσεων',
                'formats': ['xlsx']
            },
            {
                'name': 'Αναφορά απόδοσης',
                'type': 'performance',
                'description': 'Χρόνοι ολοκλήρωσης και KPIs',
                'formats': ['xlsx']
            }
        ]
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_export_download(request):
    """
    GET /api/reports/export/download/

    Download report files.
    Parameters:
    - type: clients, obligations, financial, performance
    - format: xlsx (default)
    - period: today, week, month, quarter, year, all (default: month)
    """
    from django.http import HttpResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    from .utils.report_constants import (
        get_date_range,
        get_excel_header_style,
        get_excel_border,
        auto_adjust_column_width,
        get_export_filename,
        GREEK_MONTHS_SHORT,
    )

    export_type = request.query_params.get('type', 'obligations')
    export_format = request.query_params.get('format', 'xlsx')
    period = request.query_params.get('period', 'month')

    # Get date range
    start_date, end_date = get_date_range(period)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header styling
    header_style = get_excel_header_style()
    border = get_excel_border()

    if export_type == 'obligations':
        ws.title = 'Υποχρεώσεις'

        # Build query
        query = MonthlyObligation.objects.select_related('client', 'obligation_type')
        if start_date and end_date:
            query = query.filter(deadline__gte=start_date, deadline__lte=end_date)

        # Headers
        headers = ['#', 'Προθεσμία', 'Πελάτης', 'ΑΦΜ', 'Υποχρέωση', 'Κατάσταση', 'Ώρες', 'Σημειώσεις']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = border

        # Data
        for row_idx, obl in enumerate(query.order_by('deadline')[:1000], 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws.cell(row=row_idx, column=2, value=obl.deadline.strftime('%d/%m/%Y') if obl.deadline else '').border = border
            ws.cell(row=row_idx, column=3, value=obl.client.eponimia if obl.client else '').border = border
            ws.cell(row=row_idx, column=4, value=obl.client.afm if obl.client else '').border = border
            ws.cell(row=row_idx, column=5, value=obl.obligation_type.name if obl.obligation_type else '').border = border
            ws.cell(row=row_idx, column=6, value=obl.get_status_display()).border = border
            ws.cell(row=row_idx, column=7, value=float(obl.time_spent) if obl.time_spent else '').border = border
            ws.cell(row=row_idx, column=8, value=(obl.notes or '')[:100]).border = border

        filename = get_export_filename(f'Ypoxreoseis_{period}', 'xlsx')

    elif export_type == 'financial':
        ws.title = 'Οικονομικά'

        # Financial report - hours and estimated revenue by client
        from django.db.models import Sum, Count

        query = MonthlyObligation.objects.filter(status='completed')
        if start_date and end_date:
            query = query.filter(completed_date__gte=start_date, completed_date__lte=end_date)

        client_stats = query.values(
            'client__eponimia', 'client__afm'
        ).annotate(
            total_hours=Sum('time_spent'),
            total_obligations=Count('id')
        ).order_by('-total_hours')

        # Headers
        headers = ['#', 'Πελάτης', 'ΑΦΜ', 'Υποχρεώσεις', 'Ώρες', 'Εκτ. Έσοδα (€50/ώρα)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = border

        total_hours = 0
        for row_idx, stat in enumerate(client_stats[:500], 2):
            hours = float(stat['total_hours'] or 0)
            total_hours += hours
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws.cell(row=row_idx, column=2, value=stat['client__eponimia'] or '').border = border
            ws.cell(row=row_idx, column=3, value=stat['client__afm'] or '').border = border
            ws.cell(row=row_idx, column=4, value=stat['total_obligations']).border = border
            ws.cell(row=row_idx, column=5, value=hours).border = border
            ws.cell(row=row_idx, column=6, value=hours * 50).border = border

        # Summary row
        summary_row = client_stats.count() + 2
        ws.cell(row=summary_row, column=1, value='ΣΥΝΟΛΟ').font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=total_hours).font = Font(bold=True)
        ws.cell(row=summary_row, column=6, value=total_hours * 50).font = Font(bold=True)

        filename = get_export_filename(f'Oikonomika_{period}', 'xlsx')

    elif export_type == 'performance':
        ws.title = 'Απόδοση'

        # Performance report - completion rates by obligation type
        from django.db.models import Count, Avg

        query = MonthlyObligation.objects.all()
        if start_date and end_date:
            query = query.filter(deadline__gte=start_date, deadline__lte=end_date)

        type_stats = query.values(
            'obligation_type__name', 'obligation_type__code'
        ).annotate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed')),
            avg_time=Avg('time_spent', filter=Q(status='completed'))
        ).order_by('-total')

        # Headers
        headers = ['#', 'Τύπος Υποχρέωσης', 'Κωδικός', 'Σύνολο', 'Ολοκληρωμένες', 'Ποσοστό', 'Μ.Ο. Ώρες']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = border

        for row_idx, stat in enumerate(type_stats, 2):
            total = stat['total']
            completed = stat['completed']
            rate = round((completed / total * 100), 1) if total > 0 else 0

            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws.cell(row=row_idx, column=2, value=stat['obligation_type__name'] or '').border = border
            ws.cell(row=row_idx, column=3, value=stat['obligation_type__code'] or '').border = border
            ws.cell(row=row_idx, column=4, value=total).border = border
            ws.cell(row=row_idx, column=5, value=completed).border = border
            ws.cell(row=row_idx, column=6, value=f'{rate}%').border = border
            ws.cell(row=row_idx, column=7, value=round(float(stat['avg_time'] or 0), 2)).border = border

        filename = get_export_filename(f'Apodosi_{period}', 'xlsx')

    else:
        return Response({'error': f'Unknown export type: {export_type}'}, status=400)

    # Auto-adjust columns
    auto_adjust_column_width(ws)

    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def client_statement(request, client_id):
    """
    GET /api/reports/client-statement/<client_id>/

    Generate a comprehensive client statement with all obligations.
    Parameters:
    - format: xlsx, pdf (default: xlsx)
    """
    from django.http import HttpResponse
    from django.shortcuts import get_object_or_404
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from .utils.report_constants import (
        get_excel_header_style,
        get_excel_border,
        auto_adjust_column_width,
        get_export_filename,
        BRAND_COLOR,
    )

    export_format = request.query_params.get('format', 'xlsx')
    client = get_object_or_404(ClientProfile, id=client_id)

    if export_format == 'pdf':
        # Redirect to existing PDF endpoint
        from django.shortcuts import redirect
        return redirect('client_report_pdf', client_id=client_id)

    # Excel format
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Κατάσταση Πελάτη'

    header_style = get_excel_header_style()
    border = get_excel_border()

    # Client info section
    ws.merge_cells('A1:H1')
    ws['A1'] = f'ΚΑΡΤΕΛΑ ΠΕΛΑΤΗ: {client.eponimia}'
    ws['A1'].font = Font(bold=True, size=16, color=BRAND_COLOR)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Client details
    details = [
        ('ΑΦΜ:', client.afm),
        ('ΔΟΥ:', client.doy or '-'),
        ('Email:', client.email or '-'),
        ('Τηλέφωνο:', client.kinito_tilefono or client.tilefono_epixeirisis_1 or '-'),
    ]

    for idx, (label, value) in enumerate(details, 3):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)

    # Obligations section
    start_row = 8
    ws.cell(row=start_row, column=1, value='ΥΠΟΧΡΕΩΣΕΙΣ').font = Font(bold=True, size=14, color=BRAND_COLOR)

    # Headers
    headers = ['#', 'Έτος', 'Μήνας', 'Τύπος', 'Προθεσμία', 'Κατάσταση', 'Ημ. Ολοκλήρωσης', 'Ώρες']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row + 1, column=col, value=header)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['alignment']
        cell.border = border

    # Obligations data
    obligations = MonthlyObligation.objects.filter(
        client=client
    ).select_related('obligation_type').order_by('-year', '-month', 'deadline')

    for row_idx, obl in enumerate(obligations, start_row + 2):
        ws.cell(row=row_idx, column=1, value=row_idx - start_row - 1).border = border
        ws.cell(row=row_idx, column=2, value=obl.year).border = border
        ws.cell(row=row_idx, column=3, value=obl.month).border = border
        ws.cell(row=row_idx, column=4, value=obl.obligation_type.name if obl.obligation_type else '').border = border
        ws.cell(row=row_idx, column=5, value=obl.deadline.strftime('%d/%m/%Y') if obl.deadline else '').border = border

        status_cell = ws.cell(row=row_idx, column=6, value=obl.get_status_display())
        status_cell.border = border
        if obl.status == 'completed':
            status_cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
        elif obl.status == 'overdue':
            status_cell.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')

        ws.cell(row=row_idx, column=7, value=obl.completed_date.strftime('%d/%m/%Y') if obl.completed_date else '').border = border
        ws.cell(row=row_idx, column=8, value=float(obl.time_spent) if obl.time_spent else '').border = border

    # Summary
    summary_row = start_row + obligations.count() + 3
    total = obligations.count()
    completed = obligations.filter(status='completed').count()
    rate = round((completed / total * 100), 1) if total > 0 else 0

    ws.cell(row=summary_row, column=1, value='ΣΥΝΟΨΗ:').font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f'Σύνολο: {total}')
    ws.cell(row=summary_row, column=3, value=f'Ολοκληρωμένες: {completed}')
    ws.cell(row=summary_row, column=4, value=f'Ποσοστό: {rate}%')

    auto_adjust_column_width(ws)

    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = client.afm.replace(' ', '_')
    filename = get_export_filename(f'Kartela_{safe_name}', 'xlsx')

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def vat_summary(request):
    """
    GET /api/reports/vat-summary/

    Generate VAT period summary report.
    Parameters:
    - year: Year (default: current year)
    - period_type: 'month' or 'quarter' (default: month)
    - period: Month (1-12) or Quarter (1-4) depending on period_type
    - client_id: Optional - filter by specific client
    - format: 'json' or 'xlsx' (default: json)
    """
    from django.http import HttpResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from decimal import Decimal
    from .utils.report_constants import (
        get_excel_header_style,
        get_excel_border,
        auto_adjust_column_width,
        get_export_filename,
        GREEK_MONTHS_FULL,
        BRAND_COLOR,
    )

    # Parse parameters
    current_date = timezone.now()
    year = int(request.query_params.get('year', current_date.year))
    period_type = request.query_params.get('period_type', 'month')
    period = request.query_params.get('period')
    client_id = request.query_params.get('client_id')
    export_format = request.query_params.get('format', 'json')

    # Default period to current month/quarter
    if period is None:
        if period_type == 'quarter':
            period = (current_date.month - 1) // 3 + 1
        else:
            period = current_date.month
    period = int(period)

    # Determine months to include
    if period_type == 'quarter':
        start_month = (period - 1) * 3 + 1
        end_month = start_month + 2
        months = list(range(start_month, end_month + 1))
        period_label = f'{period}ο Τρίμηνο {year}'
    else:
        months = [period]
        period_label = f'{GREEK_MONTHS_FULL.get(period, period)} {year}'

    # Try to get data from myDATA VATPeriodResult if available
    vat_data_from_mydata = []
    try:
        from mydata.models import VATPeriodResult, VATRecord

        # Get VATPeriodResults for the period
        vat_periods = VATPeriodResult.objects.filter(
            year=year,
            period_type='M' if period_type == 'month' else 'Q'
        )
        if period_type == 'month':
            vat_periods = vat_periods.filter(period__in=months)
        else:
            vat_periods = vat_periods.filter(period=period)

        if client_id:
            vat_periods = vat_periods.filter(client_id=client_id)

        for vp in vat_periods.select_related('client'):
            vat_data_from_mydata.append({
                'client_id': vp.client_id,
                'client_name': vp.client.eponimia if vp.client else 'N/A',
                'client_afm': vp.client.afm if vp.client else 'N/A',
                'vat_output': float(vp.vat_output or 0),
                'vat_input': float(vp.vat_input or 0),
                'vat_balance': float(vp.vat_difference or 0),
                'source': 'myDATA'
            })
    except ImportError:
        pass  # myDATA app not available
    except Exception:
        pass  # myDATA models might not be configured

    # Build summary from obligations data (as fallback or supplement)
    vat_obligations = MonthlyObligation.objects.filter(
        year=year,
        month__in=months,
        obligation_type__code__icontains='ΦΠΑ'
    ).select_related('client', 'obligation_type')

    if client_id:
        vat_obligations = vat_obligations.filter(client_id=client_id)

    # Group by client
    client_data = {}
    for obl in vat_obligations:
        if obl.client_id not in client_data:
            client_data[obl.client_id] = {
                'client_id': obl.client_id,
                'client_name': obl.client.eponimia if obl.client else 'N/A',
                'client_afm': obl.client.afm if obl.client else 'N/A',
                'obligations': [],
                'total_obligations': 0,
                'completed': 0,
                'pending': 0,
                'overdue': 0,
            }

        client_data[obl.client_id]['obligations'].append({
            'month': obl.month,
            'type': obl.obligation_type.name if obl.obligation_type else 'N/A',
            'status': obl.status,
            'deadline': obl.deadline.isoformat() if obl.deadline else None,
            'completed_date': obl.completed_date.isoformat() if obl.completed_date else None,
        })
        client_data[obl.client_id]['total_obligations'] += 1
        if obl.status == 'completed':
            client_data[obl.client_id]['completed'] += 1
        elif obl.status == 'overdue':
            client_data[obl.client_id]['overdue'] += 1
        else:
            client_data[obl.client_id]['pending'] += 1

    # Merge myDATA data if available
    for mydata_item in vat_data_from_mydata:
        cid = mydata_item['client_id']
        if cid in client_data:
            client_data[cid]['vat_output'] = mydata_item['vat_output']
            client_data[cid]['vat_input'] = mydata_item['vat_input']
            client_data[cid]['vat_balance'] = mydata_item['vat_balance']
            client_data[cid]['has_mydata'] = True
        else:
            client_data[cid] = {
                **mydata_item,
                'obligations': [],
                'total_obligations': 0,
                'completed': 0,
                'pending': 0,
                'overdue': 0,
                'has_mydata': True,
            }

    summary_data = list(client_data.values())

    # Calculate totals
    totals = {
        'total_clients': len(summary_data),
        'total_obligations': sum(c['total_obligations'] for c in summary_data),
        'completed': sum(c['completed'] for c in summary_data),
        'pending': sum(c['pending'] for c in summary_data),
        'overdue': sum(c['overdue'] for c in summary_data),
        'vat_output': sum(c.get('vat_output', 0) for c in summary_data),
        'vat_input': sum(c.get('vat_input', 0) for c in summary_data),
        'vat_balance': sum(c.get('vat_balance', 0) for c in summary_data),
    }

    if export_format == 'xlsx':
        # Generate Excel export
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ΦΠΑ Περίοδος'

        header_style = get_excel_header_style()
        border = get_excel_border()

        # Title
        ws.merge_cells('A1:I1')
        ws['A1'] = f'ΑΝΑΦΟΡΑ ΦΠΑ - {period_label}'
        ws['A1'].font = Font(bold=True, size=16, color=BRAND_COLOR)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['#', 'Πελάτης', 'ΑΦΜ', 'Υποχρεώσεις', 'Ολοκληρ.', 'Εκκρεμούν', 'ΦΠΑ Εκροών', 'ΦΠΑ Εισροών', 'Διαφορά']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = border

        # Data rows
        for row_idx, client in enumerate(sorted(summary_data, key=lambda x: x['client_name']), 4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3).border = border
            ws.cell(row=row_idx, column=2, value=client['client_name']).border = border
            ws.cell(row=row_idx, column=3, value=client['client_afm']).border = border
            ws.cell(row=row_idx, column=4, value=client['total_obligations']).border = border
            ws.cell(row=row_idx, column=5, value=client['completed']).border = border
            ws.cell(row=row_idx, column=6, value=client['pending'] + client['overdue']).border = border
            ws.cell(row=row_idx, column=7, value=client.get('vat_output', '-')).border = border
            ws.cell(row=row_idx, column=8, value=client.get('vat_input', '-')).border = border
            ws.cell(row=row_idx, column=9, value=client.get('vat_balance', '-')).border = border

        # Totals row
        total_row = len(summary_data) + 4
        ws.cell(row=total_row, column=1, value='ΣΥΝΟΛΟ').font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=totals['total_obligations']).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=totals['completed']).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=totals['pending'] + totals['overdue']).font = Font(bold=True)
        if totals['vat_output'] > 0:
            ws.cell(row=total_row, column=7, value=totals['vat_output']).font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=totals['vat_input']).font = Font(bold=True)
            ws.cell(row=total_row, column=9, value=totals['vat_balance']).font = Font(bold=True)

        auto_adjust_column_width(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = get_export_filename(f'FPA_{year}_{period}', 'xlsx')
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # JSON response
    return Response({
        'period': {
            'year': year,
            'type': period_type,
            'period': period,
            'label': period_label,
            'months': months,
        },
        'totals': totals,
        'clients': summary_data,
        'generated_at': timezone.now().isoformat(),
    })
