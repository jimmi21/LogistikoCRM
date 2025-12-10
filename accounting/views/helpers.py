# accounting/views/helpers.py
"""
Helper Functions for Accounting Views

This module contains private helper functions used across the accounting views.
All functions are prefixed with underscore to indicate they are internal.
"""

from django.utils import timezone
from django.db.models import Count, Q, Sum, Avg
from django.db.models.functions import TruncMonth
from django.core.files.base import ContentFile
from openpyxl.styles import Font, Alignment, PatternFill

from datetime import timedelta
from decimal import Decimal
from collections import defaultdict
import json

from ..models import (
    ClientProfile, MonthlyObligation, VoIPCallLog
)


# ============================================
# GENERAL UTILITIES
# ============================================

def _safe_int(value):
    """Safely convert string to integer"""
    try:
        return int(value) if value else None
    except (ValueError, TypeError):
        return None


# ============================================
# DASHBOARD HELPERS
# ============================================

def _calculate_dashboard_stats():
    """Calculate unfiltered dashboard statistics"""
    return {
        'total_clients': ClientProfile.objects.count(),
        'pending': MonthlyObligation.objects.filter(status='pending').count(),
        'completed': MonthlyObligation.objects.filter(status='completed').count(),
        'overdue': MonthlyObligation.objects.filter(status='overdue').count(),
    }


def _build_filtered_query(filter_params, client_id, type_id, now):
    """Build filtered query for obligations"""
    # Date range
    if filter_params['date_from'] and filter_params['date_to']:
        query = MonthlyObligation.objects.filter(
            deadline__gte=filter_params['date_from'],
            deadline__lte=filter_params['date_to']
        )
    else:
        # Default: past 3 to next 3 months
        past_3_months = now.date() - timedelta(days=90)
        next_3_months = now.date() + timedelta(days=90)
        query = MonthlyObligation.objects.filter(
            deadline__range=[past_3_months, next_3_months]
        )

    # Apply filters
    if filter_params['status']:
        query = query.filter(status=filter_params['status'])

    if client_id:
        query = query.filter(client_id=client_id)

    if type_id:
        query = query.filter(obligation_type_id=type_id)

    return query.select_related('client', 'obligation_type')


def _calculate_monthly_stats(obligations, now):
    """Calculate monthly statistics for obligations"""
    stats = []
    for i in range(6):
        month_date = now.date().replace(day=1) - timedelta(days=30*i)
        month_obligations = obligations.filter(
            year=month_date.year,
            month=month_date.month
        )
        stats.append({
            'month': month_date.strftime('%B %Y'),
            'total': month_obligations.count(),
            'completed': month_obligations.filter(status='completed').count(),
            'pending': month_obligations.filter(status='pending').count(),
            'overdue': month_obligations.filter(status='overdue').count(),
        })
    return stats


# ============================================
# OBLIGATION PROCESSING HELPERS
# ============================================

def _process_individual_obligations(obligation_ids, files, notes, user):
    """Process obligations with individual files"""
    completed = 0
    errors = []
    details = []

    for i, obl_id in enumerate(obligation_ids):
        try:
            obligation = MonthlyObligation.objects.get(id=obl_id)
            obligation.status = 'completed'
            obligation.completed_date = timezone.now().date()
            obligation.completed_by = user

            # Add notes
            if notes:
                timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
                new_note = f"[{timestamp}] {notes}"
                if obligation.notes:
                    obligation.notes += f"\n{new_note}"
                else:
                    obligation.notes = new_note

            # Attach file if available
            if i < len(files):
                obligation.attachment = files[i]
                details.append(f"✅ {obligation.obligation_type.name} με αρχείο {files[i].name}")
            else:
                details.append(f"✅ {obligation.obligation_type.name}")

            obligation.save()
            completed += 1

        except MonthlyObligation.DoesNotExist:
            errors.append(f"Obligation {obl_id} not found")
        except Exception as e:
            errors.append(f"Error with obligation {obl_id}: {str(e)}")

    return {'completed': completed, 'errors': errors, 'details': details}


def _process_grouped_obligations(obligation_ids, files, notes, group_num, user):
    """Process obligations with shared group file"""
    completed = 0
    errors = []
    details = []

    file_to_use = files[0] if files else None

    for obl_id in obligation_ids:
        try:
            obligation = MonthlyObligation.objects.get(id=obl_id)
            obligation.status = 'completed'
            obligation.completed_date = timezone.now().date()
            obligation.completed_by = user

            # Add notes with group indicator
            timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
            group_note = f"[{timestamp}] [Ομάδα {group_num}] {notes}" if notes else f"[{timestamp}] [Ομάδα {group_num}]"

            if obligation.notes:
                obligation.notes += f"\n{group_note}"
            else:
                obligation.notes = group_note

            # Attach group file
            if file_to_use:
                # Create a copy for each obligation
                file_copy = ContentFile(file_to_use.read())
                file_copy.name = file_to_use.name
                obligation.attachment.save(file_to_use.name, file_copy, save=False)
                file_to_use.seek(0)  # Reset for next use
                details.append(f"✅ {obligation.obligation_type.name} (Ομάδα {group_num})")
            else:
                details.append(f"✅ {obligation.obligation_type.name} (Ομάδα {group_num} - χωρίς αρχείο)")

            obligation.save()
            completed += 1

        except MonthlyObligation.DoesNotExist:
            errors.append(f"Obligation {obl_id} not found")
        except Exception as e:
            errors.append(f"Error with obligation {obl_id}: {str(e)}")

    return {'completed': completed, 'errors': errors, 'details': details}


# ============================================
# VOIP HELPERS
# ============================================

def _match_client_by_phone_standalone(phone_number):
    """Match phone number to client (standalone function for webhook)"""
    clean_number = phone_number.replace(" ", "").replace("-", "")

    return ClientProfile.objects.filter(
        Q(tilefono_oikias_1__icontains=clean_number) |
        Q(tilefono_oikias_2__icontains=clean_number) |
        Q(kinito_tilefono__icontains=clean_number) |
        Q(tilefono_epixeirisis_1__icontains=clean_number) |
        Q(tilefono_epixeirisis_2__icontains=clean_number)
    ).first()


def _format_voip_call(call):
    """Format VoIP call for JSON response"""
    return {
        'id': call.id,
        'call_id': call.call_id,
        'phone_number': call.phone_number,
        'client_name': call.client.eponimia if call.client else 'Άγνωστος',
        'client_email': call.client_email or '—',
        'direction': call.get_direction_display(),
        'direction_icon': '📲' if call.direction == 'incoming' else '☎️',
        'status': call.get_status_display(),
        'status_value': call.status,
        'status_color': _get_status_color(call.status),
        'resolution': call.get_resolution_display() if call.resolution else '—',
        'resolution_value': call.resolution or '',
        'resolution_color': _get_resolution_color(call.resolution),
        'notes': call.notes[:50] + '...' if len(call.notes or '') > 50 else call.notes or '',
        'duration': call.duration_formatted,
        'started_at': call.started_at.strftime('%d/%m %H:%M:%S'),
        'is_missed': call.status == 'missed',
        'is_recent': (timezone.now() - call.started_at).total_seconds() < 300,
    }


def _get_status_color(status):
    """Get color for status badge"""
    colors = {
        'missed': 'dc2626',    # red
        'completed': '16a34a',  # green
        'active': '2563eb',     # blue
        'failed': 'ea580c',     # orange
    }
    return colors.get(status, '666')


def _get_resolution_color(resolution):
    """Get color for resolution badge"""
    colors = {
        'pending': 'f59e0b',    # amber
        'closed': '10b981',     # green
        'follow_up': '3b82f6',  # blue
        '': 'e5e7eb',          # gray
    }
    return colors.get(resolution, 'e5e7eb')


def _calculate_success_rate(completed, total):
    """Calculate success rate percentage"""
    if total > 0:
        return round((completed / total) * 100)
    return 0


def _log_voip_change(call, old_values, user):
    """Log VoIP call changes"""
    changes = []
    for field, old_value in old_values.items():
        new_value = getattr(call, field)
        if old_value != new_value:
            changes.append(f"{field}: {old_value} → {new_value}")

    if changes:
        VoIPCallLog.objects.create(
            call=call,
            action='updated',
            description=f"Updated by {user.username}: {', '.join(changes)}"
        )


# ============================================
# EMAIL HELPERS
# ============================================

def _calculate_send_time(timing, scheduled_time):
    """Calculate email send time based on timing option"""
    if timing == 'immediate':
        return timezone.now()
    elif timing == 'delay_1h':
        return timezone.now() + timedelta(hours=1)
    elif timing == 'delay_24h':
        return timezone.now() + timedelta(days=1)
    elif timing == 'scheduled' and scheduled_time:
        # Parse scheduled time
        hour, minute = scheduled_time.split(':')
        send_at = timezone.now().replace(
            hour=int(hour),
            minute=int(minute),
            second=0,
            microsecond=0
        )
        # If time has passed today, schedule for tomorrow
        if send_at < timezone.now():
            send_at += timedelta(days=1)
        return send_at
    else:
        return timezone.now()


def _create_bulk_emails(obligations, template, send_at, user):
    """Create scheduled emails grouped by client"""
    client_obligations = defaultdict(list)
    for obl in obligations:
        client_obligations[obl.client.id].append(obl)

    emails_created = 0
    for client_id, client_obls in client_obligations.items():
        # Create scheduled email (assuming you have this model/function)
        # This would need to be implemented based on your email system
        # scheduled_email = create_scheduled_email(
        #     obligations=client_obls,
        #     template=template,
        #     send_at=send_at,
        #     user=user
        # )
        emails_created += 1

    return emails_created


# ============================================
# EXCEL EXPORT HELPERS
# ============================================

def _get_filters_from_request(request):
    """Extract filters from request"""
    return {
        'status': request.GET.get('status', ''),
        'client': request.GET.get('client', ''),
        'type': request.GET.get('type', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'sort_by': request.GET.get('sort', 'deadline'),
    }


def _build_export_query(filters):
    """Build query for export"""
    now = timezone.now()

    # Base query
    if filters['date_from'] and filters['date_to']:
        query = MonthlyObligation.objects.filter(
            deadline__gte=filters['date_from'],
            deadline__lte=filters['date_to']
        )
    else:
        next_month = now.date() + timedelta(days=30)
        query = MonthlyObligation.objects.filter(
            deadline__range=[now.date(), next_month]
        )

    # Apply filters
    if filters['status']:
        query = query.filter(status=filters['status'])

    if filters['client']:
        query = query.filter(client_id=filters['client'])

    if filters['type']:
        query = query.filter(obligation_type_id=filters['type'])

    # Sort
    sort_mapping = {
        'deadline': 'deadline',
        'client': 'client__eponimia',
        'type': 'obligation_type__name',
        'status': 'status'
    }

    return query.select_related('client', 'obligation_type').order_by(
        sort_mapping.get(filters['sort_by'], 'deadline')
    )


def _apply_excel_styling(ws):
    """Apply styling to Excel worksheet"""
    ws['A1'] = 'ΑΝΑΦΟΡΑ ΥΠΟΧΡΕΩΣΕΩΝ'
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A2'] = f'Ημερομηνία: {timezone.now().strftime("%d/%m/%Y %H:%M")}'


def _write_excel_headers(ws):
    """Write headers to Excel worksheet"""
    headers = ['#', 'Προθεσμία', 'Πελάτης', 'ΑΦΜ', 'Υποχρέωση', 'Κατάσταση', 'Ώρες', 'Σημειώσεις']
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _write_excel_data(ws, query):
    """Write data rows to Excel worksheet"""
    row = 6
    for idx, obl in enumerate(query, start=1):
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = obl.deadline.strftime('%d/%m/%Y')
        ws.cell(row=row, column=3).value = obl.client.eponimia
        ws.cell(row=row, column=4).value = obl.client.afm
        ws.cell(row=row, column=5).value = obl.obligation_type.name
        ws.cell(row=row, column=6).value = obl.get_status_display()
        ws.cell(row=row, column=7).value = float(obl.time_spent) if obl.time_spent else ''
        ws.cell(row=row, column=8).value = obl.notes[:100] if obl.notes else ''

        # Color coding for status
        status_cell = ws.cell(row=row, column=6)
        if obl.status == 'completed':
            status_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        elif obl.status == 'overdue':
            status_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")

        row += 1


def _auto_adjust_excel_columns(ws):
    """Auto-adjust column widths in Excel"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# ============================================
# ANALYTICS HELPERS
# ============================================

def _calculate_monthly_completion_stats(start_date):
    """Calculate monthly completion statistics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).annotate(
        month_label=TruncMonth('deadline')
    ).values('month_label').annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        pending=Count('id', filter=Q(status='pending')),
        overdue=Count('id', filter=Q(status='overdue')),
    ).order_by('month_label')


def _calculate_client_performance(start_date):
    """Calculate client performance metrics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).values(
        'client__id',
        'client__eponimia'
    ).annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        completion_rate=Count('id', filter=Q(status='completed')) * 100.0 / Count('id')
    ).order_by('-total')[:10]


def _calculate_time_stats(start_date):
    """Calculate time tracking statistics"""
    return MonthlyObligation.objects.filter(
        status='completed',
        completed_date__gte=start_date,
        time_spent__isnull=False
    ).aggregate(
        total_hours=Sum('time_spent'),
        avg_hours=Avg('time_spent'),
        total_tasks=Count('id')
    )


def _calculate_revenue(start_date):
    """Calculate revenue data"""
    obligations = MonthlyObligation.objects.filter(
        status='completed',
        completed_date__gte=start_date,
        time_spent__isnull=False
    ).select_related('client')

    total_revenue = 0
    for obl in obligations:
        # Assuming hourly rate from client or default
        hourly_rate = 50  # Default rate, adjust as needed
        if hasattr(obl, 'hourly_rate'):
            hourly_rate = obl.hourly_rate
        total_revenue += Decimal(obl.time_spent or 0) * hourly_rate

    return {'total': total_revenue}


def _calculate_type_stats(start_date):
    """Calculate obligation type statistics"""
    return MonthlyObligation.objects.filter(
        deadline__gte=start_date
    ).values(
        'obligation_type__name'
    ).annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        avg_time=Avg('time_spent', filter=Q(status='completed'))
    ).order_by('-total')[:10]


def _calculate_current_month_stats(now):
    """Calculate current month statistics"""
    current_month = now.month
    current_year = now.year

    return {
        'total': MonthlyObligation.objects.filter(
            year=current_year, month=current_month
        ).count(),
        'completed': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='completed'
        ).count(),
        'pending': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='pending'
        ).count(),
        'overdue': MonthlyObligation.objects.filter(
            year=current_year, month=current_month, status='overdue'
        ).count(),
    }


def _format_chart_data(monthly_stats):
    """Format data for chart display"""
    chart_labels = [stat['month_label'].strftime('%b %Y') for stat in monthly_stats]

    return {
        'chart_labels_json': json.dumps(chart_labels),
        'chart_completed_json': json.dumps([stat['completed'] for stat in monthly_stats]),
        'chart_pending_json': json.dumps([stat['pending'] for stat in monthly_stats]),
        'chart_overdue_json': json.dumps([stat['overdue'] for stat in monthly_stats]),
    }
