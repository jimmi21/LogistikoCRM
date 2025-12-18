# -*- coding: utf-8 -*-
"""
accounting/api_obligations.py
Author: Claude
Description: REST API ViewSet for MonthlyObligation management
"""

from rest_framework import viewsets, status, filters, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import (
    DjangoFilterBackend, FilterSet, CharFilter,
    NumberFilter, DateFilter
)
from django.utils import timezone
from django.db.models import Q

from .models import MonthlyObligation, ClientProfile, ObligationType, ClientDocument


class ObligationPagination(PageNumberPagination):
    """Pagination for obligation list"""
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 100


class ObligationFilter(FilterSet):
    """Filter for MonthlyObligation"""
    client = NumberFilter(field_name='client_id')
    status = CharFilter(field_name='status')
    month = NumberFilter(field_name='month')
    year = NumberFilter(field_name='year')
    type = CharFilter(field_name='obligation_type__code')
    deadline_from = DateFilter(field_name='deadline', lookup_expr='gte')
    deadline_to = DateFilter(field_name='deadline', lookup_expr='lte')
    search = CharFilter(method='filter_search')

    class Meta:
        model = MonthlyObligation
        fields = ['client', 'status', 'month', 'year', 'type']

    def filter_search(self, queryset, name, value):
        """Search by client name, AFM, or obligation type"""
        if value:
            return queryset.filter(
                Q(client__eponimia__icontains=value) |
                Q(client__afm__icontains=value) |
                Q(obligation_type__name__icontains=value) |
                Q(obligation_type__code__icontains=value)
            )
        return queryset


# ============================================
# OBLIGATION SERIALIZERS
# ============================================

class ObligationTypeSerializer(serializers.ModelSerializer):
    """Serializer for ObligationType"""

    class Meta:
        model = ObligationType
        fields = ['id', 'code', 'name', 'frequency', 'deadline_type', 'deadline_day']


class ObligationListSerializer(serializers.ModelSerializer):
    """Serializer for obligation list view"""
    client_name = serializers.CharField(source='client.eponimia', read_only=True)
    client_afm = serializers.CharField(source='client.afm', read_only=True)
    type_name = serializers.CharField(source='obligation_type.name', read_only=True)
    type_code = serializers.CharField(source='obligation_type.code', read_only=True)
    period = serializers.SerializerMethodField()
    days_until_deadline = serializers.IntegerField(read_only=True)
    is_overdue = serializers.BooleanField(read_only=True)
    assigned_to_name = serializers.SerializerMethodField()
    documents_count = serializers.SerializerMethodField()

    class Meta:
        model = MonthlyObligation
        fields = [
            'id', 'client', 'client_name', 'client_afm',
            'obligation_type', 'type_name', 'type_code',
            'year', 'month', 'period', 'deadline', 'status',
            'assigned_to', 'assigned_to_name',
            'days_until_deadline', 'is_overdue', 'documents_count'
        ]

    def get_period(self, obj):
        return f"{obj.month:02d}/{obj.year}"

    def get_assigned_to_name(self, obj):
        if obj.assigned_to:
            if obj.assigned_to.first_name or obj.assigned_to.last_name:
                return f"{obj.assigned_to.first_name} {obj.assigned_to.last_name}".strip()
            return obj.assigned_to.username
        return None

    def get_documents_count(self, obj):
        """Return count of documents attached to this obligation"""
        # Use annotated value if available (avoids N+1 queries)
        if hasattr(obj, '_documents_count'):
            return obj._documents_count
        # Fallback for detail view or non-annotated querysets
        return ClientDocument.objects.filter(
            client=obj.client,
            obligation=obj
        ).count()


class ObligationDetailSerializer(serializers.ModelSerializer):
    """Detailed serializer for single obligation view"""
    client_name = serializers.CharField(source='client.eponimia', read_only=True)
    client_afm = serializers.CharField(source='client.afm', read_only=True)
    client_email = serializers.CharField(source='client.email', read_only=True)
    type_name = serializers.CharField(source='obligation_type.name', read_only=True)
    type_code = serializers.CharField(source='obligation_type.code', read_only=True)
    period = serializers.SerializerMethodField()
    days_until_deadline = serializers.IntegerField(read_only=True)
    is_overdue = serializers.BooleanField(read_only=True)
    deadline_status = serializers.CharField(read_only=True)
    cost = serializers.DecimalField(
        max_digits=10, decimal_places=2, read_only=True
    )
    completed_by_username = serializers.CharField(
        source='completed_by.username', read_only=True
    )
    assigned_to_name = serializers.SerializerMethodField()
    documents = serializers.SerializerMethodField()
    attachment_url = serializers.SerializerMethodField()

    class Meta:
        model = MonthlyObligation
        fields = [
            'id', 'client', 'client_name', 'client_afm', 'client_email',
            'obligation_type', 'type_name', 'type_code',
            'year', 'month', 'period', 'deadline', 'status',
            'completed_date', 'completed_by', 'completed_by_username',
            'assigned_to', 'assigned_to_name',
            'notes', 'time_spent', 'hourly_rate', 'cost',
            'days_until_deadline', 'is_overdue', 'deadline_status',
            'attachment', 'attachment_url', 'attachments', 'documents',
            'created_at', 'updated_at'
        ]
        read_only_fields = ['created_at', 'updated_at', 'completed_by']

    def get_assigned_to_name(self, obj):
        if obj.assigned_to:
            if obj.assigned_to.first_name or obj.assigned_to.last_name:
                return f"{obj.assigned_to.first_name} {obj.assigned_to.last_name}".strip()
            return obj.assigned_to.username
        return None

    def get_period(self, obj):
        return f"{obj.month:02d}/{obj.year}"

    def get_attachment_url(self, obj):
        if obj.attachment:
            request = self.context.get('request')
            if request:
                return request.build_absolute_uri(obj.attachment.url)
            return obj.attachment.url
        return None

    def get_documents(self, obj):
        """Get related documents for this obligation"""
        docs = ClientDocument.objects.filter(
            client=obj.client,
            obligation=obj
        )
        from .serializers import ClientDocumentSerializer
        return ClientDocumentSerializer(
            docs, many=True, context=self.context
        ).data


class ObligationCreateUpdateSerializer(serializers.ModelSerializer):
    """Serializer for creating/updating obligations"""

    class Meta:
        model = MonthlyObligation
        fields = [
            'client', 'obligation_type', 'year', 'month',
            'deadline', 'status', 'notes', 'time_spent', 'hourly_rate',
            'assigned_to'
        ]

    def validate(self, data):
        """Validate unique together constraint"""
        client = data.get('client')
        obligation_type = data.get('obligation_type')
        year = data.get('year')
        month = data.get('month')

        # Check for existing obligation (only on create)
        if not self.instance:
            exists = MonthlyObligation.objects.filter(
                client=client,
                obligation_type=obligation_type,
                year=year,
                month=month
            ).exists()
            if exists:
                raise serializers.ValidationError({
                    'non_field_errors': [
                        f"Υπάρχει ήδη υποχρέωση για αυτόν τον πελάτη, "
                        f"τύπο και περίοδο ({month:02d}/{year})."
                    ]
                })

        # Validate month range
        if month and (month < 1 or month > 12):
            raise serializers.ValidationError({
                'month': "Ο μήνας πρέπει να είναι από 1 έως 12."
            })

        return data


# ============================================
# OBLIGATION VIEWSET
# ============================================

class ObligationViewSet(viewsets.ModelViewSet):
    """
    REST API ViewSet for MonthlyObligation

    Endpoints:
    - GET /api/obligations/ - List all (with filters)
    - GET /api/obligations/{id}/ - Get single
    - POST /api/obligations/ - Create
    - PUT /api/obligations/{id}/ - Update
    - PATCH /api/obligations/{id}/complete/ - Mark as complete
    - DELETE /api/obligations/{id}/ - Delete
    """
    queryset = MonthlyObligation.objects.all()
    permission_classes = [IsAuthenticated]
    pagination_class = ObligationPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = ObligationFilter
    ordering_fields = ['deadline', 'client__eponimia', 'status', 'created_at']
    ordering = ['deadline']

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return ObligationListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return ObligationCreateUpdateSerializer
        return ObligationDetailSerializer

    def get_queryset(self):
        """Optimize queryset with select_related and annotations"""
        from django.db.models import Count

        queryset = super().get_queryset().select_related(
            'client', 'obligation_type', 'completed_by', 'assigned_to'
        )

        # Add document count annotation to avoid N+1 queries
        if self.action == 'list':
            queryset = queryset.annotate(
                _documents_count=Count('documents', distinct=True)
            )

        return queryset

    @action(detail=True, methods=['patch'])
    def complete(self, request, pk=None):
        """
        PATCH /api/obligations/{id}/complete/
        Mark an obligation as completed
        """
        obligation = self.get_object()

        if obligation.status == 'completed':
            return Response(
                {'error': 'Η υποχρέωση είναι ήδη ολοκληρωμένη.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update obligation
        obligation.status = 'completed'
        obligation.completed_date = timezone.now().date()
        obligation.completed_by = request.user

        # Optional: time spent and notes from request
        if 'time_spent' in request.data:
            obligation.time_spent = request.data['time_spent']
        if 'notes' in request.data:
            obligation.notes = request.data.get('notes', '')

        obligation.save()

        serializer = ObligationDetailSerializer(
            obligation, context={'request': request}
        )
        return Response({
            'message': 'Η υποχρέωση ολοκληρώθηκε επιτυχώς.',
            'obligation': serializer.data
        })

    @action(detail=False, methods=['get'])
    def types(self, request):
        """
        GET /api/obligations/types/
        Get all active obligation types
        """
        types = ObligationType.objects.filter(is_active=True).order_by('priority', 'name')
        serializer = ObligationTypeSerializer(types, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """
        GET /api/obligations/overdue/
        Get all overdue obligations
        """
        today = timezone.now().date()
        overdue = self.get_queryset().filter(
            status__in=['pending', 'overdue'],
            deadline__lt=today
        ).order_by('deadline')

        # Update status to overdue if needed
        overdue.update(status='overdue')

        page = self.paginate_queryset(overdue)
        if page is not None:
            serializer = ObligationListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = ObligationListSerializer(overdue, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        """
        GET /api/obligations/upcoming/
        Get obligations with deadline in next 7 days
        """
        today = timezone.now().date()
        from datetime import timedelta
        next_week = today + timedelta(days=7)

        upcoming = self.get_queryset().filter(
            status='pending',
            deadline__gte=today,
            deadline__lte=next_week
        ).order_by('deadline')

        page = self.paginate_queryset(upcoming)
        if page is not None:
            serializer = ObligationListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = ObligationListSerializer(upcoming, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def bulk_complete(self, request):
        """
        POST /api/obligations/bulk_complete/
        Mark multiple obligations as completed

        Body: {"obligation_ids": [1, 2, 3]}
        """
        obligation_ids = request.data.get('obligation_ids', [])

        if not obligation_ids:
            return Response(
                {'error': 'Δεν δόθηκαν IDs υποχρεώσεων.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        obligations = MonthlyObligation.objects.filter(
            id__in=obligation_ids,
            status__in=['pending', 'overdue']
        )

        updated_count = obligations.update(
            status='completed',
            completed_date=timezone.now().date(),
            completed_by=request.user
        )

        return Response({
            'message': f'{updated_count} υποχρεώσεις ολοκληρώθηκαν.',
            'updated_count': updated_count
        })

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """
        POST /api/obligations/bulk_create/
        Create obligations for multiple clients at once

        Body: {
            "client_ids": [1, 2, 3],
            "obligation_type": 1,
            "month": 12,
            "year": 2025
        }
        """
        client_ids = request.data.get('client_ids', [])
        obligation_type_id = request.data.get('obligation_type')
        month = request.data.get('month')
        year = request.data.get('year')

        # Validation
        if not client_ids:
            return Response(
                {'error': 'Δεν δόθηκαν IDs πελατών.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not obligation_type_id:
            return Response(
                {'error': 'Δεν δόθηκε τύπος υποχρέωσης.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not month or not year:
            return Response(
                {'error': 'Δεν δόθηκε μήνας ή έτος.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            obligation_type = ObligationType.objects.get(id=obligation_type_id)
        except ObligationType.DoesNotExist:
            return Response(
                {'error': 'Ο τύπος υποχρέωσης δεν βρέθηκε.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get clients
        clients = ClientProfile.objects.filter(id__in=client_ids, is_active=True)
        if not clients.exists():
            return Response(
                {'error': 'Δεν βρέθηκαν ενεργοί πελάτες.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Calculate deadline
        deadline = obligation_type.get_deadline_for_month(year, month)
        if not deadline:
            # Default to last day of month
            from calendar import monthrange
            last_day = monthrange(year, month)[1]
            deadline = timezone.datetime(year, month, last_day).date()

        created_count = 0
        skipped_count = 0
        created_obligations = []

        for client in clients:
            # Check if obligation already exists
            existing = MonthlyObligation.objects.filter(
                client=client,
                obligation_type=obligation_type,
                year=year,
                month=month
            ).exists()

            if existing:
                skipped_count += 1
                continue

            # Create new obligation
            obligation = MonthlyObligation.objects.create(
                client=client,
                obligation_type=obligation_type,
                year=year,
                month=month,
                deadline=deadline,
                status='pending'
            )
            created_obligations.append(obligation)
            created_count += 1

        serializer = ObligationListSerializer(created_obligations, many=True)
        return Response({
            'message': f'Δημιουργήθηκαν {created_count} υποχρεώσεις. Παραλείφθηκαν {skipped_count} (υπήρχαν ήδη).',
            'created_count': created_count,
            'skipped_count': skipped_count,
            'obligations': serializer.data
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """
        POST /api/obligations/bulk_update/
        Update status for multiple obligations

        Body: {"obligation_ids": [1, 2, 3], "status": "completed"}
        """
        obligation_ids = request.data.get('obligation_ids', [])
        new_status = request.data.get('status')

        if not obligation_ids:
            return Response(
                {'error': 'Δεν δόθηκαν IDs υποχρεώσεων.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        valid_statuses = ['pending', 'completed', 'overdue']
        if new_status not in valid_statuses:
            return Response(
                {'error': f'Μη έγκυρη κατάσταση. Επιλέξτε από: {", ".join(valid_statuses)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        obligations = MonthlyObligation.objects.filter(id__in=obligation_ids)

        update_data = {'status': new_status}
        if new_status == 'completed':
            update_data['completed_date'] = timezone.now().date()
            update_data['completed_by'] = request.user

        updated_count = obligations.update(**update_data)

        return Response({
            'message': f'{updated_count} υποχρεώσεις ενημερώθηκαν σε "{new_status}".',
            'updated_count': updated_count
        })

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """
        POST /api/obligations/bulk_delete/
        Delete multiple obligations

        Body: {"obligation_ids": [1, 2, 3]}
        """
        obligation_ids = request.data.get('obligation_ids', [])

        if not obligation_ids:
            return Response(
                {'error': 'Δεν δόθηκαν IDs υποχρεώσεων.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        deleted_count, _ = MonthlyObligation.objects.filter(
            id__in=obligation_ids
        ).delete()

        return Response({
            'message': f'{deleted_count} υποχρεώσεις διαγράφηκαν.',
            'deleted_count': deleted_count
        })

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        GET /api/obligations/export/?format=excel
        Export obligations to Excel with filters applied
        """
        import io
        from django.http import HttpResponse

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return Response(
                {'error': 'Η βιβλιοθήκη openpyxl δεν είναι εγκατεστημένη.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Get filtered queryset using the same filter as list
        queryset = self.filter_queryset(self.get_queryset())

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Υποχρεώσεις"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "ID", "Πελάτης", "ΑΦΜ", "Τύπος Υποχρέωσης", "Κωδικός",
            "Μήνας", "Έτος", "Προθεσμία", "Κατάσταση", "Ημ/νία Ολοκλήρωσης",
            "Σημειώσεις"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Status translations
        status_map = {
            'pending': 'Εκκρεμεί',
            'completed': 'Ολοκληρώθηκε',
            'overdue': 'Καθυστερεί',
        }

        # Data rows
        for row_num, obl in enumerate(queryset, 2):
            data = [
                obl.id,
                obl.client.eponimia,
                obl.client.afm,
                obl.obligation_type.name,
                obl.obligation_type.code,
                obl.month,
                obl.year,
                obl.deadline.strftime('%d/%m/%Y') if obl.deadline else '',
                status_map.get(obl.status, obl.status),
                obl.completed_date.strftime('%d/%m/%Y') if obl.completed_date else '',
                obl.notes or ''
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        column_widths = [8, 30, 12, 25, 10, 8, 8, 12, 15, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename with current date
        filename = f"υποχρεώσεις_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['get'])
    def calendar(self, request):
        """
        GET /api/v1/obligations/calendar/
        Get obligations organized by day for calendar view

        Query params:
        - month (1-12, required)
        - year (YYYY, required)
        - client_id (optional) - filter by specific client
        - type_id (optional) - filter by obligation type ID
        - status (optional) - filter by status

        Returns:
        {
            "month": 12,
            "year": 2025,
            "days": {
                "15": {
                    "total": 5,
                    "pending": 3,
                    "completed": 1,
                    "overdue": 1,
                    "obligations": [...]
                }
            },
            "summary": {
                "total": 45,
                "pending": 30,
                "completed": 10,
                "overdue": 5
            }
        }
        """
        from calendar import monthrange
        from collections import defaultdict

        today = timezone.now().date()

        # Get and validate parameters
        try:
            month = int(request.query_params.get('month', today.month))
            year = int(request.query_params.get('year', today.year))
        except (ValueError, TypeError):
            return Response(
                {'error': 'Μη έγκυρος μήνας ή έτος.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if month < 1 or month > 12:
            return Response(
                {'error': 'Ο μήνας πρέπει να είναι από 1 έως 12.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Optional filters
        client_id = request.query_params.get('client_id')
        type_id = request.query_params.get('type_id')
        filter_status = request.query_params.get('status')

        # Get date range for the month
        first_day = timezone.datetime(year, month, 1).date()
        last_day_num = monthrange(year, month)[1]
        last_day = timezone.datetime(year, month, last_day_num).date()

        # Build queryset
        queryset = self.get_queryset().filter(
            deadline__gte=first_day,
            deadline__lte=last_day
        )

        # Apply optional filters
        if client_id:
            try:
                queryset = queryset.filter(client_id=int(client_id))
            except (ValueError, TypeError):
                pass

        if type_id:
            try:
                queryset = queryset.filter(obligation_type_id=int(type_id))
            except (ValueError, TypeError):
                pass

        if filter_status:
            queryset = queryset.filter(status=filter_status)

        # Update overdue status for pending obligations past deadline
        queryset.filter(
            status='pending',
            deadline__lt=today
        ).update(status='overdue')

        # Refresh queryset after update with select_related to avoid N+1 queries
        queryset = queryset.select_related(
            'client', 'obligation_type'
        ).order_by('deadline')

        # Valid statuses for counting
        VALID_STATUSES = {'pending', 'completed', 'overdue', 'in_progress', 'cancelled'}

        # Group obligations by day
        days = defaultdict(lambda: {
            'total': 0,
            'pending': 0,
            'completed': 0,
            'overdue': 0,
            'in_progress': 0,
            'cancelled': 0,
            'obligations': []
        })

        for obl in queryset:
            day_str = str(obl.deadline.day)
            days[day_str]['total'] += 1
            # Only increment known status counters
            if obl.status in VALID_STATUSES:
                days[day_str][obl.status] += 1
            days[day_str]['obligations'].append({
                'id': obl.id,
                'client_name': obl.client.eponimia,
                'client_id': obl.client.id,
                'type_name': obl.obligation_type.name,
                'type_code': obl.obligation_type.code,
                'status': obl.status,
                'deadline': obl.deadline.isoformat(),
                'notes': obl.notes or '',
            })

        # Calculate summary
        summary = {
            'total': 0,
            'pending': 0,
            'completed': 0,
            'overdue': 0,
            'in_progress': 0,
            'cancelled': 0,
        }
        for day_data in days.values():
            summary['total'] += day_data['total']
            summary['pending'] += day_data['pending']
            summary['completed'] += day_data['completed']
            summary['overdue'] += day_data['overdue']
            summary['in_progress'] += day_data['in_progress']
            summary['cancelled'] += day_data['cancelled']

        return Response({
            'month': month,
            'year': year,
            'days': dict(days),
            'summary': summary,
        })


# ============================================
# OBLIGATION TYPE VIEWSET
# ============================================

class ObligationTypeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    REST API ViewSet for ObligationType (read-only)

    Endpoints:
    - GET /api/v1/obligation-types/ - List all active types
    - GET /api/v1/obligation-types/{id}/ - Get single type
    """
    queryset = ObligationType.objects.filter(is_active=True).order_by('priority', 'name')
    serializer_class = ObligationTypeSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None  # Return all types without pagination
