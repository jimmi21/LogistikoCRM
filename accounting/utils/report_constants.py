# -*- coding: utf-8 -*-
"""
accounting/utils/report_constants.py
Author: Claude
Description: Centralized constants and utilities for report generation.
             Consolidates duplicate code from views, API endpoints, and exports.
"""

from datetime import timedelta
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================
# GREEK MONTH NAMES
# ============================================

GREEK_MONTHS_FULL = {
    1: 'Ιανουάριος',
    2: 'Φεβρουάριος',
    3: 'Μάρτιος',
    4: 'Απρίλιος',
    5: 'Μάιος',
    6: 'Ιούνιος',
    7: 'Ιούλιος',
    8: 'Αύγουστος',
    9: 'Σεπτέμβριος',
    10: 'Οκτώβριος',
    11: 'Νοέμβριος',
    12: 'Δεκέμβριος',
}

GREEK_MONTHS_SHORT = [
    'Ιαν', 'Φεβ', 'Μαρ', 'Απρ', 'Μαι', 'Ιουν',
    'Ιουλ', 'Αυγ', 'Σεπ', 'Οκτ', 'Νοε', 'Δεκ'
]

GREEK_MONTHS_GENITIVE = {
    1: 'Ιανουαρίου',
    2: 'Φεβρουαρίου',
    3: 'Μαρτίου',
    4: 'Απριλίου',
    5: 'Μαΐου',
    6: 'Ιουνίου',
    7: 'Ιουλίου',
    8: 'Αυγούστου',
    9: 'Σεπτεμβρίου',
    10: 'Οκτωβρίου',
    11: 'Νοεμβρίου',
    12: 'Δεκεμβρίου',
}


def get_greek_month_name(month_num, form='full'):
    """
    Get Greek month name by number.

    Args:
        month_num: Month number (1-12)
        form: 'full', 'short', or 'genitive'

    Returns:
        Greek month name string
    """
    if form == 'short':
        return GREEK_MONTHS_SHORT[month_num - 1] if 1 <= month_num <= 12 else str(month_num)
    elif form == 'genitive':
        return GREEK_MONTHS_GENITIVE.get(month_num, str(month_num))
    else:
        return GREEK_MONTHS_FULL.get(month_num, str(month_num))


# ============================================
# DATE RANGE UTILITIES
# ============================================

REPORT_PERIODS = [
    ('today', 'Σήμερα'),
    ('week', 'Εβδομάδα'),
    ('month', 'Μήνας'),
    ('quarter', 'Τρίμηνο'),
    ('year', 'Έτος'),
    ('all', 'Όλα'),
]


def get_date_range(period: str):
    """
    Returns start_date and end_date based on period filter.

    Args:
        period: 'today', 'week', 'month', 'quarter', 'year', 'all'

    Returns:
        Tuple of (start_date, end_date) or (None, None) for 'all'
    """
    today = timezone.now().date()

    if period == 'today':
        return today, today
    elif period == 'week':
        start_of_week = today - timedelta(days=today.weekday())
        return start_of_week, today
    elif period == 'month':
        start_of_month = today.replace(day=1)
        return start_of_month, today
    elif period == 'quarter':
        current_quarter = (today.month - 1) // 3
        start_month = current_quarter * 3 + 1
        start_of_quarter = today.replace(month=start_month, day=1)
        return start_of_quarter, today
    elif period == 'year':
        start_of_year = today.replace(month=1, day=1)
        return start_of_year, today
    else:  # 'all' or any other value
        return None, None


def get_previous_period_range(period: str, current_start, current_end):
    """
    Get the previous period's date range for comparison.

    Args:
        period: Current period type
        current_start: Current period start date
        current_end: Current period end date

    Returns:
        Tuple of (prev_start, prev_end)
    """
    today = timezone.now().date()

    if period == 'today':
        prev_start = prev_end = today - timedelta(days=1)
    elif period == 'week':
        prev_end = current_start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=6)
    elif period == 'month':
        prev_end = current_start - timedelta(days=1)
        prev_start = prev_end.replace(day=1)
    elif period == 'quarter':
        prev_end = current_start - timedelta(days=1)
        prev_quarter = (prev_end.month - 1) // 3
        prev_start_month = prev_quarter * 3 + 1
        prev_start = prev_end.replace(month=prev_start_month, day=1)
    elif period == 'year':
        prev_end = current_start - timedelta(days=1)
        prev_start = prev_end.replace(month=1, day=1)
    else:
        return None, None

    return prev_start, prev_end


# ============================================
# EXCEL STYLING CONSTANTS
# ============================================

# Primary brand color (purple)
BRAND_COLOR = "667EEA"
BRAND_COLOR_LIGHT = "E8EAF6"

# Status colors
STATUS_COLORS = {
    'completed': 'd4edda',   # Light green
    'pending': 'fff3cd',     # Light yellow
    'overdue': 'f8d7da',     # Light red
    'in_progress': 'cce5ff', # Light blue
}

# Status text colors
STATUS_TEXT_COLORS = {
    'completed': '155724',
    'pending': '856404',
    'overdue': '721c24',
    'in_progress': '004085',
}


def get_excel_header_style():
    """
    Get standard Excel header styling.

    Returns:
        Dict with font, fill, alignment, border
    """
    return {
        'font': Font(bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid"),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': get_excel_border(),
    }


def get_excel_border():
    """Get standard thin border for Excel cells."""
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def get_excel_title_style():
    """Get styling for Excel title row."""
    return {
        'font': Font(bold=True, size=14, color=BRAND_COLOR),
        'alignment': Alignment(horizontal='center'),
    }


def get_status_fill(status):
    """
    Get PatternFill for a given status.

    Args:
        status: 'completed', 'pending', 'overdue', 'in_progress'

    Returns:
        PatternFill object
    """
    color = STATUS_COLORS.get(status, 'ffffff')
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def get_alternating_row_fill(row_num):
    """
    Get fill for alternating row coloring.

    Args:
        row_num: Row number (0-indexed or 1-indexed)

    Returns:
        PatternFill for even rows, None for odd rows
    """
    if row_num % 2 == 0:
        return PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    return None


def apply_header_style(cell):
    """Apply header styling to a cell."""
    style = get_excel_header_style()
    cell.font = style['font']
    cell.fill = style['fill']
    cell.alignment = style['alignment']
    cell.border = style['border']


def auto_adjust_column_width(worksheet, min_width=8, max_width=50):
    """
    Auto-adjust column widths based on content.

    Args:
        worksheet: openpyxl worksheet
        min_width: Minimum column width
        max_width: Maximum column width
    """
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, worksheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)

            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue

            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        adjusted_width = max(min_width, min(max_length + 2, max_width))
        worksheet.column_dimensions[column_letter].width = adjusted_width


# ============================================
# REPORT EXPORT FORMATS
# ============================================

EXPORT_FORMATS = {
    'csv': {
        'content_type': 'text/csv; charset=utf-8-sig',
        'extension': '.csv',
    },
    'xlsx': {
        'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'extension': '.xlsx',
    },
    'pdf': {
        'content_type': 'application/pdf',
        'extension': '.pdf',
    },
}


def get_export_filename(base_name, format_type='xlsx'):
    """
    Generate a timestamped export filename.

    Args:
        base_name: Base name for the file (e.g., 'Clients', 'Obligations')
        format_type: 'csv', 'xlsx', or 'pdf'

    Returns:
        Filename string with timestamp
    """
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    extension = EXPORT_FORMATS.get(format_type, EXPORT_FORMATS['xlsx'])['extension']
    return f"{base_name}_{timestamp}{extension}"


def get_content_type(format_type='xlsx'):
    """Get HTTP content type for export format."""
    return EXPORT_FORMATS.get(format_type, EXPORT_FORMATS['xlsx'])['content_type']


# ============================================
# OBLIGATION STATUS CONSTANTS
# ============================================

OBLIGATION_STATUS_CHOICES = [
    ('pending', 'Εκκρεμεί'),
    ('in_progress', 'Σε εξέλιξη'),
    ('completed', 'Ολοκληρώθηκε'),
    ('overdue', 'Εκπρόθεσμη'),
    ('cancelled', 'Ακυρώθηκε'),
]

OBLIGATION_STATUS_ICONS = {
    'pending': '🟡',
    'in_progress': '🔵',
    'completed': '✅',
    'overdue': '🔴',
    'cancelled': '⚫',
}


def get_status_display(status):
    """Get Greek display text for status."""
    return dict(OBLIGATION_STATUS_CHOICES).get(status, status)


def get_status_icon(status):
    """Get emoji icon for status."""
    return OBLIGATION_STATUS_ICONS.get(status, '⚪')


# ============================================
# OBLIGATION TYPE COLORS (for charts)
# ============================================

OBLIGATION_TYPE_COLORS = {
    'ΦΠΑ': '#3B82F6',      # Blue
    'ΑΠΔ': '#10B981',      # Green
    'ΕΝΦΙΑ': '#F59E0B',    # Yellow
    'Ε1': '#8B5CF6',       # Purple
    'Ε3': '#EC4899',       # Pink
    'ΜΥΦ': '#F97316',      # Orange
    'default': '#6B7280',  # Gray
}


def get_type_color(type_code):
    """Get chart color for obligation type."""
    return OBLIGATION_TYPE_COLORS.get(type_code, OBLIGATION_TYPE_COLORS['default'])
